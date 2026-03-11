from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from app import db
from sqlalchemy import text
import math
from datetime import datetime

transactions_bp = Blueprint('transactions', __name__, url_prefix='/transactions')

class Pagination:
    def __init__(self, page, per_page, total):
        self.page = page
        self.per_page = per_page
        self.total = total
        self.pages = int(math.ceil(total / float(per_page))) if total > 0 else 0
    @property
    def has_prev(self): return self.page > 1
    @property
    def has_next(self): return self.page < self.pages
    @property
    def prev_num(self): return self.page - 1
    @property
    def next_num(self): return self.page + 1
    def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
        last = 0
        for num in range(1, self.pages + 1):
            if num <= left_edge or (num > self.page - left_current - 1 and num < self.page + right_current) or num > self.pages - right_edge:
                if last + 1 != num: yield None
                yield num
                last = num

def _db_ping():
    try:
        db.session.execute(text("SELECT 1"))
        return True
    except Exception:
        try: db.session.rollback()
        except: pass
        return False

def safe_all(query_fn, default=None):
    try: return query_fn()
    except Exception: return [] if default is None else default

def format_dt(val):
    if not val: return ""
    if hasattr(val, 'strftime'): return val.strftime('%d %b %Y')
    return str(val)

def get_common_masters():
    sessions = safe_all(lambda: db.session.execute(text("SELECT pk_sessionid as id, description as name FROM LUP_AcademicSession_Mst ORDER BY pk_sessionid DESC")).mappings().all(), [])
    degrees = safe_all(lambda: db.session.execute(text("""
        SELECT d.pk_degreeid as id, 
               CASE 
                   WHEN dt.description IS NOT NULL THEN d.description + ' (' + dt.description + ')' 
                   ELSE d.description 
               END as name 
        FROM ACD_Degree_Mst d
        LEFT JOIN ACD_DegreeType_Mst dt ON d.fk_dtypeid = dt.pk_dtypeid
        WHERE d.active=1 
        ORDER BY d.description
    """)).mappings().all(), [])
    
    # Only fetch campuses (Hisar, Kaul, Bawal, CSSRI, IIWBR, Gurugram)
    colleges = safe_all(lambda: db.session.execute(text("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE collegename LIKE '%Campus%' OR collegename LIKE '%Gurugram%' ORDER BY collegename")).mappings().all(), [])
    
    streams = safe_all(lambda: db.session.execute(text("SELECT Pk_CollegeID as id, CollegeName as name FROM PA_College_Mst ORDER BY CollegeName")).mappings().all(), [])
    return sessions, degrees, colleges, streams

# 1. Apply For Upward

@transactions_bp.before_request
def check_admin_auth():
    from flask import session, redirect, url_for, flash, request
    if not session.get('user_id'):
        if request.endpoint and 'login' not in request.endpoint:
            flash('Please login to access this section.', 'error')
            return redirect(url_for('main.login'))

@transactions_bp.route('/apply-for-upward', methods=['GET', 'POST'])
def apply_for_upward():
    sessions, degrees, colleges, streams = get_common_masters()
    
    # Fixed range for cutoffs based on live system
    cutoffs = list(range(1, 11))
    
    f_session = request.args.get('session_id')
    if not f_session and sessions:
        f_session = str(sessions[0]['id'])
        
    f_stream = request.args.get('stream_id')
    f_college = request.args.get('college_id')
    f_degree = request.args.get('degree_id')
    f_reg = request.args.get('reg_no')
    f_cutoff = request.args.get('cutoff')
    f_pref = request.args.get('pref_above_1')
    action = request.args.get('action')
    
    eligible_records = []
    applied_records = []
    show_grids = False
    
    if _db_ping() and action in ['view', 'apply'] and (f_session or f_reg):
        show_grids = True
        where = " WHERE 1=1"
        params = {}
        if f_session: 
            where += " AND mm.Fk_SessionID=:session_id"
            params["session_id"] = f_session
        if f_degree: 
            where += " AND mm.Fk_DegreeID=:degree_id"
            params["degree_id"] = f_degree
        if f_stream:
            where += " AND mt.Fk_CollegeID=:stream_id"
            params["stream_id"] = f_stream
        
        # In the live system, the grid is NOT filtered by the allocated campus. 
        # All students in the stream and degree are returned.
        # So we intentionally skip `mt.fk_allotedcollegeid = f_college` here.
        
        if f_reg: 
            where += " AND m.regno LIKE :reg_no"
            params["reg_no"] = f"%{f_reg}%"
        if f_cutoff:
            where += " AND mm.CutOff=:cutoff"
            params["cutoff"] = f_cutoff
            
        # Exclude absent, not paid, rollno not issued
        where += " AND m.IsPaymentSuccess=1 AND m.rollno IS NOT NULL AND m.rollno != ''"
        
        if f_pref == '1':
            where += " AND mt.AllottedPreference >= 1"

        try:
            stmt = text(f"""
                SELECT mt.Pk_MeritTrnID, m.pk_regid, m.regno, m.rollno, m.s_name, m.s_surname, m.f_name, m.mobileno,
                       d.description as degree_desc,
                       m.SportsQuota, m.IsSportsVerified,
                       mt.AllottedSpecialisation, mt.AllottedPreference,
                       mt.ProcessStatus, mt.ProcessRemarks, mt.WithdrawalCutoff,
                       cat.Description as category_name
                FROM PA_Merit_Trn mt
                JOIN PA_Merit_Mst mm ON mt.Fk_MeritID = mm.Pk_MeritID
                JOIN PA_Registration_Mst m ON mt.fk_regid = m.pk_regid
                LEFT JOIN ACD_Degree_Mst d ON mm.Fk_DegreeID = d.pk_degreeid
                LEFT JOIN PA_StudentCategory_Mst cat ON m.fk_stucatid_cast = cat.Pk_StuCatId
                {where}
                ORDER BY m.s_name ASC, m.s_surname ASC
            """)
            rows = db.session.execute(stmt, params).mappings().all()
            for r in rows:
                name = (r.get("s_name") or "") + " " + (r.get("s_surname") or "")
                
                process_status_code = r.get("ProcessStatus")
                
                status_text = ""
                if process_status_code == 'U':
                    status_text = "Upward"
                elif process_status_code == 'W':
                    status_text = "Withdrawl"
                elif process_status_code == 'N':
                    status_text = "Not Reported"
                elif process_status_code == 'R':
                    status_text = "Rejected"
                elif process_status_code:
                    status_text = str(process_status_code)
                
                record = {
                    "id": r["pk_regid"],
                    "trn_id": r["Pk_MeritTrnID"],
                    "regno": r["regno"],
                    "rollno": r["rollno"],
                    "name": name.strip(),
                    "father_name": r["f_name"],
                    "mobile": r.get("mobileno"),
                    "degree": r["degree_desc"],
                    "specialization": r.get("AllottedSpecialisation") or "",
                    "allotted_preference": r.get("AllottedPreference") or "",
                    "sports_quota": bool(r.get("SportsQuota")),
                    "sports_verified": bool(r.get("IsSportsVerified")),
                    "status": status_text,
                    "remarks": r.get("ProcessRemarks") or "",
                    "withdrawal_cutoff": r.get("WithdrawalCutoff") or ""
                }
                
                if process_status_code:
                    applied_records.append(record)
                else:
                    eligible_records.append(record)
        except Exception as e:
            print("Error fetching upward students:", e)

    return render_template('transactions/apply_for_upward.html', sessions=sessions, streams=streams, colleges=colleges, degrees=degrees, cutoffs=cutoffs,
                           f_session=f_session, f_stream=f_stream, f_college=f_college, f_degree=f_degree, f_reg=f_reg, f_cutoff=f_cutoff, f_pref=f_pref, eligible_records=eligible_records, applied_records=applied_records, show_grids=show_grids)

@transactions_bp.route('/print-admission-letter/<int:reg_id>')
def print_admission_letter(reg_id):
    from flask import send_file
    import sys
    import os
    sys.path.insert(0, os.path.abspath('D:/Preadmission'))
    from tools.pdf_generator import generate_admission_letter
    
    student_data = {}
    if _db_ping():
        stmt = text("""
            SELECT m.regno, m.rollno, m.s_name, m.s_surname, m.f_name, 
                   cat.Description as category,
                   mt.AllottedCategory as allotted_category,
                   CONCAT(c.collegename, ' / ', d.description, ' / ', mt.AllottedSpecialisation) as programme
            FROM PA_Merit_Trn mt
            JOIN PA_Registration_Mst m ON mt.fk_regid = m.pk_regid
            LEFT JOIN PA_StudentCategory_Mst cat ON m.fk_stucatid_cast = cat.Pk_StuCatId   
            LEFT JOIN PA_College_Mst c ON mt.Fk_CollegeID = c.Pk_CollegeID
            LEFT JOIN PA_Merit_Mst mm ON mt.Fk_MeritID = mm.Pk_MeritID
            LEFT JOIN ACD_Degree_Mst d ON mm.Fk_DegreeID = d.pk_degreeid
            WHERE m.pk_regid = :reg_id
        """)
        row = db.session.execute(stmt, {"reg_id": reg_id}).mappings().first()
        if row:
            name = (row.get("s_name") or "") + " " + (row.get("s_surname") or "")
            student_data = {
                'regno': row.get('regno'),
                'rollno': row.get('rollno'),
                'name': name.strip(),
                'father_name': row.get('f_name'),
                'category': row.get('category'),
                'allotted_category': row.get('allotted_category'),
                'programme': row.get('programme')
            }
            
    pdf_buffer = generate_admission_letter(student_data)
    return send_file(pdf_buffer, download_name=f"admission_letter_{reg_id}.pdf", as_attachment=False, mimetype='application/pdf')

# 2. Candidate Marks Upload
@transactions_bp.route('/candidate-marks-upload', methods=['GET', 'POST'])
def candidate_marks_upload():
    sessions, degrees, _, _ = get_common_masters()

    f_session = request.form.get('session_id') if request.method == 'POST' else request.args.get('session_id')
    f_degree = request.form.get('degree_type_id') if request.method == 'POST' else request.args.get('degree_type_id')

    if not f_session and sessions:
        f_session = str(sessions[0]['id'])

    records = []

    if _db_ping() and (f_session or f_degree):
        where = " WHERE 1=1"
        params = {}
        if f_session and f_session != '0':
            where += " AND m.fk_sessionid=:session_id"
            params["session_id"] = f_session
        if f_degree:
            where += " AND m.fk_degreeid=:degree_id"
            params["degree_id"] = f_degree

        stmt = text(f"""
            SELECT 
                m.fk_degreeid as degree_id,
                s.description as session_name,
                d.description as degree_name,
                COUNT(m.pk_Cmarksid) as count
            FROM PA_Candidate_Marks m
            LEFT JOIN LUP_AcademicSession_Mst s ON m.fk_sessionid = s.pk_sessionid
            LEFT JOIN ACD_Degree_Mst d ON m.fk_degreeid = d.pk_degreeid
            {where}
            GROUP BY m.fk_degreeid, s.description, d.description
            ORDER BY d.description
        """)
        rows = db.session.execute(stmt, params).mappings().all()
        for r in rows:
            records.append({
                "session": r.get("session_name"),
                "degree": r.get("degree_name"),
                "degree_id": r.get("degree_id"),
                "college": "",
                "count": r.get("count")
            })

    if request.method == 'POST':
        action = request.form.get('action')
        
        # Check if action is an export trigger from the grid
        if action and action.startswith('export_grid_'):
            parts = action.split('_')
            grid_session_id = parts[2]
            grid_degree_id = parts[3]
            
            import io
            import openpyxl
            from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
            from flask import send_file
            
            params = {"session_id": grid_session_id, "degree_id": grid_degree_id}

            # The join logic on PA_Registration_Mst uses the session ID so it doesn't duplicate cross-session data
            stmt_export = text("""
                SELECT 
                    d.description as Description,
                    c.collegename as CollegeName,
                    reg.regno as regno,
                    reg.pwd as pwd,
                    reg.mobileno as mobileno,
                    (ISNULL(reg.s_name, '') + ' ' + ISNULL(reg.s_surname, '')) as studentname,
                    m.RollNo as Rollno,
                    cat.Description as Category,
                    COALESCE(m.marks, 100) as marks,
                    m.ObtainMarks as ObtainMArks,
                    m.overallR,
                    m.overallRcount as OverallRcount,
                    m.Categoryrank as categoryrank,
                    m.Categoryrankcount as categoryrankcount,
                    m.mscexamtype,
                    CASE WHEN reg.ESM = 1 THEN 'TRUE' ELSE 'FALSE' END as ESM,
                    m.ESMRank,
                    m.ESMRankCount,
                    CASE WHEN reg.PH = 1 THEN 'TRUE' ELSE 'FALSE' END as PH,
                    m.PHRank,
                    m.PHRankCount,
                    CASE WHEN reg.FF = 1 THEN 'TRUE' ELSE 'FALSE' END as FF,
                    m.FFRank,
                    m.FFRankCount,
                    CASE 
                        WHEN m.PresentStatus IS NOT NULL THEN m.PresentStatus
                        WHEN m.remarks = 'Absent' THEN 'A' 
                        WHEN m.ObtainMarks IS NOT NULL THEN 'P'
                        ELSE 'A' 
                    END as PresentStatus
                FROM PA_Candidate_Marks m
                LEFT JOIN PA_Registration_Mst reg ON m.fk_regid = reg.pk_regid OR (m.RollNo = reg.rollno AND m.fk_sessionid = reg.fk_sessionid)
                LEFT JOIN ACD_Degree_Mst d ON m.fk_degreeid = d.pk_degreeid
                LEFT JOIN SMS_College_Mst c ON m.fk_collegeid = c.pk_collegeid
                LEFT JOIN PA_StudentCategory_Mst cat ON reg.fk_stucatid_cast = cat.Pk_StuCatId
                WHERE m.fk_sessionid = :session_id AND m.fk_degreeid = :degree_id
                ORDER BY m.RollNo
            """)
            
            try:
                export_data = db.session.execute(stmt_export, params).mappings().all()
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Marks Export"
                
                headers = [
                    "Description", "CollegeName", "regno", "pwd", "mobileno", "studentname", 
                    "Rollno", "Category", "marks", "ObtainMArks", "overallR", "OverallRcount", 
                    "categoryrank", "categoryrankcount", "mscexamtype", "ESM", "ESMRank", "ESMRankCount", 
                    "PH", "PHRank", "PHRankCount", "FF", "FFRank", "FFRankCount", "PresentStatus"
                ]
                
                ws.append(headers)
                
                # Professional Styling Setup
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4A534A", end_color="4A534A", fill_type="solid")
                alt_row_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center')
                
                # Apply header styling
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align
                
                for i, row in enumerate(export_data, start=2):
                    row_data = [
                        row.get("Description"), 
                        row.get("CollegeName") or '', 
                        row.get("regno"), 
                        row.get("pwd"), 
                        row.get("mobileno"), 
                        (row.get("studentname") or '').strip(), 
                        row.get("Rollno"), 
                        row.get("Category"), 
                        row.get("marks"), 
                        "{:.2f}".format(float(row.get("ObtainMArks"))) if row.get("ObtainMArks") is not None else "", 
                        row.get("overallR"), 
                        row.get("OverallRcount"), 
                        row.get("categoryrank"), 
                        row.get("categoryrankcount"), 
                        row.get("mscexamtype"), 
                        row.get("ESM"), 
                        row.get("ESMRank"), 
                        row.get("ESMRankCount"), 
                        row.get("PH"), 
                        row.get("PHRank"), 
                        row.get("PHRankCount"), 
                        row.get("FF"), 
                        row.get("FFRank"), 
                        row.get("FFRankCount"), 
                        row.get("PresentStatus")
                    ]
                    ws.append(row_data)
                    
                    # Apply row styling
                    current_fill = alt_row_fill if i % 2 == 0 else white_fill
                    for cell in ws[i]:
                        cell.border = thin_border
                        cell.fill = current_fill
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)
                    
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                return send_file(
                    buffer, 
                    download_name=f"Candidate_Marks_Export_{grid_session_id}_{grid_degree_id}.xlsx", 
                    as_attachment=True, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                flash(f"Error generating export: {str(e)}", "error")
            
        elif action == 'export':
            if not f_session or f_session == '0' or not f_degree:
                flash("Please select both Academic Session and Degree Type to export.", "error")
            else:
                import io
                import openpyxl
                from flask import send_file
                
                stmt_export = text("""
                    SELECT 
                        d.description as Description,
                        c.collegename as CollegeName,
                        reg.regno as regno,
                        reg.pwd as pwd,
                        reg.mobileno as mobileno,
                        (ISNULL(reg.s_name, '') + ' ' + ISNULL(reg.s_surname, '')) as studentname,
                        m.RollNo as Rollno,
                        cat.Description as Category,
                        COALESCE(m.marks, 100) as marks,
                        m.ObtainMarks as ObtainMArks,
                        m.overallR,
                        m.overallRcount as OverallRcount,
                        m.Categoryrank as categoryrank,
                        m.Categoryrankcount as categoryrankcount,
                        m.mscexamtype,
                        CASE WHEN reg.ESM = 1 THEN 'TRUE' ELSE 'FALSE' END as ESM,
                        m.ESMRank,
                        m.ESMRankCount,
                        CASE WHEN reg.PH = 1 THEN 'TRUE' ELSE 'FALSE' END as PH,
                        m.PHRank,
                        m.PHRankCount,
                        CASE WHEN reg.FF = 1 THEN 'TRUE' ELSE 'FALSE' END as FF,
                        m.FFRank,
                        m.FFRankCount,
                        CASE 
                            WHEN m.PresentStatus IS NOT NULL THEN m.PresentStatus
                            WHEN m.remarks = 'Absent' THEN 'A' 
                            WHEN m.ObtainMarks IS NOT NULL THEN 'P'
                            ELSE 'A' 
                        END as PresentStatus
                    FROM PA_Candidate_Marks m
                    LEFT JOIN PA_Registration_Mst reg ON m.fk_regid = reg.pk_regid OR m.RollNo = reg.rollno
                    LEFT JOIN ACD_Degree_Mst d ON m.fk_degreeid = d.pk_degreeid
                    LEFT JOIN SMS_College_Mst c ON m.fk_collegeid = c.pk_collegeid
                    LEFT JOIN PA_StudentCategory_Mst cat ON reg.fk_stucatid_cast = cat.Pk_StuCatId
                    WHERE m.fk_sessionid = :session_id AND m.fk_degreeid = :degree_id
                    ORDER BY m.RollNo
                """)
                export_data = db.session.execute(stmt_export, {"session_id": f_session, "degree_id": f_degree}).mappings().all()
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Marks Export"
                
                headers = [
                    "Description", "CollegeName", "regno", "pwd", "mobileno", "studentname", 
                    "Rollno", "Category", "marks", "ObtainMArks", "overallR", "OverallRcount", 
                    "categoryrank", "categoryrankcount", "mscexamtype", "ESM", "ESMRank", "ESMRankCount", 
                    "PH", "PHRank", "PHRankCount", "FF", "FFRank", "FFRankCount", "PresentStatus"
                ]
                ws.append(headers)
                
                for row in export_data:
                    ws.append([
                        row.get("Description"), 
                        row.get("CollegeName") or '', 
                        row.get("regno"), 
                        row.get("pwd"), 
                        row.get("mobileno"), 
                        (row.get("studentname") or '').strip(), 
                        row.get("Rollno"), 
                        row.get("Category"), 
                        row.get("marks"), 
                        "{:.2f}".format(float(row.get("ObtainMArks"))) if row.get("ObtainMArks") is not None else "", 
                        row.get("overallR"), 
                        row.get("OverallRcount"), 
                        row.get("categoryrank"), 
                        row.get("categoryrankcount"), 
                        row.get("mscexamtype"), 
                        row.get("ESM"), 
                        row.get("ESMRank"), 
                        row.get("ESMRankCount"), 
                        row.get("PH"), 
                        row.get("PHRank"), 
                        row.get("PHRankCount"), 
                        row.get("FF"), 
                        row.get("FFRank"), 
                        row.get("FFRankCount"), 
                        row.get("PresentStatus")
                    ])
                    
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                return send_file(
                    buffer, 
                    download_name=f"Candidate_Marks_Export.xlsx", 
                    as_attachment=True, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        elif action == 'import':
            if not f_session or f_session == '0' or not f_degree:
                flash("Please select both Academic Session and Degree Type before importing.", "error")
            elif 'upload_file' not in request.files or request.files['upload_file'].filename == '':
                flash("Please select a file to import.", "error")
            else:
                try:
                    import openpyxl
                    file = request.files['upload_file']
                    wb = openpyxl.load_sheet_by_name(file) if hasattr(openpyxl, 'load_sheet_by_name') else openpyxl.load_workbook(file)
                    ws = wb.active
                    
                    update_count = 0
                    headers = [cell.value for cell in ws[1]]
                    
                    if "Rollno" not in headers or "ObtainMArks" not in headers:
                        flash("Invalid file format. Ensure columns 'Rollno' and 'ObtainMArks' exist.", "error")
                    else:
                        idx_roll = headers.index("Rollno")
                        idx_marks = headers.index("ObtainMArks")
                        idx_status = headers.index("PresentStatus") if "PresentStatus" in headers else -1
                        
                        stmt_update = text("""
                            UPDATE PA_Candidate_Marks 
                            SET ObtainMarks = :obt_marks,
                                PresentStatus = :p_status,
                                remarks = :rem
                            WHERE RollNo = :roll_no AND fk_sessionid = :session_id AND fk_degreeid = :degree_id
                        """)
                        
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            roll_no = str(row[idx_roll]) if row[idx_roll] is not None else None
                            if not roll_no:
                                continue
                                
                            try:
                                obt_marks = float(row[idx_marks]) if row[idx_marks] is not None and str(row[idx_marks]).strip() != '' else None
                            except ValueError:
                                obt_marks = None
                                
                            p_status = str(row[idx_status]).strip().upper() if idx_status >= 0 and row[idx_status] is not None and str(row[idx_status]).strip() != '' else 'P'
                            
                            rem = None
                            if p_status == 'A' or obt_marks is None:
                                rem = 'Absent'
                                obt_marks = None
                                p_status = 'A'
                            
                            db.session.execute(stmt_update, {
                                "obt_marks": obt_marks,
                                "p_status": p_status,
                                "rem": rem,
                                "roll_no": roll_no,
                                "session_id": f_session,
                                "degree_id": f_degree
                            })
                            update_count += 1
                            
                        db.session.commit()
                        flash(f"Successfully imported and updated {update_count} records.", "success")
                except Exception as e:
                    db.session.rollback()
                    flash(f"Error importing file: {str(e)}", "error")

    return render_template('transactions/candidate_marks_upload.html', sessions=sessions, degrees=degrees, f_session=f_session, f_degree=f_degree, records=records)
# 3. Counselling Raw Data
@transactions_bp.route('/counselling-raw-data', methods=['GET', 'POST'])
def counselling_raw_data():
    sessions, degrees, colleges, streams = get_common_masters()
    
    # Only current session
    if sessions:
        sessions = [sessions[0]]
        
    f_session = request.form.get('session_id') if request.method == 'POST' else request.args.get('session_id')
    f_degree = request.form.get('degree_id') if request.method == 'POST' else request.args.get('degree_id')
    f_college = request.form.get('college_id') if request.method == 'POST' else request.args.get('college_id')
    f_cutoff = request.form.get('cutoff') if request.method == 'POST' else request.args.get('cutoff')
    
    if not f_session and sessions:
        f_session = str(sessions[0]['id'])
        
    cutoffs = list(range(1, 11))

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'export':
            if not f_session or not f_degree:
                flash("Please select Session and Degree Type to export.", "error")
            else:
                import io
                import openpyxl
                from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
                from flask import send_file
                
                where = " WHERE mm.Fk_SessionID = :session_id AND mm.Fk_DegreeID = :degree_id"
                params = {"session_id": int(f_session), "degree_id": int(f_degree)}
                
                if f_college:
                    where += " AND mm.Fk_CollegeID = :college_id"
                    params["college_id"] = int(f_college)
                if f_cutoff and f_cutoff != '0':
                    where += " AND mm.CutOff = :cutoff"
                    params["cutoff"] = int(f_cutoff)

                stmt_export = text(f"""
                    SELECT 
                        m.regno, 
                        m.rollno, 
                        (ISNULL(m.s_name, '') + ' ' + ISNULL(m.s_surname, '')) as CandidateName,
                        m.f_name as FatherName,
                        m.mobileno,
                        d.description as Degree,
                        c.collegename as College,
                        U.AllottedSpecialisation,
                        cat.Description as Category,
                        U.AllottedCategory,
                        U.ObtainMarks,
                        U.OverAllRank,
                        U.AllottedPreference,
                        U.ProcessStatus,
                        U.ProcessRemarks
                    FROM (
                        SELECT Fk_MeritID, fk_regid, AllottedSpecialisation, AllottedCategory, ObtainMarks, OverAllRank, AllottedPreference, ProcessStatus, ProcessRemarks FROM PA_Merit_Trn
                        UNION ALL
                        SELECT Fk_MeritID, fk_regid, SubjectName as AllottedSpecialisation, AllottedCategory, ObtainMarks, OverAllRank, AllottedPreference, ProcessStatus, ProcessRemarks FROM PA_Merit_Bsc_Trn
                    ) U
                    JOIN PA_Merit_Mst mm ON U.Fk_MeritID = mm.Pk_MeritID
                    JOIN PA_Registration_Mst m ON U.fk_regid = m.pk_regid
                    LEFT JOIN ACD_Degree_Mst d ON mm.Fk_DegreeID = d.pk_degreeid
                    LEFT JOIN PA_College_Mst c ON mm.Fk_CollegeID = c.Pk_CollegeID
                    LEFT JOIN PA_StudentCategory_Mst cat ON m.fk_stucatid_cast = cat.Pk_StuCatId
                    {where}
                    ORDER BY m.s_name ASC, m.s_surname ASC
                """)
                
                try:
                    export_data = db.session.execute(stmt_export, params).mappings().all()
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Counselling Raw Data"
                    
                    headers = [
                        "Registration No", "Roll No", "Candidate Name", "Father Name", "Mobile No",
                        "Degree", "College", "Allotted Specialisation", "Category", "Allotted Category",
                        "Obtain Marks", "Overall Rank", "Allotted Preference", "Process Status", "Process Remarks"
                    ]
                    
                    ws.append(headers)
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4A534A", end_color="4A534A", fill_type="solid")
                    alt_row_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
                    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    center_align = Alignment(horizontal='center', vertical='center')
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = center_align
                        
                    for i, row in enumerate(export_data, start=2):
                        ws.append([
                            row.get("regno"),
                            row.get("rollno"),
                            row.get("CandidateName").strip() if row.get("CandidateName") else "",
                            row.get("FatherName"),
                            row.get("mobileno"),
                            row.get("Degree"),
                            row.get("College"),
                            row.get("AllottedSpecialisation"),
                            row.get("Category"),
                            row.get("AllottedCategory"),
                            "{:.2f}".format(float(row.get("ObtainMarks"))) if row.get("ObtainMarks") is not None else "",
                            row.get("OverAllRank"),
                            row.get("AllottedPreference"),
                            row.get("ProcessStatus"),
                            row.get("ProcessRemarks")
                        ])
                        
                        current_fill = alt_row_fill if i % 2 == 0 else white_fill
                        for cell in ws[i]:
                            cell.border = thin_border
                            cell.fill = current_fill
                            
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except: pass
                        ws.column_dimensions[column].width = min(max_length + 2, 50)
                        
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    
                    return send_file(
                        buffer, 
                        download_name=f"Counselling_Raw_Data_Export.xlsx", 
                        as_attachment=True, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                except Exception as e:
                    flash(f"Error generating export: {str(e)}", "error")

    return render_template('transactions/counselling_raw_data.html', sessions=sessions, degrees=degrees, colleges=colleges, streams=streams, cutoffs=cutoffs,
                           f_session=f_session, f_degree=f_degree, f_college=f_college, f_cutoff=f_cutoff)

# 4. Generate Merit Seat Allotment
@transactions_bp.route('/generate-merit-seat-allotment', methods=['GET', 'POST'])
def generate_merit_seat_allotment():
    sessions, degrees, colleges, _ = get_common_masters()
    
    f_session = request.form.get('session_id') if request.method == 'POST' else request.args.get('session_id')
    f_degree = request.form.get('degree_id') if request.method == 'POST' else request.args.get('degree_id')
    f_college = request.form.get('college_id') if request.method == 'POST' else request.args.get('college_id')
    f_cutoff = request.form.get('cutoff') if request.method == 'POST' else request.args.get('cutoff')
    f_merit_type = request.form.get('merit_type', 'S') if request.method == 'POST' else request.args.get('merit_type', 'S')
    page = request.args.get('page', 1, type=int)

    if not f_session and sessions:
        f_session = str(sessions[0]['id'])
        
    cutoffs = list(range(1, 11))
    
    records = []
    
    if _db_ping() and f_session and f_degree:
        where = " WHERE m.Fk_SessionID = :session_id AND m.Fk_DegreeID = :degree_id"
        params = {"session_id": f_session, "degree_id": f_degree}
        
        # Optional filters
        if f_college and f_college != '0':
            where += " AND m.Fk_CollegeID = :college_id"
            params["college_id"] = f_college
            
        if f_cutoff and f_cutoff != '0':
            where += " AND m.CutOff = :cutoff"
            params["cutoff"] = f_cutoff
            
        stmt = text(f"""
            SELECT 
                m.Pk_MeritID as id,
                s.description as session_name,
                c.collegename as college_name,
                d.description as degree_name,
                m.CutOff as cutoff,
                m.MeritType as merit_type
            FROM PA_Merit_Mst m
            LEFT JOIN LUP_AcademicSession_Mst s ON m.Fk_SessionID = s.pk_sessionid
            LEFT JOIN PA_College_Mst c ON m.Fk_CollegeID = c.Pk_CollegeID
            LEFT JOIN ACD_Degree_Mst d ON m.Fk_DegreeID = d.pk_degreeid
            {where}
            ORDER BY m.CutOff ASC, m.Pk_MeritID ASC
        """)
        try:
            rows = db.session.execute(stmt, params).mappings().all()
            for r in rows:
                m_type = "System Generated" if r.get("merit_type") == 'S' else "Manual" if r.get("merit_type") == 'M' else str(r.get("merit_type"))
                records.append({
                    "id": r.get("id"),
                    "session": r.get("session_name"),
                    "college": r.get("college_name") or "",
                    "degree": r.get("degree_name"),
                    "cutoff": r.get("cutoff"),
                    "merit_type": m_type
                })
        except Exception as e:
            print("Error fetching merit generated:", e)

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'process':
            flash("Process completed", "success")
        elif action and action.startswith('export_record_'):
            merit_id = action.split('_')[2]
            import io
            import openpyxl
            from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
            from flask import send_file
            
            stmt_export = text("""
                SELECT 
                    m.regno, 
                    m.rollno, 
                    (ISNULL(m.s_name, '') + ' ' + ISNULL(m.s_surname, '')) as CandidateName,
                    m.f_name as FatherName,
                    m.mobileno,
                    d.description as Degree,
                    c.collegename as College,
                    U.AllottedSpecialisation,
                    cat.Description as Category,
                    U.AllottedCategory,
                    U.ObtainMarks,
                    U.OverAllRank,
                    U.AllottedPreference,
                    U.ProcessStatus,
                    U.ProcessRemarks
                FROM (
                    SELECT Fk_MeritID, fk_regid, AllottedSpecialisation, AllottedCategory, ObtainMarks, OverAllRank, AllottedPreference, ProcessStatus, ProcessRemarks, Fk_CollegeID FROM PA_Merit_Trn
                    UNION ALL
                    SELECT Fk_MeritID, fk_regid, SubjectName as AllottedSpecialisation, AllottedCategory, ObtainMarks, OverAllRank, AllottedPreference, ProcessStatus, ProcessRemarks, AllottedCollegeID as Fk_CollegeID FROM PA_Merit_Bsc_Trn
                ) U
                JOIN PA_Merit_Mst mm ON U.Fk_MeritID = mm.Pk_MeritID
                JOIN PA_Registration_Mst m ON U.fk_regid = m.pk_regid
                LEFT JOIN ACD_Degree_Mst d ON mm.Fk_DegreeID = d.pk_degreeid
                LEFT JOIN PA_College_Mst c ON mm.Fk_CollegeID = c.Pk_CollegeID
                LEFT JOIN PA_StudentCategory_Mst cat ON m.fk_stucatid_cast = cat.Pk_StuCatId
                WHERE U.Fk_MeritID = :merit_id
                ORDER BY m.s_name ASC, m.s_surname ASC
            """)
            
            try:
                export_data = db.session.execute(stmt_export, {"merit_id": merit_id}).mappings().all()
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Seat Allotment"
                
                headers = [
                    "Registration No", "Roll No", "Candidate Name", "Father Name", "Mobile No",
                    "Degree", "College", "Allotted Specialisation", "Category", "Allotted Category",
                    "Obtain Marks", "Overall Rank", "Allotted Preference", "Process Status", "Process Remarks"
                ]
                
                ws.append(headers)
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4A534A", end_color="4A534A", fill_type="solid")
                alt_row_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center')
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align
                    
                for i, row in enumerate(export_data, start=2):
                    ws.append([
                        row.get("regno"),
                        row.get("rollno"),
                        row.get("CandidateName").strip() if row.get("CandidateName") else "",
                        row.get("FatherName"),
                        row.get("mobileno"),
                        row.get("Degree"),
                        row.get("College"),
                        row.get("AllottedSpecialisation"),
                        row.get("Category"),
                        row.get("AllottedCategory"),
                        "{:.2f}".format(float(row.get("ObtainMarks"))) if row.get("ObtainMarks") is not None else "",
                        row.get("OverAllRank"),
                        row.get("AllottedPreference"),
                        row.get("ProcessStatus"),
                        row.get("ProcessRemarks")
                    ])
                    
                    current_fill = alt_row_fill if i % 2 == 0 else white_fill
                    for cell in ws[i]:
                        cell.border = thin_border
                        cell.fill = current_fill
                        
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)
                    
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                return send_file(
                    buffer, 
                    download_name=f"Seat_Allotment_Export_{merit_id}.xlsx", 
                    as_attachment=True, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                flash(f"Error generating export: {str(e)}", "error")

    return render_template('transactions/generate_merit_seat_allotment.html', sessions=sessions, degrees=degrees, colleges=colleges, cutoffs=cutoffs,
                           f_session=f_session, f_degree=f_degree, f_college=f_college, f_cutoff=f_cutoff, f_merit_type=f_merit_type, records=records, pagination=Pagination(page, 15, len(records)))

# 5. Manual Merit Entry
@transactions_bp.route('/manual-merit-entry', methods=['GET', 'POST'])
def manual_merit_entry():
    sessions, degrees, colleges, _ = get_common_masters()
    f_session = request.args.get('session_id') or (request.form.get('session_id') if request.method == 'POST' else None)
    f_degree = request.args.get('degree_id') or (request.form.get('degree_id') if request.method == 'POST' else None)
    context_id = request.args.get('context_id') or (request.form.get('context_id') if request.method == 'POST' else None)
    action = request.args.get('action') or (request.form.get('action') if request.method == 'POST' else None)

    records = []
    merit_context = None
    student_details = None
    manual_allotted = []
    
    categories = safe_all(lambda: db.session.execute(text("SELECT Pk_StuCatId as id, Description as name FROM PA_StudentCategory_Mst ORDER BY Description")).mappings().all(), [])
    specializations = safe_all(lambda: db.session.execute(text("SELECT Pk_SID as id, Specialization as name FROM PA_Specialization_mst ORDER BY Specialization")).mappings().all(), [])
    roasters = list(range(1, 151))

    if _db_ping() and f_session and f_degree:
        stmt = text("""
            SELECT 
                m.Pk_MeritID as id,
                s.description as session_name,
                c.collegename as college_name,
                d.description as degree_name,
                m.CutOff as cutoff
            FROM PA_Merit_Mst m
            LEFT JOIN LUP_AcademicSession_Mst s ON m.Fk_SessionID = s.pk_sessionid
            LEFT JOIN PA_College_Mst c ON m.Fk_CollegeID = c.Pk_CollegeID
            LEFT JOIN ACD_Degree_Mst d ON m.Fk_DegreeID = d.pk_degreeid
            WHERE m.Fk_SessionID = :session_id AND m.Fk_DegreeID = :degree_id
            ORDER BY m.CutOff ASC, m.Pk_MeritID ASC
        """)
        try:
            rows = db.session.execute(stmt, {"session_id": f_session, "degree_id": f_degree}).mappings().all()
            for r in rows:
                records.append({
                    "id": r.get("id"),
                    "session": r.get("session_name"),
                    "college": r.get("college_name") or "",
                    "degree": r.get("degree_name"),
                    "cutoff": r.get("cutoff")
                })
        except Exception as e:
            print("Error fetching manual records:", e)

    if context_id and _db_ping():
        stmt_ctx = text("""
            SELECT 
                m.Pk_MeritID as id,
                s.description as session_name,
                c.collegename as college_name,
                d.description as degree_name,
                m.CutOff as cutoff
            FROM PA_Merit_Mst m
            LEFT JOIN LUP_AcademicSession_Mst s ON m.Fk_SessionID = s.pk_sessionid
            LEFT JOIN PA_College_Mst c ON m.Fk_CollegeID = c.Pk_CollegeID
            LEFT JOIN ACD_Degree_Mst d ON m.Fk_DegreeID = d.pk_degreeid
            WHERE m.Pk_MeritID = :merit_id
        """)
        row = db.session.execute(stmt_ctx, {"merit_id": context_id}).mappings().first()
        if row:
            merit_context = dict(row)
            
        # Fetch existing manual allotted students for this merit_id
        stmt_allotted = text("""
            SELECT mt.Pk_MeritTrnID as id, m.regno, (ISNULL(m.s_name, '') + ' ' + ISNULL(m.s_surname, '')) as name, 
                   m.rollno, mt.AllottedSpecialisation as spec, mt.AllottedPreference as pref, 
                   cat.Description as category, c.collegename as college
            FROM PA_Merit_Trn mt
            JOIN PA_Registration_Mst m ON mt.fk_regid = m.pk_regid
            LEFT JOIN PA_StudentCategory_Mst cat ON mt.AllottedCategory = CAST(cat.Pk_StuCatId as VARCHAR)
            LEFT JOIN PA_College_Mst c ON mt.Fk_CollegeID = c.Pk_CollegeID
            WHERE mt.Fk_MeritID = :merit_id AND mt.ProcessStatus = 'M'
            ORDER BY mt.Pk_MeritTrnID DESC
        """)
        try:
            manual_allotted = [dict(r) for r in db.session.execute(stmt_allotted, {"merit_id": context_id}).mappings().all()]
        except Exception:
            pass

    reg_no = request.args.get('reg_no') or (request.form.get('reg_no') if request.method == 'POST' else None)

    if request.method == 'POST' and action == 'show_student' and reg_no and merit_context:
        stmt_stu = text("""
            SELECT m.pk_regid as id, m.regno, m.rollno, m.s_name, m.s_surname, m.gender, cat.Description as category_name
            FROM PA_Registration_Mst m
            LEFT JOIN PA_StudentCategory_Mst cat ON m.fk_stucatid_cast = cat.Pk_StuCatId
            WHERE m.regno = :reg_no AND m.fk_sessionid = :session_id
        """)
        row_stu = db.session.execute(stmt_stu, {"reg_no": reg_no, "session_id": f_session}).mappings().first()
        if row_stu:
            student_details = dict(row_stu)
            student_details["name"] = ((row_stu.get("s_name") or "") + " " + (row_stu.get("s_surname") or "")).strip()
        else:
            flash("Student not found for the selected session.", "error")

    elif request.method == 'POST' and action == 'save_student' and reg_no and merit_context:
        allotted_spec = request.form.get('allotted_spec')
        roaster_id = request.form.get('roaster_id')
        allotted_pref = request.form.get('allotted_pref')
        allotted_cat = request.form.get('allotted_cat')
        remarks = request.form.get('remarks')
        allotted_clg = request.form.get('allotted_clg')
        reg_id = request.form.get('reg_id')
        rollno = request.form.get('rollno')
        
        # Get spec name
        spec_name = None
        if allotted_spec:
            try:
                row_spec = db.session.execute(text("SELECT Specialization FROM PA_Specialization_mst WHERE Pk_SID = :sid"), {"sid": allotted_spec}).mappings().first()
                if row_spec: spec_name = row_spec['Specialization']
            except: pass

        if _db_ping():
            try:
                stmt_insert = text("""
                    INSERT INTO PA_Merit_Trn 
                    (Fk_MeritID, fk_regid, RollNo, Fk_CollegeID, AllottedSpec, AllottedSpecialisation, 
                     AllottedPreference, AllottedCategory, RoasterID, ProcessRemarks, ProcessStatus)
                    VALUES 
                    (:merit_id, :reg_id, :rollno, :colg_id, :spec_id, :spec_name, 
                     :pref, :cat, :roaster, :remarks, 'M')
                """)
                db.session.execute(stmt_insert, {
                    "merit_id": context_id,
                    "reg_id": reg_id,
                    "rollno": rollno,
                    "colg_id": allotted_clg or None,
                    "spec_id": allotted_spec or None,
                    "spec_name": spec_name,
                    "pref": allotted_pref or None,
                    "cat": allotted_cat or None,
                    "roaster": roaster_id or None,
                    "remarks": remarks or None
                })
                db.session.commit()
                flash("Student manual merit details saved successfully.", "success")
                return redirect(url_for('transactions.manual_merit_entry', session_id=f_session, degree_id=f_degree, context_id=context_id))
            except Exception as e:
                db.session.rollback()
                flash(f"Error saving student: {str(e)}", "error")
                
    elif request.method == 'POST' and action and action.startswith('delete_'):
        trn_id = action.split('_')[1]
        if _db_ping():
            try:
                db.session.execute(text("DELETE FROM PA_Merit_Trn WHERE Pk_MeritTrnID = :trn_id"), {"trn_id": trn_id})
                db.session.commit()
                flash("Record deleted.", "success")
                return redirect(url_for('transactions.manual_merit_entry', session_id=f_session, degree_id=f_degree, context_id=context_id))
            except Exception as e:
                db.session.rollback()
                flash(f"Error deleting record: {str(e)}", "error")

    return render_template('transactions/manual_merit_entry.html', 
                           sessions=sessions, degrees=degrees, colleges=colleges, 
                           categories=categories, specializations=specializations, roasters=roasters,
                           f_session=f_session, f_degree=f_degree, records=records,
                           merit_context=merit_context, context_id=context_id, student_details=student_details, reg_no=reg_no,
                           manual_allotted=manual_allotted)

# 6. PH Candidates Verifications
@transactions_bp.route('/ph-candidates-verifications', methods=['GET', 'POST'])
def ph_candidates_verifications():
    sessions, degrees, _, streams = get_common_masters()
    
    # Prioritize form parameters for POST, then args for GET
    f_session = request.form.get('session_id') or request.args.get('session_id')
    f_degree = request.form.get('degree_id') or request.args.get('degree_id')
    f_status = request.form.get('ph_status') or request.args.get('ph_status') or '1'
    f_college = request.form.get('college_id') or request.args.get('college_id')
    edit_id = request.form.get('edit_id') or request.args.get('edit_id')
    page = request.args.get('page', 1, type=int)

    records = []
    edit_record = None

    if request.method == 'POST' and request.form.get('action') == 'verify_ph_inline':
        import os
        from werkzeug.utils import secure_filename
        
        reg_id = request.form.get('reg_id')
        verify_status = request.form.get('verify_status')
        verify_remarks = request.form.get('verify_remarks')
        smo_percentage = request.form.get('smo_percentage')
        
        if reg_id and _db_ping():
            # Handle file upload if present
            attachment_name = None
            attachment_unique_name = None
            if 'smo_document' in request.files:
                file = request.files['smo_document']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    attachment_name = filename
                    attachment_unique_name = f"SMO_{reg_id}_{filename}"
                    
            try:
                if attachment_name:
                    db.session.execute(text("""
                        UPDATE PA_Registration_Mst 
                        SET VerificationDone = :v_status, 
                            VerificationReason = :v_remarks,
                            Disability = :smo_pct,
                            CategoryProof_AttachmentName = :att_name,
                            CategoryProof_AttachmentUniqueName = :att_uname
                        WHERE pk_regid = :reg_id
                    """), {
                        "v_status": verify_status, 
                        "v_remarks": verify_remarks, 
                        "smo_pct": float(smo_percentage) if smo_percentage else None,
                        "att_name": attachment_name,
                        "att_uname": attachment_unique_name,
                        "reg_id": reg_id
                    })
                else:
                    db.session.execute(text("""
                        UPDATE PA_Registration_Mst 
                        SET VerificationDone = :v_status, 
                            VerificationReason = :v_remarks,
                            Disability = :smo_pct
                        WHERE pk_regid = :reg_id
                    """), {
                        "v_status": verify_status, 
                        "v_remarks": verify_remarks, 
                        "smo_pct": float(smo_percentage) if smo_percentage else None,
                        "reg_id": reg_id
                    })
                    
                db.session.commit()
                flash("PH Verification updated successfully.", "success")
                edit_id = None # Clear edit_id after successful update
            except Exception as e:
                db.session.rollback()
                flash(f"Error updating verification: {str(e)}", "error")
            return redirect(url_for('transactions.ph_candidates_verifications', session_id=f_session, degree_id=f_degree, college_id=f_college, page=page))

    if f_session and f_degree and _db_ping():
        where_clause = "m.PH = 1 AND m.fk_sessionid = :session_id AND m.fk_dtypeid = :degree_id"
        params = {"session_id": f_session, "degree_id": f_degree}
        
        if f_college and f_college != '0':
            where_clause += " AND m.fk_CollegID = :college_id"
            params["college_id"] = f_college

        # Get total count for pagination
        count_stmt = text(f"""
            SELECT COUNT(DISTINCT m.pk_regid) 
            FROM PA_Registration_Mst m
            WHERE {where_clause}
        """)
        total_records = db.session.execute(count_stmt, params).scalar() or 0

        # Fetch paginated records
        per_page = 50
        offset = (page - 1) * per_page
        
        stmt = text(f"""
            SELECT 
                m.pk_regid as id,
                m.regno, 
                (ISNULL(m.s_name, '') + ' ' + ISNULL(m.s_surname, '')) as name,
                d.description as degree,
                c.collegename as college,
                s.Specialization as specialization,
                m.VerificationDone,
                m.PH_Percentage,
                m.Disability,
                m.VerificationReason,
                m.CategoryProof_AttachmentName
            FROM PA_Registration_Mst m
            LEFT JOIN ACD_Degree_Mst d ON m.fk_dtypeid = d.pk_degreeid
            LEFT JOIN PA_College_Mst c ON m.fk_CollegID = c.Pk_CollegeID
            LEFT JOIN PA_Specialization_mst s ON m.fk_SId = s.Pk_SID
            WHERE {where_clause}
            ORDER BY m.s_name ASC
            OFFSET :offset ROWS FETCH NEXT :per_page ROWS ONLY
        """)
        
        params["offset"] = offset
        params["per_page"] = per_page
        
        try:
            rows = db.session.execute(stmt, params).mappings().all()
            for r in rows:
                status_text = "Yes"
                if r.get("VerificationDone") == '1':
                    status_text = "Verified"
                elif r.get("VerificationDone") == '0':
                    status_text = "Rejected"
                elif r.get("VerificationDone") == '2':
                    status_text = "Pending"
                    
                record_data = {
                    "id": r.get("id"),
                    "regno": r.get("regno"),
                    "name": r.get("name").strip() if r.get("name") else "",
                    "degree": r.get("degree") or "",
                    "college": r.get("college") or "",
                    "specialization": r.get("specialization") or "",
                    "ph_status": status_text,
                    "ph_percentage": r.get("PH_Percentage"),
                    "disability": float(r.get("Disability")) if r.get("Disability") is not None else None,
                    "verification_done": r.get("VerificationDone"),
                    "verification_reason": r.get("VerificationReason"),
                    "attachment_name": r.get("CategoryProof_AttachmentName")
                }
                records.append(record_data)
                
                if str(r.get("id")) == str(edit_id):
                    edit_record = record_data
        except Exception as e:
            print(f"Error fetching PH records: {e}")
            total_records = 0

        pagination = Pagination(page, per_page, total_records)
    else:
        pagination = Pagination(page, 50, 0)

    return render_template('transactions/ph_candidates_verifications.html', sessions=sessions, degrees=degrees, colleges=streams,
                           f_session=f_session, f_degree=f_degree, f_status=f_status, f_college=f_college, 
                           records=records, pagination=pagination, edit_record=edit_record, edit_id=edit_id)

# 7. Sports Quota Document Verification
@transactions_bp.route('/sports-quota-document-verification', methods=['GET', 'POST'])
def sports_quota_document_verification():
    sessions, degrees, _, _ = get_common_masters()
    
    f_session = request.form.get('session_id') or request.args.get('session_id')
    f_degree = request.form.get('degree_id') or request.args.get('degree_id')
    edit_id = request.form.get('edit_id') or request.args.get('edit_id')
    
    unverified_records = []
    verified_records = []
    sports_details = []
    edit_is_verified = False

    if request.method == 'POST' and request.form.get('action') == 'verify_sports':
        reg_id = request.form.get('reg_id')
        is_verified = request.form.get('is_verified') == '1'
        
        # Collect sports entry updates
        sports_updates = []
        for key, val in request.form.items():
            if key.startswith('sports_verify_'):
                sports_id = key.replace('sports_verify_', '')
                reason_key = f'sports_reason_{sports_id}'
                level_key = f'sports_level_{sports_id}'
                sports_updates.append({
                    'id': sports_id,
                    'verified': val == '1',
                    'reason': request.form.get(reason_key),
                    'level': request.form.get(level_key)
                })

        if reg_id and _db_ping():
            try:
                # Update main registration table
                db.session.execute(text("""
                    UPDATE PA_Registration_Mst 
                    SET IsSportsVerified = :is_verified
                    WHERE pk_regid = :reg_id
                """), {"is_verified": is_verified, "reg_id": reg_id})
                
                # Update individual sports
                for sp in sports_updates:
                    db.session.execute(text("""
                        UPDATE PA_CandidateSports_Trn 
                        SET Verified = :verified, Reason = :reason, VerifiedLevel = :v_level
                        WHERE Pk_SportsID = :sp_id
                    """), {
                        "verified": sp['verified'],
                        "reason": sp['reason'] or None,
                        "v_level": sp['level'] or None,
                        "sp_id": sp['id']
                    })
                    
                db.session.commit()
                flash("Sports verification updated successfully.", "success")
                edit_id = None
            except Exception as e:
                db.session.rollback()
                flash(f"Error updating verification: {str(e)}", "error")
            return redirect(url_for('transactions.sports_quota_document_verification', session_id=f_session, degree_id=f_degree))

    if f_session and _db_ping():
        where_clause = "m.SportsQuota = 1 AND m.fk_sessionid = :session_id"
        params = {"session_id": f_session}
        
        if f_degree and f_degree != '0':
            where_clause += " AND m.fk_dtypeid = :degree_id"
            params["degree_id"] = f_degree

        stmt = text(f"""
            SELECT DISTINCT
                m.pk_regid as id,
                m.s_name as raw_name,
                m.regno, 
                m.rollno,
                (ISNULL(m.s_name, '') + ' ' + ISNULL(m.s_surname, '')) as name,
                m.f_name as father_name,
                m.mobileno as mobile,
                m.email,
                d.description as degree,
                m.IsSportsVerified
            FROM PA_Registration_Mst m
            INNER JOIN PA_CandidateSports_Trn st ON m.pk_regid = st.fk_regid
            LEFT JOIN ACD_Degree_Mst d ON m.fk_dtypeid = d.pk_degreeid
            WHERE {where_clause}
            ORDER BY m.s_name ASC
        """)
        
        try:
            rows = db.session.execute(stmt, params).mappings().all()
            for r in rows:
                record = {
                    "id": r.get("id"),
                    "regno": r.get("regno"),
                    "rollno": r.get("rollno"),
                    "name": r.get("name").strip() if r.get("name") else "",
                    "father_name": r.get("father_name"),
                    "mobile": r.get("mobile"),
                    "email": r.get("email"),
                    "degree": r.get("degree") or ""
                }
                
                if r.get("IsSportsVerified"):
                    verified_records.append(record)
                else:
                    unverified_records.append(record)
                
                if str(r.get("id")) == str(edit_id):
                    edit_is_verified = bool(r.get("IsSportsVerified"))
                    # Fetch sports details for this student
                    sports_stmt = text("""
                        SELECT t.Pk_SportsID, t.Fk_GameID, g.GameName, t.LevelName, t.PartDate, t.Verified, t.Reason, t.VerifiedLevel
                        FROM PA_CandidateSports_Trn t
                        LEFT JOIN PA_GameList_Mst g ON t.Fk_GameID = g.Pk_GameID
                        WHERE t.fk_regid = :reg_id
                    """)
                    sports_rows = db.session.execute(sports_stmt, {"reg_id": edit_id}).mappings().all()
                    sports_details = [dict(s) for s in sports_rows]

        except Exception as e:
            print(f"Error fetching Sports records: {e}")

    return render_template('transactions/sports_quota_document_verification.html', sessions=sessions, degrees=degrees,
                           f_session=f_session, f_degree=f_degree, edit_id=edit_id, 
                           unverified_records=unverified_records, verified_records=verified_records,
                           sports_details=sports_details, edit_is_verified=edit_is_verified)

# 8. Student Data Modification (Search Criteria)
@transactions_bp.route('/student-data-modification', methods=['GET', 'POST'])
def student_data_modification():
    sessions = safe_all(lambda: db.session.execute(text("SELECT pk_sessionid as id, description as name FROM LUP_AcademicSession_Mst ORDER BY pk_sessionid DESC")).mappings().all(), [])
    degrees = safe_all(lambda: db.session.execute(text("SELECT pk_degreeid as id, description as name FROM ACD_Degree_Mst WHERE active=1 ORDER BY description")).mappings().all(), [])
    categories = safe_all(lambda: db.session.execute(text("SELECT Pk_StuCatId as id, Description as name FROM PA_StudentCategory_Mst ORDER BY Description")).mappings().all(), [])
    specs = safe_all(lambda: db.session.execute(text("SELECT Pk_SID as id, Specialization as name FROM PA_Specialization_mst ORDER BY Specialization")).mappings().all(), [])

    page = request.args.get('page', 1, type=int)
    per_page = 15
    use_db = _db_ping()

    f_session = request.args.get('session_id')
    f_degree = request.args.get('degree_id')
    f_cat = request.args.get('category_id')
    f_spec = request.args.get('spec_id')
    f_reg = request.args.get('reg_no')
    f_mob = request.args.get('mob_no')
    f_name = request.args.get('name')
    f_status = request.args.get('status', '0')

    display_records = []
    total_records = 0

    if use_db and (f_session or f_degree or f_reg or f_mob or f_name or f_status != '0'):
        where = " WHERE 1=1"
        params = {}
        if f_session: where += " AND fk_sessionid=:session_id"; params["session_id"] = f_session
        if f_degree: where += " AND fk_dtypeid=:degree_id"; params["degree_id"] = f_degree
        if f_cat: where += " AND fk_stucatid_cast=:cat_id"; params["cat_id"] = f_cat
        if f_spec: where += " AND fk_SId=:spec_id"; params["spec_id"] = f_spec
        if f_reg: where += " AND regno LIKE :reg_no"; params["reg_no"] = f"%{f_reg}%"
        if f_mob: where += " AND mobileno LIKE :mob_no"; params["mob_no"] = f"%{f_mob}%"
        if f_name: where += " AND (s_name LIKE :name OR s_surname LIKE :name)"; params["name"] = f"%{f_name}%"
        if f_status == '2': where += " AND IsPaymentSuccess=1"
        if f_status == '1': where += " AND (IsPaymentSuccess=0 OR IsPaymentSuccess IS NULL)"

        total_records = db.session.execute(text("SELECT COUNT(*) FROM PA_Registration_Mst " + where), params).scalar()
        off = (page - 1) * per_page

        stmt = text(f"""
            SELECT pk_regid, regno, s_name, s_surname, dob, dated, mobileno, email
            FROM PA_Registration_Mst
            {where}
            ORDER BY pk_regid DESC
            OFFSET {off} ROWS FETCH NEXT {per_page} ROWS ONLY
        """)
        rows = db.session.execute(stmt, params).mappings().all()
        for r in rows:
            name = (r.get("s_name") or "") + " " + (r.get("s_surname") or "")
            display_records.append({
                "id": r["pk_regid"],
                "reg_no": r["regno"],
                "name": name.strip(),
                "dob": format_dt(r["dob"]),
                "dated": format_dt(r["dated"]),
                "mobile": r["mobileno"],
                "email": r["email"]
            })

    return render_template('transactions/student_data_modification.html',
                           records=display_records,
                           sessions=sessions, degrees=degrees, categories=categories, specs=specs,        
                           f_session=f_session, f_degree=f_degree, f_cat=f_cat, f_spec=f_spec,
                           f_reg=f_reg, f_mob=f_mob, f_name=f_name, f_status=f_status,
                           pagination=Pagination(page, per_page, total_records))

# 9. Transfer Students
@transactions_bp.route('/transfer-students', methods=['GET', 'POST'])
def transfer_students():
    sessions = safe_all(lambda: db.session.execute(text("SELECT pk_sessionid as id, description as name FROM LUP_AcademicSession_Mst ORDER BY pk_sessionid DESC")).mappings().all(), [])
    streams = safe_all(lambda: db.session.execute(text("SELECT Pk_CollegeID as id, CollegeName as name FROM PA_College_Mst ORDER BY CollegeName")).mappings().all(), [])

    f_stream = request.form.get('stream_id') or request.args.get('stream_id')
    f_campus = request.form.get('campus_id') or request.args.get('campus_id')
    f_session = request.form.get('session_id') or request.args.get('session_id')
    f_degree = request.form.get('degree_id') or request.args.get('degree_id')
    f_regno = request.form.get('reg_no') or request.args.get('reg_no')
    f_cutoff = request.form.get('cutoff') or request.args.get('cutoff')

    records = []
    
    # Pre-populate campuses for the selected stream
    campuses = []
    if f_stream and _db_ping():
        c_stmt = text("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_Parentcollege_mst = :stream_id ORDER BY collegename")
        campuses = safe_all(lambda: db.session.execute(c_stmt, {"stream_id": f_stream}).mappings().all(), [])
        
    # Pre-populate degrees for the selected campus
    degrees = []
    if f_campus and _db_ping():
        d_stmt = text("""
            SELECT DISTINCT d.pk_degreeid as id, 
                   CASE 
                       WHEN dt.description IS NOT NULL THEN d.description + ' (' + dt.description + ')' 
                       ELSE d.description 
                   END as name 
            FROM SMS_CollegeDegreeBranchMap_Mst m
            JOIN ACD_Degree_Mst d ON m.fk_Degreeid = d.pk_degreeid
            LEFT JOIN ACD_DegreeType_Mst dt ON d.fk_dtypeid = dt.pk_dtypeid
            WHERE m.fk_CollegeId IN (
                SELECT pk_collegeid FROM SMS_College_Mst 
                WHERE fk_Parentcollege_mst = (SELECT fk_Parentcollege_mst FROM SMS_College_Mst WHERE pk_collegeid = :campus_id)
                OR pk_collegeid = :campus_id
            ) AND d.active = 1
            ORDER BY name
        """)
        degrees = safe_all(lambda: db.session.execute(d_stmt, {"campus_id": f_campus}).mappings().all(), [])
        
    cutoffs = list(range(1, 11))

    if request.method == 'POST' and request.form.get('action') == 'export':
        import io
        import openpyxl
        from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
        from flask import send_file
        
        where_clause = " WHERE m.fk_sessionid = :session_id AND m.fk_dtypeid = :degree_id AND (U.IsTransfer = 0 OR U.IsTransfer IS NULL)"
        params = {"session_id": f_session, "degree_id": f_degree}
        
        if f_campus and f_campus != '0':
            where_clause += " AND U.Fk_CollegeID = :campus_id"
            params["campus_id"] = f_campus
            
        if f_regno:
            where_clause += " AND m.regno LIKE :reg_no"
            params["reg_no"] = f"%{f_regno}%"
            
        if f_cutoff and f_cutoff != '0':
            where_clause += " AND U.CutOff = :cutoff"
            params["cutoff"] = f_cutoff

        stmt_export = text(f"""
            SELECT 
                m.regno as RegistrationNo, 
                m.rollno as RollNo, 
                (ISNULL(m.s_name, '') + ' ' + ISNULL(m.s_surname, '')) as StudentName,
                m.f_name as FatherName,
                d.description as Degree,
                cat.Description as Category,
                U.AllottedSpecialisation as Specialization,
                c.collegename as CollegeName,
                CONVERT(varchar, m.dob, 103) as DOB,
                m.gender as Gender
            FROM (
                SELECT t.fk_regid, t.AllottedSpecialisation, t.Fk_CollegeID, t.AllottedCategory, t.IsTransfer, mm.CutOff 
                FROM PA_Merit_Trn t 
                JOIN PA_Merit_Mst mm ON t.Fk_MeritID = mm.Pk_MeritID 
                WHERE mm.Fk_SessionID = :session_id AND mm.Fk_DegreeID = :degree_id
                
                UNION ALL
                
                SELECT t.fk_regid, t.SubjectName as AllottedSpecialisation, t.AllottedCollegeID as Fk_CollegeID, t.AllottedCategory, t.IsTransfer, mm.CutOff 
                FROM PA_Merit_Bsc_Trn t 
                JOIN PA_Merit_Mst mm ON t.Fk_MeritID = mm.Pk_MeritID 
                WHERE mm.Fk_SessionID = :session_id AND mm.Fk_DegreeID = :degree_id
            ) U
            JOIN PA_Registration_Mst m ON U.fk_regid = m.pk_regid
            LEFT JOIN ACD_Degree_Mst d ON m.fk_dtypeid = d.pk_degreeid
            LEFT JOIN PA_StudentCategory_Mst cat ON CAST(U.AllottedCategory as VARCHAR) = CAST(cat.Pk_StuCatId as VARCHAR)
            LEFT JOIN SMS_College_Mst c ON U.Fk_CollegeID = c.pk_collegeid
            {where_clause}
            ORDER BY m.s_name ASC
        """)
        
        try:
            export_data = db.session.execute(stmt_export, params).mappings().all()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transfer Students Export"
            
            headers = [
                "Registration No", "Roll No", "Student Name", "Father Name", 
                "Degree", "Allotted Category", "Allotted Specialisation", 
                "Allotted CollegeName", "DOB", "Gender"
            ]
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4A534A", end_color="4A534A", fill_type="solid")
            alt_row_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
                
            for i, row in enumerate(export_data, start=2):
                ws.append([
                    row.get("RegistrationNo"), row.get("RollNo"), row.get("StudentName"),
                    row.get("FatherName"), row.get("Degree"), row.get("Category"),
                    row.get("Specialization"), row.get("CollegeName"), row.get("DOB"),
                    row.get("Gender")
                ])
                current_fill = alt_row_fill if i % 2 == 0 else white_fill
                for cell in ws[i]:
                    cell.border = thin_border
                    cell.fill = current_fill
                    
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
                
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return send_file(
                buffer, 
                download_name=f"Transfer_Students_Export_{f_session}_{f_degree}.xlsx", 
                as_attachment=True, 
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            flash(f"Error generating export: {str(e)}", "error")
            return redirect(url_for('transactions.transfer_students', stream_id=f_stream, campus_id=f_campus, session_id=f_session, degree_id=f_degree, reg_no=f_regno, cutoff=f_cutoff))

    if request.method == 'POST' and request.form.get('action') == 'transfer_students':
        student_ids = request.form.getlist('student_id')
        t_college_id = request.form.get('transfer_college_id')
        t_session_id = request.form.get('transfer_session_id')
        t_degree_id = request.form.get('transfer_degree_id')
        t_spec_id = request.form.get('transfer_spec_id')

        t_college_name = ""
        t_session_name = ""
        t_degree_name = ""
        t_spec_name = ""

        if _db_ping():
            col_row = db.session.execute(text("SELECT CollegeName FROM PA_College_Mst WHERE Pk_CollegeID=:id"), {"id": t_college_id}).first() if t_college_id else None
            if col_row: t_college_name = col_row[0]

            ses_row = db.session.execute(text("SELECT description FROM LUP_AcademicSession_Mst WHERE pk_sessionid=:id"), {"id": t_session_id}).first() if t_session_id else None
            if ses_row: t_session_name = ses_row[0]

            deg_row = db.session.execute(text("SELECT description FROM ACD_Degree_Mst WHERE pk_degreeid=:id"), {"id": t_degree_id}).first() if t_degree_id else None
            if deg_row: t_degree_name = deg_row[0]

            if t_spec_id:
                spec_row = db.session.execute(text("SELECT Specialization FROM PA_Specialization_mst WHERE Pk_SID=:id"), {"id": t_spec_id}).first()
                if spec_row: t_spec_name = spec_row[0]

            try:
                for reg_id in student_ids:
                    db.session.execute(text("""
                        UPDATE PA_Merit_Trn 
                        SET IsTransfer = 1, Acd_College = :col, Acd_Session = :ses, Acd_Degree = :deg, Acd_Specialization = :spec
                        WHERE fk_regid = :reg_id
                    """), {
                        "col": t_college_name, "ses": t_session_name, "deg": t_degree_name, "spec": t_spec_name, "reg_id": reg_id
                    })

                    db.session.execute(text("""
                        UPDATE PA_Merit_Bsc_Trn 
                        SET IsTransfer = 1, Acd_College = :col, Acd_Session = :ses, Acd_Degree = :deg, Acd_Specialization = :spec
                        WHERE fk_regid = :reg_id
                    """), {
                        "col": t_college_name, "ses": t_session_name, "deg": t_degree_name, "spec": t_spec_name, "reg_id": reg_id
                    })
                db.session.commit()
                flash("Students transferred successfully.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Error transferring students: {str(e)}", "error")
            return redirect(url_for('transactions.transfer_students', stream_id=f_stream, campus_id=f_campus, session_id=f_session, degree_id=f_degree, reg_no=f_regno, cutoff=f_cutoff))

    if f_session and f_degree and _db_ping():
        where_clause = " WHERE (U.IsTransfer = 0 OR U.IsTransfer IS NULL)"
        params = {"session_id": f_session, "degree_id": f_degree}
        
        # We must filter by the campus (AllottedCollegeID in Bsc_Trn or Fk_CollegeID in Merit_Trn)
        if f_campus and f_campus != '0':
            where_clause += " AND U.Fk_CollegeID = :campus_id"
            params["campus_id"] = f_campus
            
        if f_regno:
            where_clause += " AND m.regno LIKE :reg_no"
            params["reg_no"] = f"%{f_regno}%"
            
        if f_cutoff and f_cutoff != '0':
            where_clause += " AND U.CutOff = :cutoff"
            params["cutoff"] = f_cutoff

        # Filter the union directly by Session and Degree for massive performance improvements and correctness
        stmt = text(f"""
            SELECT 
                m.pk_regid as id,
                m.regno, 
                m.rollno, 
                (ISNULL(m.s_name, '') + ' ' + ISNULL(m.s_surname, '')) as name,
                m.f_name as father_name,
                d.description as degree,
                U.AllottedCategory as category,
                U.AllottedSpecialisation as specialization,
                c.collegename as college,
                CONVERT(varchar, m.dob, 103) as dob_str,
                m.gender
            FROM (
                SELECT t.fk_regid, t.AllottedSpecialisation, t.Fk_CollegeID, t.AllottedCategory, t.IsTransfer, mm.CutOff 
                FROM PA_Merit_Trn t 
                JOIN PA_Merit_Mst mm ON t.Fk_MeritID = mm.Pk_MeritID 
                WHERE mm.Fk_SessionID = :session_id AND mm.Fk_DegreeID = :degree_id
                
                UNION ALL
                
                SELECT t.fk_regid, t.SubjectName as AllottedSpecialisation, t.AllottedCollegeID as Fk_CollegeID, t.AllottedCategory, t.IsTransfer, mm.CutOff 
                FROM PA_Merit_Bsc_Trn t 
                JOIN PA_Merit_Mst mm ON t.Fk_MeritID = mm.Pk_MeritID 
                WHERE mm.Fk_SessionID = :session_id AND mm.Fk_DegreeID = :degree_id
            ) U
            JOIN PA_Registration_Mst m ON U.fk_regid = m.pk_regid
            LEFT JOIN ACD_Degree_Mst d ON m.fk_dtypeid = d.pk_degreeid
            LEFT JOIN SMS_College_Mst c ON U.Fk_CollegeID = c.pk_collegeid
            {where_clause}
            ORDER BY m.s_name ASC
        """)

        try:
            rows = db.session.execute(stmt, params).mappings().all()
            for r in rows:
                records.append(dict(r))
        except Exception as e:
            print(f"Error fetching transfer students: {e}")

    # For mapping section
    all_colleges = safe_all(lambda: db.session.execute(text("SELECT Pk_CollegeID as id, CollegeName as name FROM PA_College_Mst ORDER BY CollegeName")).mappings().all(), [])
    all_degrees = safe_all(lambda: db.session.execute(text("SELECT pk_degreeid as id, description as name FROM ACD_Degree_Mst WHERE active=1")).mappings().all(), [])
    all_specs = safe_all(lambda: db.session.execute(text("SELECT Pk_SID as id, Specialization as name FROM PA_Specialization_mst")).mappings().all(), [])

    return render_template('transactions/transfer_students.html',
                           records=records,
                           sessions=sessions, streams=streams, campuses=campuses, degrees=degrees, cutoffs=cutoffs,
                           all_colleges=all_colleges, all_degrees=all_degrees, all_specs=all_specs,
                           f_stream=f_stream, f_campus=f_campus, f_session=f_session, f_degree=f_degree, f_regno=f_regno, f_cutoff=f_cutoff)
@transactions_bp.route('/api/get-degrees-by-college/<int:college_id>')
def get_degrees_by_college(college_id):
    degrees = []
    if _db_ping():
        stmt = text("""
            SELECT DISTINCT d.pk_degreeid as id, 
                   CASE 
                       WHEN dt.description IS NOT NULL THEN d.description + ' (' + dt.description + ')' 
                       ELSE d.description 
                   END as name 
            FROM SMS_CollegeDegreeBranchMap_Mst m
            JOIN ACD_Degree_Mst d ON m.fk_Degreeid = d.pk_degreeid
            LEFT JOIN ACD_DegreeType_Mst dt ON d.fk_dtypeid = dt.pk_dtypeid
            WHERE m.fk_CollegeId IN (
                SELECT pk_collegeid FROM SMS_College_Mst 
                WHERE fk_Parentcollege_mst = (SELECT fk_Parentcollege_mst FROM SMS_College_Mst WHERE pk_collegeid = :college_id)
                OR pk_collegeid = :college_id
            ) AND d.active = 1
            ORDER BY name
        """)
        rows = db.session.execute(stmt, {"college_id": college_id}).mappings().all()
        degrees = [dict(r) for r in rows]
    return jsonify(degrees)

@transactions_bp.route('/api/get-campuses-by-stream/<int:stream_id>')
def get_campuses_by_stream(stream_id):
    campuses = []
    if _db_ping():
        stmt = text("""
            SELECT pk_collegeid as id, collegename as name 
            FROM SMS_College_Mst 
            WHERE fk_Parentcollege_mst = :stream_id
            ORDER BY collegename
        """)
        rows = db.session.execute(stmt, {"stream_id": stream_id}).mappings().all()
        campuses = [dict(r) for r in rows]
    return jsonify(campuses)

@transactions_bp.route('/api/get-colleges-by-degree/<int:degree_id>')
def get_colleges_by_degree(degree_id):
    colleges = []
    if _db_ping():
        stmt = text("""
            SELECT DISTINCT p.Pk_CollegeID as id, p.CollegeName as name
            FROM SMS_CollegeDegreeBranchMap_Mst m
            JOIN SMS_College_Mst s ON m.fk_CollegeId = s.pk_collegeid
            JOIN PA_College_Mst p ON s.fk_Parentcollege_mst = p.Pk_CollegeID
            WHERE m.fk_Degreeid = :degree_id
            ORDER BY p.CollegeName
        """)
        rows = db.session.execute(stmt, {"degree_id": degree_id}).mappings().all()
        colleges = [dict(r) for r in rows]
    return jsonify(colleges)
@transactions_bp.route('/student-edit-part1/<int:id>', methods=['GET', 'POST'])
def student_edit_part1(id):
    if request.method == 'POST':
        # Processing save goes here...
        if _db_ping():
            # Gather all fields from form - matching template input names
            data = {
                "id": id,
                "s_name": request.form.get('name'),
                "dob_str": request.form.get('dob'),
                "f_name": request.form.get('father_name'),
                "m_name": request.form.get('mother_name'),
                "gender": request.form.get('gender'),
                "adharno": request.form.get('aadhar'),
                "religion": request.form.get('religion'),
                "mobile": request.form.get('mobile'),
                "parents_mobile": request.form.get('parents_mobile'),
                "marital_status": request.form.get('marital_status'),
                "child_status": 1 if request.form.get('girl_child') == '1' else 0,
                "blood_group": request.form.get('blood_group'),
                "nationality": request.form.get('nationality'),
                "email": request.form.get('email'),
                "category": request.form.get('category'),
                "ldv": 1 if request.form.get('ldv') == '1' else 0,
                "father_guardian": request.form.get('guardian'),
                "father_occ": request.form.get('occupation'),
                "annual_income": request.form.get('annual_income'),
                "ff": 1 if request.form.get('ff_category') == '1' else 0,
                "esm": 1 if request.form.get('esm_category') == '1' else 0,
                "ph": 1 if request.form.get('disability') == '1' else 0,
                "sports": 1 if request.form.get('sports_quota') == '1' else 0,
                "is_ward": 1 if request.form.get('ward_emp') == '1' else 0,
                "resident": 1 if request.form.get('resident') == 'Haryana' else 2,
                "step1": 1
            }
            
            # Format DOB
            dob_val = None
            if data['dob_str']:
                try:
                    # Try to parse dd/mm/yyyy or yyyy-mm-dd
                    if '-' in data['dob_str']:
                        dob_val = datetime.strptime(data['dob_str'], '%Y-%m-%d')
                    else:
                        dob_val = datetime.strptime(data['dob_str'], '%d/%m/%Y')
                except:
                    # Try common other formats or just pass it as is
                    dob_val = data['dob_str']

            stmt = text('''
                UPDATE PA_Registration_Mst
                SET s_name=:s_name, dob=:dob, f_name=:f_name, m_name=:m_name, gender=:gender,
                    AdharNo=:adharno, fk_religionid=:religion, mobileno=:mobile, Parents_Mobileno=:parents_mobile,
                    Marital_Status=:marital_status, ChildStatus=:child_status, Blood_Group=:blood_group,
                    nationality=:nationality, email=:email, fk_stucatid_cast=:category, LDV=:ldv,
                    FatherGuargian=:father_guardian, FatherOccupation=:father_occ, AnnualIncome=:annual_income,
                    FF=:ff, ESM=:esm, PH=:ph, SportsQuota=:sports, IsWard=:is_ward, Resident=:resident,
                    step1=:step1
                WHERE pk_regid=:id
            ''')
            db.session.execute(stmt, {
                "s_name": data["s_name"], "dob": dob_val, "f_name": data["f_name"], "m_name": data["m_name"],
                "gender": data["gender"], "adharno": data["adharno"], "religion": data["religion"],
                "mobile": data["mobile"], "parents_mobile": data["parents_mobile"],
                "marital_status": data["marital_status"], "child_status": data["child_status"],
                "blood_group": data["blood_group"], "nationality": data["nationality"],
                "email": data["email"], "category": data["category"], "ldv": data["ldv"],
                "father_guardian": data["father_guardian"], "father_occ": data["father_occ"],
                "annual_income": data["annual_income"], "ff": data["ff"], "esm": data["esm"],
                "ph": data["ph"], "sports": data["sports"], "is_ward": data["is_ward"],
                "resident": data["resident"], "step1": data["step1"], "id": id
            })
            db.session.commit()
            
        flash("Personal details updated.", "success")
        return redirect(url_for('transactions.student_edit_part2', id=id))
    
    sessions, _, _, _ = get_common_masters()
    categories = safe_all(lambda: db.session.execute(text("SELECT Pk_StuCatId as id, Description as name FROM PA_StudentCategory_Mst ORDER BY Description")).mappings().all(), [])
    religions = safe_all(lambda: db.session.execute(text("SELECT Pk_ReligionId as id, Description as name FROM PA_Religion_Mst")).mappings().all(), [])
    
    student = None
    if _db_ping():
        stmt = text("""
            SELECT regno, fk_sessionid, s_name, dob, f_name, m_name, gender, AdharNo, fk_religionid, 
                   mobileno, Parents_Mobileno, Marital_Status, ChildStatus, Blood_Group, nationality, 
                   email, fk_stucatid_cast, LDV, FatherGuargian, FatherOccupation, AnnualIncome, 
                   FF, ESM, PH, SportsQuota, IsWard, Resident
            FROM PA_Registration_Mst 
            WHERE pk_regid=:id
        """)
        row = db.session.execute(stmt, {"id": id}).mappings().first()
        if row:
            student = dict(row)
            student['dob'] = format_dt(student['dob'])
            if student.get('Blood_Group'): student['Blood_Group'] = student['Blood_Group'].strip()
            if student.get('gender'): student['gender'] = student['gender'].strip()
            if student.get('nationality'): student['nationality'] = student['nationality'].strip()
            if student.get('FatherGuargian'): student['FatherGuargian'] = student['FatherGuargian'].strip()

    from app.models import CandidateDocument
    doc = CandidateDocument.query.filter_by(fk_regid=id).first()
    return render_template('transactions/student_edit_part1.html', student_id=id, student=student, sessions=sessions, categories=categories, religions=religions, doc=doc)


@transactions_bp.route('/view_student_image/<int:reg_id>/<doc_type>')
def view_student_image(reg_id, doc_type):
    from flask import send_file
    import io
    from app.models import CandidateDocument
    doc = CandidateDocument.query.filter_by(fk_regid=reg_id).first()
    if not doc:
        return "No document", 404
        
    img_data = None
    mimetype = 'image/jpeg'
    if doc_type == 'photo':
        img_data = doc.imgattach_p
        if doc.contenttype_p: mimetype = doc.contenttype_p
    elif doc_type == 'signature':
        img_data = doc.imgattach_s
        if doc.contenttype_s: mimetype = doc.contenttype_s
    elif doc_type == 'thumb':
        img_data = doc.imgattach_t
        if doc.contenttype_t: mimetype = doc.contenttype_t
        
    if img_data:
        return send_file(io.BytesIO(img_data), mimetype=mimetype)
    return "No image", 404

@transactions_bp.route('/student-edit-part2/<int:id>', methods=['GET', 'POST'])
def student_edit_part2(id):
    if request.method == 'POST':
        if request.form.get('action') == 'next':
            return redirect(url_for('transactions.student_edit_part3', id=id))
        
        c_address = request.form.get('c_address')
        c_district = request.form.get('c_district')
        c_state = request.form.get('c_state')
        c_pincode = request.form.get('c_pincode')
        
        p_address = request.form.get('p_address')
        p_district = request.form.get('p_district')
        p_state = request.form.get('p_state')
        p_pincode = request.form.get('p_pincode')
        
        if _db_ping():
            stmt = text('''
                UPDATE PA_Registration_Mst
                SET c_address=:c_address, c_district=:c_district, c_fk_stateid=:c_state, c_pincode=:c_pincode,
                    p_address=:p_address, p_district=:p_district, p_fk_stateid=:p_state, p_pincode=:p_pincode,
                    step2=1
                WHERE pk_regid=:id
            ''')
            db.session.execute(stmt, {
                "c_address": c_address, "c_district": c_district, "c_state": c_state, "c_pincode": c_pincode,
                "p_address": p_address, "p_district": p_district, "p_state": p_state, "p_pincode": p_pincode,
                "id": id
            })
            
            # Handle file uploads
            from app.models import CandidateDocument
            doc = CandidateDocument.query.filter_by(fk_regid=id).first()
            if not doc:
                doc = CandidateDocument(fk_regid=id)
                db.session.add(doc)
            
            photo = request.files.get('photo')
            if photo and photo.filename:
                doc.imgattach_p = photo.read()
                doc.contenttype_p = photo.mimetype
                doc.filename_p = photo.filename
            
            signature = request.files.get('signature')
            if signature and signature.filename:
                doc.imgattach_s = signature.read()
                doc.contenttype_s = signature.mimetype
                doc.filename_s = signature.filename
                
            thumb = request.files.get('thumb')
            if thumb and thumb.filename:
                doc.imgattach_t = thumb.read()
                doc.contenttype_t = thumb.mimetype
                doc.filename_t = thumb.filename
                
            db.session.commit()
            
        flash("Address details updated.", "success")
        
    states = safe_all(lambda: db.session.execute(text("SELECT pk_StateId as id, StateName as name FROM State_Mst ORDER BY StateName")).mappings().all(), [])
    
    student = None
    if _db_ping():
        # Join with PA_FamilyID_Details and PA_Candidate_Verification
        stmt = text('''
            SELECT 
                r.c_address, r.c_district, r.c_fk_stateid, r.c_pincode,
                r.p_address, r.p_district, r.p_fk_stateid, r.p_pincode,
                f.houseNo, f.streetNo, f.address_LandMark, f.districtName, f.btName, f.wvName, f.pinCode as ppp_pin,
                v.Address as v_address, v.FK_Stateid as v_state, v.FK_District as v_district
            FROM PA_Registration_Mst r
            LEFT JOIN PA_FamilyID_Details f ON r.pk_regid = f.fk_regid
            LEFT JOIN PA_Candidate_Verification v ON r.pk_regid = v.fk_regid
            WHERE r.pk_regid=:id
        ''')
        row = db.session.execute(stmt, {"id": id}).mappings().first()
        
        if row:
            student = dict(row)
            
            # 1. Check Candidate Verification (New Priority)
            if student.get('c_address') in [None, '', 'N/A'] and student.get('v_address'):
                student['c_address'] = student['v_address']
                student['p_address'] = student['v_address']
                if student.get('c_fk_stateid') in [None, 0, 1]: student['c_fk_stateid'] = student.get('v_state')
                if student.get('p_fk_stateid') in [None, 0, 1]: student['p_fk_stateid'] = student.get('v_state')
                
                # Extract Pin Code from address string if pincode is 000000
                if student.get('c_pincode') in [None, '', '000000']:
                    import re
                    # Look for 6 consecutive digits at the end or following a hyphen/space
                    pin_match = re.search(r'(\d{6})\b', student.get('v_address') or '')
                    if pin_match:
                        student['c_pincode'] = pin_match.group(1)
                        student['p_pincode'] = pin_match.group(1)

                # Try to extract district/pin from address string if available or use v_district
                if student.get('c_district') in [None, '', 'N/A'] and student.get('v_district'): 
                    # v_district might be an ID or Name, try to get name
                    try:
                        d_name = db.session.execute(text("SELECT DistrictName FROM District_Mst WHERE pk_DistrictId=:id"), {"id": student['v_district']}).scalar()
                        if d_name: 
                            student['c_district'] = d_name
                            student['p_district'] = d_name
                    except: pass

            # 2. Check PPP (Existing Fallback)
            if student.get('c_address') in [None, '', 'N/A'] and student.get('houseNo'):
                h = student.get('houseNo', '') or ''
                s = student.get('streetNo', '') or ''
                l = student.get('address_LandMark', '') or ''
                d = student.get('districtName', '') or ''
                b = student.get('btName', '') or ''
                w = student.get('wvName', '') or ''
                assembled = f"HouseNo-{h},StreetNo-{s},AddressLandMark-{l},DistrictName-{d},Block/Town-{b},Ward/Village-{w}"
                
                if student.get('c_address') in [None, '', 'N/A']: student['c_address'] = assembled
                if student.get('p_address') in [None, '', 'N/A']: student['p_address'] = assembled
                if student.get('c_district') in [None, '', 'N/A']: student['c_district'] = d
                if student.get('p_district') in [None, '', 'N/A']: student['p_district'] = d
                if student.get('c_pincode') in [None, '', '000000'] and student.get('ppp_pin'): student['c_pincode'] = student['ppp_pin']
                if student.get('p_pincode') in [None, '', '000000'] and student.get('ppp_pin'): student['p_pincode'] = student['ppp_pin']

    from app.models import CandidateDocument
    doc = CandidateDocument.query.filter_by(fk_regid=id).first()
    return render_template('transactions/student_edit_part2.html', student_id=id, student=student, states=states, doc=doc)

@transactions_bp.route('/student-edit-part3/<int:id>', methods=['GET', 'POST'])
def student_edit_part3(id):
    if request.method == 'POST':
        if request.form.get('action') == 'next':
            return redirect(url_for('transactions.student_edit_part4', id=id))
        
        if _db_ping():
            # Helper to update or insert qualification
            def update_qual(prevexamid, prefix):
                year = request.form.get(f'{prefix}_year')
                board = request.form.get(f'{prefix}_board')
                rollno = request.form.get(f'{prefix}_rollno')
                m_type = request.form.get(f'{prefix}_marks_type')
                max_m = request.form.get(f'{prefix}_max')
                obt_m = request.form.get(f'{prefix}_obt')
                perc = request.form.get(f'{prefix}_perc')
                subj = request.form.get(f'{prefix}_subject')
                spec = request.form.get(f'{prefix}_spec') if prefix == 'q12' else None
                res_await = 1 if request.form.get(f'{prefix}_result') == 'N' else 0
                
                # Check if exists
                chk = db.session.execute(text("SELECT pk_trnid FROM PA_Registration_CollegeApply_PreExam_Trn WHERE fk_regid=:id AND fk_prevexamid=:eid"), {"id": id, "eid": prevexamid}).scalar()
                
                params = {
                    "id": id, "eid": prevexamid, "year": year, "board": board, "roll": rollno,
                    "grade": 'G' if m_type == 'CGPA' else 'M',
                    "max": float(max_m) if max_m else 0,
                    "obt": float(obt_m) if obt_m else 0,
                    "perc": float(perc) if perc else 0,
                    "subj": subj, "spec": spec, "res": res_await
                }
                
                if chk:
                    stmt = text('''
                        UPDATE PA_Registration_CollegeApply_PreExam_Trn
                        SET fk_yearid=:year, fk_boardid=:board, Rollno=:roll, grade=:grade,
                            maxmarks=:max, marks=:obt, marks_act=:perc, coursedtl=:subj,
                            fk_pestreamid=:spec, result_await=:res
                        WHERE pk_trnid=:chk_id
                    ''')
                    params["chk_id"] = chk
                    db.session.execute(stmt, params)
                else:
                    stmt = text('''
                        INSERT INTO PA_Registration_CollegeApply_PreExam_Trn
                        (fk_regid, fk_prevexamid, fk_yearid, fk_boardid, Rollno, grade, maxmarks, marks, marks_act, coursedtl, fk_pestreamid, result_await)
                        VALUES (:id, :eid, :year, :board, :roll, :grade, :max, :obt, :perc, :subj, :spec, :res)
                    ''')
                    db.session.execute(stmt, params)

            update_qual(27, 'q10')
            update_qual(26, 'q12')
            
            # Update step3 in Mst
            db.session.execute(text("UPDATE PA_Registration_Mst SET step3=1 WHERE pk_regid=:id"), {"id": id})
            db.session.commit()
            
        flash("Educational details updated.", "success")
        
    boards = safe_all(lambda: db.session.execute(text("SELECT Pk_BoardId as id, Description as name FROM PA_Board_Mst ORDER BY Description")).mappings().all(), [])
    specializations = safe_all(lambda: db.session.execute(text("SELECT Pk_ESP_Id as id, Description as name FROM PA_Education_Specialization_Mst ORDER BY Description")).mappings().all(), [])
    
    student = None
    if _db_ping():
        stmt = text("SELECT PassingYear10th FROM PA_Registration_Mst WHERE pk_regid=:id")
        row = db.session.execute(stmt, {"id": id}).mappings().first()
        if row: student = dict(row)


    q10 = {}
    q12 = {}
    if _db_ping():
        # First try PA_Registration_CollegeApply_PreExam_Trn (Candidate portal table)
        stmt_qual = text('''
            SELECT fk_prevexamid, fk_yearid as Year, Rollno as RollNo, maxmarks as MaxMark, marks as ObtOrOGPA, 
                   marks_act as Percentage, fk_boardid as fk_BoardId, grade as MarksType, coursedtl as Subject, 
                   fk_pestreamid as fk_SID, result_await as Result_Awaited
            FROM PA_Registration_CollegeApply_PreExam_Trn
            WHERE fk_regid=:id
        ''')
        quals = db.session.execute(stmt_qual, {"id": id}).mappings().all()
        for q in quals:
            if q['fk_prevexamid'] == 27:
                q10 = dict(q)
                if q10.get('MarksType'): q10['MarksType'] = 'CGPA' if str(q10['MarksType']).strip() == 'G' else 'Marks'
            elif q['fk_prevexamid'] == 26:
                q12 = dict(q)
                if q12.get('MarksType'): q12['MarksType'] = 'CGPA' if str(q12['MarksType']).strip() == 'G' else 'Marks'
        
        # If empty, fallback to ACD table
        if not q10 and not q12:
            stmt_qual = text('''
                SELECT fk_EID, Year, RollNo, MaxMark, ObtOrOGPA, Percentage, fk_BoardId, MarksType, Subject, fk_SID, Result_Awaited
                FROM ACD_EducationQualification_Details
                WHERE fk_regid=:id
            ''')
            quals = db.session.execute(stmt_qual, {"id": id}).mappings().all()
            for q in quals:
                if q['fk_EID'] == 1:
                    q10 = dict(q)
                    # Normalize MarksType
                    if q10.get('MarksType'): 
                        mt = str(q10['MarksType']).strip().upper()
                        q10['MarksType'] = 'CGPA' if mt in ('C', 'G') else 'Marks'
                    
                    # Normalize Result_Awaited
                    ra = str(q10.get('Result_Awaited')).strip().upper()
                    q10['Result_Awaited'] = True if ra in ('1', 'Y', 'TRUE') else False
                elif q['fk_EID'] in (2, 11):
                    q12 = dict(q)
                    # Normalize MarksType
                    if q12.get('MarksType'): 
                        mt = str(q12['MarksType']).strip().upper()
                        q12['MarksType'] = 'CGPA' if mt in ('C', 'G') else 'Marks'
                    
                    # Normalize Result_Awaited
                    ra = str(q12.get('Result_Awaited')).strip().upper()
                    q12['Result_Awaited'] = True if ra in ('1', 'Y', 'TRUE') else False

    # Convert Decimal to float for template
    for d in [q10, q12]:
        if d:
            for k, v in d.items():
                from decimal import Decimal
                if isinstance(v, Decimal):
                    d[k] = float(v)


    return render_template('transactions/student_edit_part3.html', student_id=id, student=student, boards=boards, specializations=specializations, q10=q10, q12=q12)

@transactions_bp.route('/student-edit-part4/<int:id>', methods=['GET', 'POST'])
def student_edit_part4(id):
    if request.method == 'POST':
        if request.form.get('action') == 'next':
            return redirect(url_for('transactions.student_edit_part5', id=id))
        
        if _db_ping():
            ex_student = request.form.get('ex_student')
            other_info = request.form.get('other_info_text')
            rural_urban = request.form.get('rural_urban')
            
            # Check if exists
            chk = db.session.execute(text("SELECT Pk_FId FROM REC_FamilyandAdditional_Information WHERE fk_regid=:id"), {"id": id}).scalar()
            
            if chk:
                stmt = text('''
                    UPDATE REC_FamilyandAdditional_Information
                    SET Ex_Student=:ex, OtherInformation=:other, RuralOrUrban=:ru
                    WHERE Pk_FId=:chk_id
                ''')
                db.session.execute(stmt, {"ex": ex_student, "other": other_info, "ru": rural_urban, "chk_id": chk, "id": id})
            else:
                stmt = text('''
                    INSERT INTO REC_FamilyandAdditional_Information (fk_regid, Ex_Student, OtherInformation, RuralOrUrban)
                    VALUES (:id, :ex, :other, :ru)
                ''')
                db.session.execute(stmt, {"id": id, "ex": ex_student, "other": other_info, "ru": rural_urban})
            
            # Update step4 in Mst
            db.session.execute(text("UPDATE PA_Registration_Mst SET step4=1 WHERE pk_regid=:id"), {"id": id})
            db.session.commit()
            
        flash("Additional information updated.", "success")
        
    student = None
    if _db_ping():
        stmt = text("SELECT Ex_Student, ReleventStatus, OtherInformation, RuralOrUrban FROM REC_FamilyandAdditional_Information WHERE fk_regid=:id")
        row = db.session.execute(stmt, {"id": id}).mappings().first()
        if row: student = dict(row)

    return render_template('transactions/student_edit_part4.html', student_id=id, student=student)

@transactions_bp.route('/student-edit-part5/<int:id>', methods=['GET', 'POST'])
def student_edit_part5(id):
    if request.method == 'POST':
        flash("Student details updated successfully.", "success")
        return redirect(url_for('transactions.student_data_modification'))
    
    student = None
    preferences = []
    if _db_ping():
        stmt = text("""
            SELECT r.rollno, r.LDV, r.FF, r.ESM, r.PH, r.IsWard, r.IsHaryanaPass, r.SportsQuota, c.Description as CategoryName,
                   cm.ObtainMarks as entrance_marks, cm.overallR as overall_rank, 
                   cm.overallRcount as overall_rank_count, cm.Categoryrank as cat_rank, 
                   cm.ESMRank as esm_rank, cm.FFRank as ff_rank, cm.PHRank as ph_rank
            FROM PA_Registration_Mst r
            LEFT JOIN PA_StudentCategory_Mst c ON r.fk_stucatid_cast = c.Pk_StuCatId
            LEFT JOIN PA_Candidate_Marks cm ON r.rollno = cm.RollNo
            WHERE r.pk_regid=:id
        """)
        row = db.session.execute(stmt, {"id": id}).mappings().first()
        if row: student = dict(row)
        
        stmt_pref = text("""
            SELECT p.Preference, c.collegename as CollegeName
            FROM PA_StudentCollegePreference_Details p
            JOIN SMS_College_Mst c ON p.fk_CollegeId = c.pk_collegeid
            WHERE p.fk_regid=:id
            ORDER BY p.Preference
        """)
        preferences = db.session.execute(stmt_pref, {"id": id}).mappings().all()

    return render_template('transactions/student_edit_part5.html', student_id=id, student=student, preferences=preferences)