from flask import Blueprint, render_template, request, flash, make_response, jsonify, redirect, url_for
from app import db
from app.models import (AcademicSession, Degree, DegreeType, CollegeCategory,
                         UniversitySpecialization, PA_ET_Master, PA_Exam_Center_Mst,
                         PA_Exam_Center_Trn, UM_Users_Mst)
from sqlalchemy import text
import json
import io
from collections import defaultdict
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                  Spacer, HRFlowable, PageBreak, Image)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

seat_allocation_bp = Blueprint('seat_allocation', __name__, url_prefix='/seat_allocation')

@seat_allocation_bp.before_request
def check_admin_auth():
    from flask import session
    if not session.get('user_id'):
        if request.endpoint and 'login' not in request.endpoint: return redirect(url_for('main.login'))

# ─────────────────────────────────────────────────────────────
#  HELPER: build ReportLab styles (FROM NEW CODE)
# ─────────────────────────────────────────────────────────────
UNIV_HINDI  = "चौधरी चरण सिंह हरियाणा कृषि विश्वविद्यालय"
UNIV_ENG    = "CCS Haryana Agricultural University"
UNIV_CITY   = "Hisar – 125 004  (Haryana)"
GRID_HDR_BG = colors.HexColor('#2e5090')   # dark-blue header
GRID_ALT_BG = colors.HexColor('#eef2fb')   # alternate row tint
BORDER_CLR  = colors.HexColor('#4a6690')

def _styles():
    s = getSampleStyleSheet()
    reg = ParagraphStyle
    
    font_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'NotoSansDevanagari-Regular.ttf')
    hindi_font = 'Helvetica-Bold'
    
    # Register Arial fonts
    arial_path = 'C:/Windows/Fonts/arial.ttf'
    arial_bd_path = 'C:/Windows/Fonts/arialbd.ttf'
    try:
        if os.path.exists(arial_path):
            pdfmetrics.registerFont(TTFont('Arial', arial_path))
        if os.path.exists(arial_bd_path):
            pdfmetrics.registerFont(TTFont('Arial-Bold', arial_bd_path))
    except Exception as e:
        print(f"Error registering Arial: {e}")

    try:
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('Hindi', font_path))
            hindi_font = 'Hindi'
    except Exception as e:
        print(f"Font registration error: {e}")
        
    return {
        'hindi'   : reg('hindi',   fontName=hindi_font,  fontSize=13, alignment=TA_CENTER, leading=16),
        'eng'     : reg('eng',     fontName='Arial-Bold',  fontSize=11, alignment=TA_CENTER, leading=14),
        'city'    : reg('city',    fontName='Arial',        fontSize=9,  alignment=TA_CENTER, leading=12),
        'title'   : reg('title',   fontName='Arial-Bold',  fontSize=12, alignment=TA_CENTER, spaceBefore=6, spaceAfter=4),
        'sub'     : reg('sub',     fontName='Arial-Bold',  fontSize=10, alignment=TA_LEFT,   spaceBefore=4, spaceAfter=2),
        'normal'  : reg('normal',  fontName='Arial',        fontSize=9,  alignment=TA_LEFT,   leading=11),
        'bold9'   : reg('bold9',   fontName='Arial-Bold',  fontSize=9,  alignment=TA_LEFT),
        'center9' : reg('center9', fontName='Arial',        fontSize=9,  alignment=TA_CENTER),
    }

def _univ_header(els, st, report_title=None):
    """Append university header paragraphs."""
    logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
    hindi_img_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'hindi_name.png')

    if os.path.exists(hindi_img_path):
        hindi_el = Image(hindi_img_path, width=9.5*cm, height=0.6*cm)
    else:
        hindi_el = Paragraph(UNIV_HINDI, st['hindi'])

    text_col = [
        hindi_el,
        Paragraph(UNIV_ENG,   st['eng']),
        Paragraph(UNIV_CITY,  st['city'])
    ]
    if report_title:
        text_col.append(Spacer(1, 4))
        text_col.append(Paragraph(report_title, st['title']))

    img = ''
    if os.path.exists(logo_path):
        img = Image(logo_path, width=2.0*cm, height=2.0*cm)

    header_table = Table([[img, text_col, '']], colWidths=[3.5*cm, 11*cm, 3.5*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('ALIGN', (2,0), (2,0), 'RIGHT')
    ]))
    els.append(header_table)

    els.append(Spacer(1, 4))
    els.append(HRFlowable(width="100%", thickness=1.0, color=colors.black, spaceAfter=8))

def _grid_style(hdr_rows=1):
    return TableStyle([
        ('FONTNAME',        (0, 0),        (-1, hdr_rows-1), 'Helvetica-Bold'),
        ('FONTSIZE',        (0, 0),        (-1, hdr_rows-1), 9),
        ('BACKGROUND',      (0, 0),        (-1, hdr_rows-1), GRID_HDR_BG),
        ('TEXTCOLOR',       (0, 0),        (-1, hdr_rows-1), colors.white),
        ('ALIGN',           (0, 0),        (-1, hdr_rows-1), 'CENTER'),
        ('VALIGN',          (0, 0),        (-1, -1),          'MIDDLE'),
        ('FONTNAME',        (0, hdr_rows), (-1, -1),          'Helvetica'),
        ('FONTSIZE',        (0, hdr_rows), (-1, -1),          8),
        ('ROWBACKGROUNDS',  (0, hdr_rows), (-1, -1),          [colors.white, GRID_ALT_BG]),
        ('GRID',            (0, 0),        (-1, -1),          0.4, BORDER_CLR),
        ('TOPPADDING',      (0, 0),        (-1, -1),          3),
        ('BOTTOMPADDING',   (0, 0),        (-1, -1),          3),
        ('LEFTPADDING',     (0, 0),        (-1, -1),          4),
        ('RIGHTPADDING',    (0, 0),        (-1, -1),          4),
    ])

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(20*cm, 1*cm, f"Page {self._pageNumber} of {page_count}")

def _header_footer(canvas, doc, title_info):
    canvas.saveState()
    st = _styles()
    
    # Header Area drawing directly on Canvas
    logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
    
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, 1.5*cm, 26.5*cm, width=2.2*cm, height=2.2*cm, preserveAspectRatio=True, mask='auto')

    # Univ Texts
    hindi_img_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'hindi_name.png')
    if os.path.exists(hindi_img_path):
        p1 = Image(hindi_img_path, width=9.5*cm, height=0.6*cm) # Aspect ratio roughly matches 60pt font over long text
    else:
        p1 = Paragraph(UNIV_HINDI, st['hindi'])
        
    p1.wrapOn(canvas, 13*cm, 2*cm)
    
    p2 = Paragraph(UNIV_ENG, st['eng'])
    p3 = Paragraph(UNIV_CITY, st['city'])
    
    p2.wrapOn(canvas, 13*cm, 2*cm)
    p3.wrapOn(canvas, 13*cm, 2*cm)
    
    if os.path.exists(hindi_img_path):
        p1.drawOn(canvas, 5.75*cm, 28.0*cm) # Centered
    else:
        p1.drawOn(canvas, 4*cm, 28.0*cm)
        
    p2.drawOn(canvas, 4*cm, 27.3*cm)
    p3.drawOn(canvas, 4*cm, 26.8*cm)
    
    # Separator Line
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(1)
    canvas.line(1.5*cm, 26.3*cm, 19.5*cm, 26.3*cm)
    
    # Report Title
    pt = Paragraph(title_info.get('title', ''), st['title'])
    pt.wrapOn(canvas, 16*cm, 1*cm)
    pt.drawOn(canvas, 2.5*cm, 25.4*cm)
    
    # Report Sub info
    canvas.setFont("Helvetica-Bold", 12)
    canvas.drawString(1.5*cm, 24.6*cm, f"Exam Centre :")
    canvas.setFont("Helvetica", 11)
    canvas.drawString(4.5*cm, 24.6*cm, f"{title_info.get('center', '')}")

    canvas.restoreState()

# ─────────────────────────────────────────────────────────────
#  APIs (FROM NEW CODE + OLD ONES TO PRESERVE SEAT ALLOTMENT)
# ─────────────────────────────────────────────────────────────
@seat_allocation_bp.route('/api/get_et_by_session', methods=['GET'])
def get_et_by_session():
    session_id = request.args.get('session_id', type=int)
    if not session_id:
        ets = PA_ET_Master.query.order_by(PA_ET_Master.id).all()
    else:
        ets = PA_ET_Master.query.filter_by(fk_session_id=session_id).order_by(PA_ET_Master.id).all()
    if not ets:
        ets = PA_ET_Master.query.order_by(PA_ET_Master.id).all()
    return jsonify([{"id": e.id, "name": e.description} for e in ets])

@seat_allocation_bp.route('/api/get_degrees_by_session', methods=['GET'])
def get_degrees_by_session():
    degrees = Degree.query.filter_by(active=True).order_by(Degree.name).all()
    return jsonify([{
        "id": d.id, 
        "name": f"{d.name} ({d.degree_type.description})" if d.degree_type else d.name
    } for d in degrees])

@seat_allocation_bp.route('/api/get_exam_centers_by_et', methods=['GET'])
def get_exam_centers_by_et():
    et_id = request.args.get('et_id', type=int)
    if not et_id:
        centers = PA_Exam_Center_Mst.query.filter_by(is_active=True).order_by(PA_Exam_Center_Mst.order_by).all()
    else:
        centers = PA_Exam_Center_Mst.query.filter_by(fk_et_id=et_id, is_active=True).order_by(PA_Exam_Center_Mst.order_by).all()
    if not centers:
        centers = PA_Exam_Center_Mst.query.filter_by(is_active=True).order_by(PA_Exam_Center_Mst.order_by).all()
    return jsonify([{"id": c.id, "name": c.name, "code": c.code or ''} for c in centers])

@seat_allocation_bp.route('/api/get_exam_types', methods=['GET'])
def get_exam_types():
    degree_id = request.args.get('degree_id', type=int)
    if not degree_id:
        return jsonify([])
    try:
        degree = Degree.query.get(degree_id)
        if not degree or not degree.is_entrance_exam:
            return jsonify([])
        rows = db.session.execute(text(
            "SELECT pk_ID, ExamType FROM PA_AnswerKeyExamType_mst ORDER BY pk_ID"
        )).fetchall()
        return jsonify([{"id": r[0], "name": r[1]} for r in rows])
    except Exception as e:
        print(f"Error fetching exam types: {e}")
        return jsonify([])

@seat_allocation_bp.route('/api/get_subjects_by_exam_type', methods=['GET'])
def get_subjects_by_exam_type():
    exam_type_ids = request.args.getlist('exam_type_ids[]')
    if not exam_type_ids:
        return jsonify([])
    ids_csv = ','.join(str(int(x)) for x in exam_type_ids if x.isdigit())
    if not ids_csv:
        return jsonify([])
    try:
        rows = db.session.execute(text(f"""
            SELECT PK_SubID, Subject
            FROM PA_Degree_Subject_Mst
            WHERE FK_ExamID IN ({ids_csv})
            ORDER BY Sequence
        """)).fetchall()
        return jsonify([{"id": r[0], "name": r[1]} for r in rows])
    except Exception as e:
        print(f"Error fetching subjects: {e}")
        return jsonify([])

# PRESERVED OLD GET_SUBJECTS FOR EXISTING TEMPLATE
@seat_allocation_bp.route('/api/get_subjects', methods=['GET'])
def get_subjects():
    ids = request.args.getlist('exam_type_ids[]')
    if not ids: return jsonify([])
    try:
        ids_clean = [int(x) for x in ids if x.isdigit()]
        if not ids_clean: return jsonify([])
        sql = f"SELECT PK_SubID, Subject FROM PA_Degree_Subject_Mst WHERE FK_ExamID IN ({','.join(map(str, ids_clean))}) ORDER BY Sequence"
        rows = db.session.execute(text(sql)).fetchall()
        return jsonify([{"id": r[0], "name": r[1]} for r in rows])
    except: return jsonify([])

@seat_allocation_bp.route('/api/get_rooms_by_center', methods=['GET'])
def get_rooms_by_center():
    center_id = request.args.get('center_id', type=int)
    et_id     = request.args.get('et_id', type=int)
    if not center_id:
        return jsonify([])
    query = PA_Exam_Center_Trn.query.filter_by(fk_exam_center_id=center_id)
    if et_id:
        used_rooms = db.session.execute(text("""
            SELECT DISTINCT fk_roomId 
            FROM PA_SeatAllotment_Details 
            WHERE fk_examCenterId = :cid AND Fk_ETID = :etid
        """), {'cid': center_id, 'etid': et_id}).fetchall()
        used_room_ids = [r[0] for r in used_rooms]
        if used_room_ids:
            query = query.filter(PA_Exam_Center_Trn.id.in_(used_room_ids))
        else:
            return jsonify([])
    rooms = query.order_by(PA_Exam_Center_Trn.order_by).all()
    return jsonify([{
        "id": r.id,
        "room_no": r.room_no,
        "capacity": r.room_capacity,
        "paper_dist": r.paper_dist or ''
    } for r in rooms])

@seat_allocation_bp.route('/api/get_colleges', methods=['GET'])
def get_colleges():
    degree_id  = request.args.get('degree_id', type=int)
    if not degree_id:
        return jsonify([])
    try:
        rows = db.session.execute(text("""
            SELECT DISTINCT c.Pk_CollegeID, c.CollegeName
            FROM PA_College_Mst c
            INNER JOIN PA_Degree_SpecializationMapping_mst m ON c.Pk_CollegeID = m.fk_CollegeId
            WHERE m.fk_DegreeId = :did
            ORDER BY c.CollegeName
        """), {'did': degree_id}).fetchall()
        return jsonify([{"id": r[0], "name": r[1]} for r in rows])
    except Exception as e:
        print(f"Error fetching colleges: {e}")
        return jsonify([])

@seat_allocation_bp.route('/api/get_specializations', methods=['GET'])
def get_specializations():
    college_ids = request.args.getlist('college_ids[]')
    degree_id   = request.args.get('degree_id', type=int)
    if not college_ids or not degree_id:
        return jsonify([])
    ids_csv = ','.join(str(int(x)) for x in college_ids if x.isdigit())
    if not ids_csv:
        return jsonify([])
    try:
        rows = db.session.execute(text(f"""
            SELECT DISTINCT s.Pk_SID, s.Specialization
            FROM PA_Specialization_mst s
            INNER JOIN PA_Degree_SpecializationMapping_mst m ON s.Pk_SID = m.fk_SID
            WHERE m.fk_CollegeId IN ({ids_csv}) AND m.fk_DegreeId = :did
            ORDER BY s.Specialization
        """), {'did': degree_id}).fetchall()
        return jsonify([{"id": r[0], "name": r[1]} for r in rows])
    except Exception as e:
        print(f"Error fetching specializations: {e}")
        return jsonify([])

@seat_allocation_bp.route('/api/get_allocation_status', methods=['GET'])
def get_allocation_status():
    et_id     = request.args.get('et_id', type=int)
    center_id = request.args.get('exam_center_id', type=int)
    session_id = request.args.get('session_id', type=int, default=71)
    if not et_id or not center_id:
        return jsonify([])
    center = PA_Exam_Center_Mst.query.get(center_id)
    if not center:
        return jsonify([])
    rooms = (PA_Exam_Center_Trn.query
             .filter_by(fk_exam_center_id=center_id)
             .order_by(PA_Exam_Center_Trn.order_by)
             .all())
    try:
        allotted_rows = db.session.execute(text("""
            SELECT fk_roomId, COUNT(*) as allotted_count 
            FROM PA_SeatAllotment_Details 
            WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
            GROUP BY fk_roomId
        """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
        allotted_map = {row[0]: row[1] for row in allotted_rows}
    except Exception as e:
        print(f"Error fetching allotted counts: {e}")
        allotted_map = {}
    data = []
    for i, r in enumerate(rooms, 1):
        allotted = allotted_map.get(r.id, 0)
        data.append({
            "sno":      i,
            "room_id":  r.id,
            "center":   center.name,
            "room":     r.room_no or str(i),
            "capacity": r.room_capacity,
            "allotted": allotted,
            "available": max(0, r.room_capacity - allotted),
        })
    return jsonify(data)

@seat_allocation_bp.route('/api/get_candidates', methods=['GET'])
def get_candidates():
    et_id      = request.args.get('et_id', type=int)
    center_id  = request.args.get('exam_center_id', type=int)
    room_id    = request.args.get('room_id', type=int)
    session_id = request.args.get('session_id', type=int, default=71)
    status     = request.args.get('status', type=str, default='all')
    if not et_id:
        return jsonify([])
    try:
        sql = """
            SELECT 
                r.regno, 
                r.s_name, 
                r.f_name, 
                r.mobileno,
                s.Specialization, 
                c.Name AS center_name, 
                rm.RoomNo, 
                rm.RoomCapacity,
                rm.NoRow,
                rm.NoColumn,
                rm.Paper_Dist,
                d.SeatNo, 
                d.RollNo,
                deg.description AS degree_desc,
                et.Description AS et_desc
            FROM PA_SeatAllotment_Details d
            INNER JOIN PA_Registration_Mst r ON d.fk_regid = r.pk_regid
            INNER JOIN PA_Exam_Center_Mst c ON d.fk_examCenterId = c.pk_examCenterId
            INNER JOIN PA_Exam_Center_Trn rm ON d.fk_roomId = rm.pk_id
            INNER JOIN PA_ET_Master et ON d.Fk_ETID = et.Pk_ETID
            LEFT JOIN PA_Specialization_mst s ON r.fk_SId = s.Pk_SID
            LEFT JOIN ACD_Degree_Mst deg ON r.fk_dtypeid = deg.pk_degreeid
            WHERE d.Fk_ETID = :etid AND d.fk_sessionid = :sid
        """
        params = {'etid': et_id, 'sid': session_id}
        if center_id:
            sql += " AND d.fk_examCenterId = :cid"
            params['cid'] = center_id
        if room_id:
            sql += " AND d.fk_roomId = :roomid"
            params['roomid'] = room_id
        if status == 'unprocessed':
            sql += " AND (d.RollNo IS NULL OR d.RollNo = '' OR d.RollNo = '0')"
        elif status == 'processed':
            sql += " AND (d.RollNo IS NOT NULL AND d.RollNo != '' AND d.RollNo != '0')"
        sql += " ORDER BY c.OrderBy, rm.RoomNo, d.SeatNo"
        rows = db.session.execute(text(sql), params).fetchall()
        data = []
        for row in rows:
            data.append({
                "reg_no": row.regno,
                "name": row.s_name,
                "father_name": row.f_name,
                "mobile": row.mobileno,
                "subject": row.Specialization or '',
                "center": row.center_name,
                "room": row.RoomNo,
                "capacity": getattr(row, 'RoomCapacity', 0),
                "no_row": getattr(row, 'NoRow', ''),
                "no_column": getattr(row, 'NoColumn', ''),
                "paper_dist": getattr(row, 'Paper_Dist', ''),
                "seat": row.SeatNo,
                "roll_no": row.RollNo,
                "degree_type": getattr(row, 'degree_desc', ''),
                "exam_type": getattr(row, 'et_desc', '')
            })
        return jsonify(data)
    except Exception as e:
        print(f"Error fetching candidates: {e}")
        return jsonify([])

# PRESERVED OLD GET_UNALLOTTED_CANDIDATES FOR SEAT ALLOTMENT TEMPLATE
@seat_allocation_bp.route('/api/get_unallotted_candidates', methods=['GET'])
def get_unallotted_candidates():
    sid = request.args.get('session_id', type=int); did = request.args.get('degree_id', type=int)
    college_ids = request.args.getlist('college_ids[]')
    spec_ids    = request.args.getlist('specialization_ids[]')
    if not sid or not did: return jsonify([])
    try:
        sql = """SELECT DISTINCT r.pk_regid, r.regno, (ISNULL(r.s_name, '') + ' ' + ISNULL(r.s_surname, '')) as name, r.f_name as father_name, s.Specialization as subject
                 FROM PA_Registration_Mst r LEFT JOIN PA_Specialization_mst s ON r.fk_SId = s.Pk_SID
                 WHERE r.fk_sessionid = :sid AND r.fk_dtypeid = :did AND r.IsPaymentSuccess = 1
                 AND (r.pk_regid NOT IN (SELECT fk_regid FROM PA_SeatAllotment_Details WHERE fk_sessionid = :sid AND fk_regid IS NOT NULL)
                      OR r.pk_regid IN (SELECT fk_regid FROM PA_SeatAllotment_Details WHERE fk_sessionid = :sid AND fk_roomId IS NULL))"""
        params = {'sid': sid, 'did': did}
        if college_ids:
            c_ids = [int(x) for x in college_ids if x.isdigit()]
            if c_ids: sql += f" AND r.fk_CollegID IN ({','.join(map(str, c_ids))})"
        if spec_ids:
            s_ids = [int(x) for x in spec_ids if x.isdigit()]
            if s_ids: sql += f" AND r.fk_SId IN ({','.join(map(str, s_ids))})"
        sql += " ORDER BY name ASC"
        rows = db.session.execute(text(sql), params).mappings().all()
        return jsonify([dict(r) for r in rows])
    except: return jsonify([])

# ─────────────────────────────────────────────────────────────
#  OLD SEAT ALLOTMENT ROUTE & PDF (PRESERVED EXACTLY AS IS)
# ─────────────────────────────────────────────────────────────
def _pdf_seat_allotment_details(et_name, center_name, session_id, et_id, center_id, room_filter=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=5.5*cm, bottomMargin=1.5*cm)
    st = _styles(); els = []
    try:
        sql = """SELECT r.regno, r.s_name, r.f_name, s.Specialization, rm.RoomNo, d.SeatNo, d.RollNo
                 FROM PA_SeatAllotment_Details d
                 INNER JOIN PA_Registration_Mst r ON d.fk_regid = r.pk_regid
                 INNER JOIN PA_Exam_Center_Trn rm ON d.fk_roomId = rm.pk_id
                 LEFT JOIN PA_Specialization_mst s ON r.fk_SId = s.Pk_SID
                 WHERE d.Fk_ETID = :etid AND d.fk_examCenterId = :cid AND d.fk_sessionid = :sid"""
        params = {'etid': et_id, 'cid': center_id, 'sid': session_id}
        if room_filter: sql += " AND d.fk_roomId = :roomid"; params['roomid'] = room_filter
        sql += " ORDER BY rm.RoomNo, d.SeatNo"
        rows = db.session.execute(text(sql), params).fetchall()
        rooms_map = defaultdict(list)
        for row in rows:
            rooms_map[row.RoomNo].append({'reg_no': row.regno, 'name': row.s_name, 'father_name': row.f_name,
                                           'subject': row.Specialization or '', 'seat': row.SeatNo, 'roll_no': row.RollNo})
    except Exception as e:
        print(f"Error fetching real candidates for PDF: {e}"); rooms_map = {}
    if not rooms_map:
        els.append(Paragraph("No records found.", st['normal']))
    else:
        col_w = [1.2*cm, 2.5*cm, 4.2*cm, 4.2*cm, 2.5*cm, 1.5*cm, 1.8*cm]
        for room_no, clist in sorted(rooms_map.items()):
            els.append(Paragraph(f"Room.No: &nbsp;&nbsp; {room_no}", ParagraphStyle('rm', fontName='Helvetica-Bold', fontSize=12, spaceAfter=8)))
            hdr = [['Sr.No','Reg.No','Name','Father Name','Specialization','Seat.No','RollNo']]
            rows_data = hdr + [[str(idx), str(c.get('reg_no','')),
                                 Paragraph(c.get('name',''), st['normal']), Paragraph(c.get('father_name',''), st['normal']),
                                 Paragraph(c.get('subject',''), st['normal']), str(c.get('seat','')), str(c.get('roll_no',''))]
                                for idx, c in enumerate(clist, 1)]
            t = Table(rows_data, colWidths=col_w, repeatRows=1); ts = _grid_style()
            ts.add('ALIGN',(0,1),(1,-1),'CENTER'); ts.add('ALIGN',(5,1),(6,-1),'CENTER')
            t.setStyle(ts); els.append(t); els.append(Spacer(1,15))
    title_info = {'title': 'Seat Allotment Details', 'center': center_name}
    doc.build(els, canvasmaker=NumberedCanvas,
              onFirstPage=lambda cv, dc: _header_footer(cv, dc, title_info),
              onLaterPages=lambda cv, dc: _header_footer(cv, dc, title_info))
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = 'inline; filename="Seat_Allotment_Details.pdf"'
    return resp

@seat_allocation_bp.route('/seat-allotment', methods=['GET', 'POST'])
def seat_allotment():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'process_allotment':
            sid = request.form.get('session_id', type=int)
            et_id = request.form.get('et_id', type=int)
            center_id = request.form.get('exam_center_id', type=int)
            reg_ids_raw = request.form.getlist('cand_na')
            room_ids_raw = request.form.getlist('room_sel')
            if not sid or not et_id or not center_id:
                flash('Please select Session, ET Type and Exam Centre before processing.', 'error')
            elif not reg_ids_raw:
                flash('No candidates selected for allotment.', 'error')
            elif not room_ids_raw:
                flash('No rooms selected for seat allotment.', 'error')
            else:
                try:
                    reg_ids = [int(x) for x in reg_ids_raw if str(x).isdigit()]
                    room_ids = [int(x) for x in room_ids_raw if str(x).isdigit()]
                    if not reg_ids or not room_ids:
                        flash('Invalid candidate or room selection.', 'error')
                    else:
                        # Build per-room availability (capacity, allotted count, next seat no)
                        rooms_data = []
                        for room_id in room_ids:
                            row = db.session.execute(text(
                                "SELECT RoomCapacity FROM PA_Exam_Center_Trn WHERE pk_id = :rid"
                            ), {'rid': room_id}).fetchone()
                            cnt = db.session.execute(text(
                                "SELECT COUNT(*) FROM PA_SeatAllotment_Details WHERE fk_roomId=:rid AND fk_sessionid=:sid AND Fk_ETID=:etid"
                            ), {'rid': room_id, 'sid': sid, 'etid': et_id}).fetchone()[0]
                            max_seat = db.session.execute(text(
                                "SELECT ISNULL(MAX(SeatNo), 0) FROM PA_SeatAllotment_Details WHERE fk_roomId=:rid AND fk_sessionid=:sid AND Fk_ETID=:etid"
                            ), {'rid': room_id, 'sid': sid, 'etid': et_id}).fetchone()[0]
                            rooms_data.append({'id': room_id, 'capacity': (row[0] if row else 0),
                                               'allotted': cnt, 'next_seat': (max_seat or 0) + 1})

                        allotted_count = 0
                        queue = list(reg_ids)

                        for room in rooms_data:
                            available = room['capacity'] - room['allotted']
                            if available <= 0 or not queue:
                                continue
                            seats_to_assign = min(available, len(queue))
                            for i in range(seats_to_assign):
                                reg_id = queue.pop(0)
                                seat_no = room['next_seat'] + i
                                # Update existing unprocessed row, or insert new
                                existing = db.session.execute(text(
                                    "SELECT Pk_Id FROM PA_SeatAllotment_Details WHERE fk_regid=:rid AND fk_sessionid=:sid AND Fk_ETID=:etid"
                                ), {'rid': reg_id, 'sid': sid, 'etid': et_id}).fetchone()
                                if existing:
                                    db.session.execute(text(
                                        "UPDATE PA_SeatAllotment_Details SET fk_roomId=:room, fk_examCenterId=:center, SeatNo=:seat WHERE Pk_Id=:pk"
                                    ), {'room': room['id'], 'center': center_id, 'seat': seat_no, 'pk': existing[0]})
                                else:
                                    db.session.execute(text(
                                        "INSERT INTO PA_SeatAllotment_Details (fk_regid, fk_examCenterId, fk_roomId, SeatNo, fk_sessionid, Fk_ETID) VALUES (:reg, :center, :room, :seat, :sid, :etid)"
                                    ), {'reg': reg_id, 'center': center_id, 'room': room['id'], 'seat': seat_no, 'sid': sid, 'etid': et_id})
                                allotted_count += 1
                            room['next_seat'] += seats_to_assign

                        db.session.commit()
                        skipped = len(queue)
                        msg = f"{allotted_count} candidate(s) seat allotted successfully."
                        if skipped:
                            msg += f" {skipped} candidate(s) could not be allotted (insufficient room capacity in selected rooms)."
                        flash(msg, 'success' if allotted_count else 'error')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing allotment: {e}', 'error')

        elif action == 'unprocess_allotment':
            sid = request.form.get('session_id', type=int); et_id = request.form.get('allotted_et_id', type=int); reg_nos = request.form.getlist('cand_sel')
            if sid and et_id and reg_nos:
                try:
                    placeholders = ', '.join(f"'{r}'" for r in reg_nos if r.isalnum())
                    if placeholders:
                        rows = db.session.execute(text(f"SELECT pk_regid, regno, s_name FROM PA_Registration_Mst WHERE regno IN ({placeholders})")).fetchall()
                        msg_parts = []
                        reg_ids = []
                        for r in rows:
                            reg_ids.append(r[0])
                            msg_parts.append(f"{r[1]} - {r[2]}")

                        if reg_ids:
                            ids_str = ', '.join(map(str, reg_ids))
                            db.session.execute(text(f"UPDATE PA_Registration_Mst SET rollno = NULL WHERE pk_regid IN ({ids_str})"))
                            db.session.execute(text(f"UPDATE PA_SeatAllotment_Details SET fk_roomId = NULL, SeatNo = NULL, RollNo = NULL WHERE Fk_ETID = :etid AND fk_sessionid = :sid AND fk_regid IN ({ids_str})"), {'etid': et_id, 'sid': sid})
                            db.session.commit()
                            flash(f"{len(reg_ids)} student(s) unprocessed successfully.", "success")
                except Exception as e: db.session.rollback(); flash(f'Error: {e}', 'error')
        return redirect(url_for('seat_allocation.seat_allotment'))
    
    sessions = AcademicSession.query.filter_by(is_active=True).order_by(AcademicSession.id.desc()).all()
    return render_template('seat_allocation/seat_allotment.html', sessions=sessions)

# ─────────────────────────────────────────────────────────────
#  NEW PDF / EXCEL HELPERS (FROM NEW CODE)
# ─────────────────────────────────────────────────────────────

def _pdf_exam_center_report(session_name, et_name, et_id=None, session_id=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    st = _styles()
    els = []
    _univ_header(els, st)
    
    title_text = f"List of Centers ({et_name}) {session_name}"
    els.append(Paragraph(f"<b>{title_text}</b>", st['title']))
    els.append(Spacer(1, 10))

    sql = """
        SELECT 
            c.pk_examCenterId, 
            c.Name, 
            (SELECT COUNT(*) FROM PA_Exam_Center_Trn t WHERE t.fk_examCenterId = c.pk_examCenterId) as room_count,
            (SELECT COUNT(*) FROM PA_SeatAllotment_Details d WHERE d.fk_examCenterId = c.pk_examCenterId AND d.Fk_ETID = :et_id AND d.fk_sessionid = :sid) as cand_count
        FROM PA_Exam_Center_Mst c
        WHERE c.fk_ETID = :et_id AND c.fk_SessionId = :sid AND c.IsActive = 1
        ORDER BY c.OrderBy
    """
    results = db.session.execute(text(sql), {'et_id': et_id, 'sid': session_id}).fetchall() if (et_id and session_id) else []
    
    hdr = [['Sr. No.', 'Name of the Centre', 'Total No. of candidates', 'Total No. of Rooms']]
    rows = []
    for i, r in enumerate(results, 1):
        rows.append([str(i), Paragraph(r.Name, st['normal']), str(r.cand_count), str(r.room_count)])
    
    if not rows:
        rows = [['', 'No exam centers found for the selected ET and session.', '', '']]

    col_w = [1.5*cm, 10.5*cm, 3.5*cm, 2.5*cm]
    t = Table(hdr + rows, colWidths=col_w, repeatRows=1)
    ts = _grid_style()
    ts.add('ALIGN', (0,0), (0,-1), 'CENTER')
    ts.add('ALIGN', (2,0), (-1,-1), 'CENTER')
    ts.add('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
    t.setStyle(ts)
    els.append(t)

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="Exam_Center_Report_{et_id}.pdf"'
    return resp

def _excel_exam_center_detail_report(session_name, et_name, center_name, session_id, et_id, center_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Center Detail Report"
    
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    alt_fill = PatternFill(start_color="EEF2FB", end_color="EEF2FB", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "Exam Center Detail Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Session : {session_name}      ET Type : {et_name}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = center_align
    
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Exam Center : {center_name}"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = center_align
    
    headers = ['S.No', 'Room No.', 'Capacity', 'Allotted', 'Starting Roll No', 'Ending Roll No']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    rooms_data = []
    if center_id:
        rooms = (PA_Exam_Center_Trn.query
                 .filter_by(fk_exam_center_id=center_id)
                 .order_by(PA_Exam_Center_Trn.order_by)
                 .all())
        try:
            allotted_rows = db.session.execute(text("""
                SELECT fk_roomId, COUNT(*) as allotted_count, MIN(RollNo) as min_roll, MAX(RollNo) as max_roll 
                FROM PA_SeatAllotment_Details 
                WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
                GROUP BY fk_roomId
            """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
            allotted_map = {row[0]: (row[1], row[2], row[3]) for row in allotted_rows}
        except Exception:
            allotted_map = {}

        for i, r in enumerate(rooms, 1):
            allotted, min_roll, max_roll = allotted_map.get(r.id, (0, '', ''))
            rooms_data.append({
                "sno": i,
                "room_no": r.room_no or str(i),
                "capacity": r.room_capacity,
                "allotted": allotted,
                "start_roll": min_roll or '',
                "end_roll": max_roll or ''
            })
            
    row_idx = 6
    if not rooms_data:
        ws.merge_cells('A6:F6')
        cell = ws.cell(row=6, column=1, value="No rooms configured.")
        cell.alignment = center_align
    else:
        for r in rooms_data:
            fill = alt_fill if r['sno'] % 2 == 0 else PatternFill(fill_type=None)
            
            c1 = ws.cell(row=row_idx, column=1, value=r['sno'])
            c2 = ws.cell(row=row_idx, column=2, value=r['room_no'])
            c3 = ws.cell(row=row_idx, column=3, value=r['capacity'])
            c4 = ws.cell(row=row_idx, column=4, value=r['allotted'])
            c5 = ws.cell(row=row_idx, column=5, value=r['start_roll'])
            c6 = ws.cell(row=row_idx, column=6, value=r['end_roll'])
            
            for cell in [c1, c2, c3, c4, c5, c6]:
                cell.border = thin_border
                cell.fill = fill
                cell.alignment = center_align
            row_idx += 1
            
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="Exam_Center_Detail_Report_{center_id}.xlsx"'
    return resp

def _excel_exam_center_distribution_pattern(session_name, et_name, center_name, session_id, et_id, center_id, et_date_str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribution Pattern"
    
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    alt_fill = PatternFill(start_color="EEF2FB", end_color="EEF2FB", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "Exam Center Wise Detail Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"{et_name} : {et_date_str}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center_align
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Exam Center : {center_name}"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = left_align
    
    headers = ['S.No', 'Room No.', 'Room Location', 'Roll No.', 'Total Seats', 'Seating Pattern (Rows X Col.)', 'Paper Distribution Pattern', 'Alloted Seats']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    rooms_data = []
    if center_id:
        rooms = (PA_Exam_Center_Trn.query
                 .filter_by(fk_exam_center_id=center_id)
                 .order_by(PA_Exam_Center_Trn.order_by)
                 .all())
        try:
            allotted_rows = db.session.execute(text("""
                SELECT fk_roomId, COUNT(*) as allotted_count, MIN(RollNo) as min_roll, MAX(RollNo) as max_roll 
                FROM PA_SeatAllotment_Details 
                WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
                GROUP BY fk_roomId
            """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
            allotted_map = {row[0]: (row[1], row[2], row[3]) for row in allotted_rows}
        except Exception:
            allotted_map = {}

        for i, r in enumerate(rooms, 1):
            allotted, min_roll, max_roll = allotted_map.get(r.id, (0, '', ''))
            roll_str = f"{min_roll} - {max_roll}" if min_roll and max_roll else ""
            rooms_data.append({
                "sno": i,
                "room_no": r.room_no or str(i),
                "location": r.room_location or '',
                "roll_no": roll_str,
                "total_seats": r.room_capacity,
                "seating_pattern": f"{r.no_row} X {r.no_column}" if r.no_row and r.no_column else "",
                "paper_dist": r.paper_dist or '',
                "allotted": allotted
            })
            
    row_idx = 6
    if not rooms_data:
        ws.merge_cells('A6:H6')
        cell = ws.cell(row=6, column=1, value="No rooms configured.")
        cell.alignment = center_align
    else:
        for r in rooms_data:
            fill = alt_fill if r['sno'] % 2 == 0 else PatternFill(fill_type=None)
            
            c1 = ws.cell(row=row_idx, column=1, value=r['sno'])
            c2 = ws.cell(row=row_idx, column=2, value=r['room_no'])
            c3 = ws.cell(row=row_idx, column=3, value=r['location'])
            c4 = ws.cell(row=row_idx, column=4, value=r['roll_no'])
            c5 = ws.cell(row=row_idx, column=5, value=r['total_seats'])
            c6 = ws.cell(row=row_idx, column=6, value=r['seating_pattern'])
            c7 = ws.cell(row=row_idx, column=7, value=r['paper_dist'])
            c8 = ws.cell(row=row_idx, column=8, value=r['allotted'])
            
            for cell in [c1, c2, c3, c4, c5, c6, c7, c8]:
                cell.border = thin_border
                cell.fill = fill
                cell.alignment = center_align if cell.column in [1,2,4,5,6,8] else left_align
            row_idx += 1
            
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 12
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="Distribution_Pattern_{center_id}.xlsx"'
    return resp

def _pdf_exam_center_distribution_pattern(session_name, et_name, center_name, session_id=None, et_id=None, center_id=None, et_date_str='N/A'):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    st = _styles()
    els = []

    now = datetime.now()
    curr_date = now.strftime("%m/%d/%Y")
    curr_time = now.strftime("%I:%M:%S %p").lower()

    logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
    hindi_img_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'hindi_name.png')

    if os.path.exists(hindi_img_path):
        hindi_el = Image(hindi_img_path, width=9.5*cm, height=0.6*cm)
    else:
        hindi_el = Paragraph(UNIV_HINDI, st['hindi'])

    img = ''
    if os.path.exists(logo_path):
        img = Image(logo_path, width=2.2*cm, height=2.2*cm)

    text_col = [
        Paragraph(UNIV_ENG, ParagraphStyle('ue', fontName='Arial-Bold', fontSize=14, alignment=TA_CENTER, leading=16)),
        Spacer(1, 4),
        Paragraph("<u><b>Exam Center Wise Detail Report</b></u>", ParagraphStyle('et1', fontName='Arial-Bold', fontSize=12, alignment=TA_CENTER, leading=14)),
        Spacer(1, 4),
        Paragraph(f"<u><b>{et_name} : {et_date_str}</b></u>", ParagraphStyle('et2', fontName='Arial-Bold', fontSize=12, alignment=TA_CENTER, leading=14)),
    ]

    dt_col = [
        Paragraph(f"<b>Date :</b> &nbsp;&nbsp;&nbsp; {curr_date}", ParagraphStyle('dt', fontName='Arial-Bold', fontSize=8, alignment=TA_RIGHT, leading=12)),
        Paragraph(f"<b>Time :</b> &nbsp;&nbsp;&nbsp; {curr_time}", ParagraphStyle('dt', fontName='Arial-Bold', fontSize=8, alignment=TA_RIGHT, leading=12))
    ]

    header_table = Table([[img, text_col, dt_col]], colWidths=[3.0*cm, 12.5*cm, 3.5*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('ALIGN', (2,0), (2,0), 'RIGHT')
    ]))
    els.append(header_table)
    els.append(Spacer(1, 15))
    els.append(Paragraph(f"<b>Exam Center : {center_name.upper()}</b>", ParagraphStyle('ecn', fontName='Arial-Bold', fontSize=10, alignment=TA_LEFT)))
    els.append(Spacer(1, 10))

    rooms_data = []
    if center_id:
        rooms = (PA_Exam_Center_Trn.query
                 .filter_by(fk_exam_center_id=center_id)
                 .order_by(PA_Exam_Center_Trn.order_by)
                 .all())
        try:
            allotted_rows = db.session.execute(text("""
                SELECT fk_roomId, COUNT(*) as allotted_count, MIN(RollNo) as min_roll, MAX(RollNo) as max_roll 
                FROM PA_SeatAllotment_Details 
                WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
                GROUP BY fk_roomId
            """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
            allotted_map = {row[0]: (row[1], row[2], row[3]) for row in allotted_rows}
        except Exception:
            allotted_map = {}

        for i, r in enumerate(rooms, 1):
            allotted, min_roll, max_roll = allotted_map.get(r.id, (0, '', ''))
            roll_str = f"{min_roll} - {max_roll}" if min_roll and max_roll else ""
            sp = f"{r.no_row} X {r.no_column}" if r.no_row and r.no_column else ""
            rooms_data.append({
                "sno": i,
                "room_no": r.room_no or str(i),
                "location": r.room_location or '',
                "roll_no": roll_str,
                "total_seats": r.room_capacity,
                "sp": sp,
                "paper_dist": r.paper_dist or '',
                "allotted": allotted
            })

    hdr = [['S.No', 'Room No.', 'Room Location', 'Roll No.', 'Total Seats', 'Seating\nPattern\n(Rows X Col.)', 'Paper\nDistribution\nPattern', 'Alloted\nSeats']]
    rows = []
    for r in rooms_data:
        rows.append([str(r['sno']), str(r['room_no']), Paragraph(r['location'], st['normal']), 
                     str(r['roll_no']), str(r['total_seats']), str(r['sp']), 
                     Paragraph(r['paper_dist'], st['center9']), str(r['allotted'])])
    if not rows:
        rows = [['', 'No rooms configured.', '', '', '', '', '', '']]

    col_w = [1*cm, 1.8*cm, 4.2*cm, 3.5*cm, 1.5*cm, 2.5*cm, 3.5*cm, 1.5*cm]
    t = Table(hdr + rows, colWidths=col_w, repeatRows=1)
    ts = _grid_style()
    ts.add('ALIGN', (0,1), (-1,-1), 'CENTER')
    ts.add('VALIGN', (0,1), (-1,-1), 'MIDDLE')
    t.setStyle(ts)
    els.append(t)

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="Distribution_Pattern_{center_id}.pdf"'
    return resp

def _pdf_exam_center_display_report(session_name, et_name, center_name, session_id=None, et_id=None, center_id=None, et_date_str='N/A'):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1*inch, rightMargin=1*inch,
                            topMargin=0.8*inch, bottomMargin=0.8*inch)
    st = _styles()
    els = []

    logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
    hindi_img_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'hindi_name.png')

    if os.path.exists(hindi_img_path):
        hindi_el = Image(hindi_img_path, width=10.5*cm, height=0.7*cm)
    else:
        hindi_el = Paragraph(UNIV_HINDI, st['hindi'])

    img = ''
    if os.path.exists(logo_path):
        img = Image(logo_path, width=2.5*cm, height=2.5*cm)

    text_col = [
        Paragraph(UNIV_ENG, ParagraphStyle('ue', fontName='Arial-Bold', fontSize=15, alignment=TA_CENTER, leading=18)),
        Spacer(1, 6),
        Paragraph(f"<u><b>{et_name} : {et_date_str}</b></u>", ParagraphStyle('et2', fontName='Arial-Bold', fontSize=14, alignment=TA_CENTER, leading=16)),
    ]

    header_table = Table([[img, text_col]], colWidths=[3.5*cm, 12.5*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
    ]))
    els.append(header_table)
    els.append(Spacer(1, 15))
    
    els.append(Paragraph(f"<b>Exam Center : {center_name.upper()}</b>", ParagraphStyle('ecn', fontName='Arial-Bold', fontSize=12, alignment=TA_LEFT)))
    els.append(Spacer(1, 15))

    rooms_data = []
    if center_id:
        rooms = (PA_Exam_Center_Trn.query
                 .filter_by(fk_exam_center_id=center_id)
                 .order_by(PA_Exam_Center_Trn.order_by)
                 .all())
        try:
            allotted_rows = db.session.execute(text("""
                SELECT fk_roomId, MIN(RollNo) as min_roll, MAX(RollNo) as max_roll 
                FROM PA_SeatAllotment_Details 
                WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
                GROUP BY fk_roomId
            """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
            allotted_map = {row[0]: (row[1], row[2]) for row in allotted_rows}
        except Exception:
            allotted_map = {}

        for i, r in enumerate(rooms, 1):
            min_roll, max_roll = allotted_map.get(r.id, ('', ''))
            roll_str = f"{min_roll} - {max_roll}" if min_roll and max_roll else ""
            rooms_data.append({
                "sno": i,
                "room_no": r.room_no or str(i),
                "location": r.room_location or '',
                "roll_no": roll_str
            })

    hdr = [['S.No', 'Room No.', 'Room Location', 'Roll No.']]
    rows = []
    for r in rooms_data:
        rows.append([str(r['sno']), str(r['room_no']), Paragraph(r['location'], ParagraphStyle('n', fontName='Arial', fontSize=11, leading=14)), str(r['roll_no'])])
    
    if not rows:
        rows = [['', 'No rooms configured.', '', '']]

    col_w = [1.5*cm, 2.5*cm, 8.5*cm, 3.5*cm]
    t = Table(hdr + rows, colWidths=col_w, repeatRows=1)
    
    ts = TableStyle([
        ('FONTNAME',        (0, 0),        (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',        (0, 0),        (-1, 0), 11),
        ('BACKGROUND',      (0, 0),        (-1, 0), GRID_HDR_BG),
        ('TEXTCOLOR',       (0, 0),        (-1, 0), colors.white),
        ('ALIGN',           (0, 0),        (-1, 0), 'CENTER'),
        ('VALIGN',          (0, 0),        (-1, -1), 'MIDDLE'),
        ('FONTNAME',        (0, 1),        (-1, -1), 'Helvetica'),
        ('FONTSIZE',        (0, 1),        (-1, -1), 11),
        ('ALIGN',           (0, 1),        (1, -1), 'CENTER'),
        ('ALIGN',           (2, 1),        (2, -1), 'LEFT'),
        ('ALIGN',           (3, 1),        (3, -1), 'CENTER'),
        ('ROWBACKGROUNDS',  (0, 1),        (-1, -1), [colors.white, GRID_ALT_BG]),
        ('GRID',            (0, 0),        (-1, -1), 0.4, BORDER_CLR),
        ('TOPPADDING',      (0, 0),        (-1, -1), 6),
        ('BOTTOMPADDING',   (0, 0),        (-1, -1), 6),
    ])
    t.setStyle(ts)
    els.append(t)

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="Display_Report_{center_id}.pdf"'
    return resp

def _pdf_exam_center_door_placards(session_name, et_name, center_name, session_id=None, et_id=None, center_id=None, et_date_str='N/A'):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1*inch, rightMargin=1*inch,
                            topMargin=1.5*inch, bottomMargin=1*inch)
    st = _styles()
    els = []
    
    rooms_data = []
    if center_id:
        rooms = (PA_Exam_Center_Trn.query
                 .filter_by(fk_exam_center_id=center_id)
                 .order_by(PA_Exam_Center_Trn.order_by)
                 .all())
        try:
            allotted_rows = db.session.execute(text("""
                SELECT fk_roomId, MIN(RollNo) as min_roll, MAX(RollNo) as max_roll 
                FROM PA_SeatAllotment_Details 
                WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
                GROUP BY fk_roomId
            """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
            allotted_map = {row[0]: (row[1], row[2]) for row in allotted_rows}
        except Exception:
            allotted_map = {}

        for r in rooms:
            min_roll, max_roll = allotted_map.get(r.id, ('', ''))
            if min_roll and max_roll:
                rooms_data.append({
                    "room_no": r.room_no or '',
                    "start": min_roll,
                    "end": max_roll
                })

    if not rooms_data:
        els.append(Paragraph("No rooms allotted for this exam center.", st['title']))
        doc.build(els)
        buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type']        = 'application/pdf'
        resp.headers['Content-Disposition'] = f'inline; filename="Room_Wise_Placards_{center_id}.pdf"'
        return resp

    for idx, r in enumerate(rooms_data):
        els.append(Paragraph("<b>CCS Haryana Agricultural University</b>", ParagraphStyle('c1', fontName='Arial-Bold', fontSize=22, alignment=TA_CENTER, leading=26)))
        els.append(Spacer(1, 20))
        els.append(Paragraph(f"<b>{et_name} : {et_date_str}</b>", ParagraphStyle('c2', fontName='Arial-Bold', fontSize=18, alignment=TA_CENTER, leading=22)))
        els.append(Spacer(1, 30))
        els.append(Paragraph(f"<b>Exam Center : {center_name}</b>", ParagraphStyle('c3', fontName='Arial-Bold', fontSize=18, alignment=TA_CENTER, leading=22)))
        els.append(Spacer(1, 40))
        
        rt = Table([[Paragraph(f"<b>Room No. {r['room_no']}</b>", ParagraphStyle('c4', fontName='Arial-Bold', fontSize=28, alignment=TA_CENTER, leading=34))]], colWidths=[12*cm], rowHeights=[3*cm])
        rt.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 2, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        
        container1 = Table([[rt]], colWidths=[16*cm])
        container1.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
        els.append(container1)
        els.append(Spacer(1, 40))
        
        els.append(Paragraph("<b>Roll Nos.</b>", ParagraphStyle('c5', fontName='Arial-Bold', fontSize=22, alignment=TA_CENTER, leading=26)))
        els.append(Spacer(1, 20))
        
        rt2 = Table([[Paragraph(f"<b>{r['start']} - {r['end']}</b>", ParagraphStyle('c6', fontName='Arial-Bold', fontSize=34, alignment=TA_CENTER, leading=40))]], colWidths=[14*cm], rowHeights=[3.5*cm])
        rt2.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 2, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        
        container2 = Table([[rt2]], colWidths=[16*cm])
        container2.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
        els.append(container2)
        
        if idx < len(rooms_data) - 1:
            els.append(PageBreak())

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="Room_Wise_Placards_{center_id}.pdf"'
    return resp

def _pdf_exam_center_detail_report(session_name, et_name, center_name, session_id=None, et_id=None, center_id=None, et_date_str='N/A'):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    st = _styles()
    els = []

    now = datetime.now()
    curr_date = now.strftime("%m/%d/%Y")
    curr_time = now.strftime("%I:%M:%S %p").lower()

    logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
    hindi_img_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'hindi_name.png')

    if os.path.exists(hindi_img_path):
        hindi_el = Image(hindi_img_path, width=9.5*cm, height=0.6*cm)
    else:
        hindi_el = Paragraph(UNIV_HINDI, st['hindi'])

    img = ''
    if os.path.exists(logo_path):
        img = Image(logo_path, width=2.2*cm, height=2.2*cm)

    text_col = [
        Paragraph(UNIV_ENG, ParagraphStyle('ue', fontName='Arial-Bold', fontSize=14, alignment=TA_CENTER, leading=16)),
        Spacer(1, 4),
        Paragraph("<u><b>Exam Center Wise Detail Report</b></u>", ParagraphStyle('et1', fontName='Arial-Bold', fontSize=12, alignment=TA_CENTER, leading=14)),
        Spacer(1, 4),
        Paragraph(f"<u><b>{et_name} : {et_date_str}</b></u>", ParagraphStyle('et2', fontName='Arial-Bold', fontSize=12, alignment=TA_CENTER, leading=14)),
    ]

    dt_col = [
        Paragraph(f"<b>Date :</b> &nbsp;&nbsp;&nbsp; {curr_date}", ParagraphStyle('dt', fontName='Arial-Bold', fontSize=8, alignment=TA_RIGHT, leading=12)),
        Paragraph(f"<b>Time :</b> &nbsp;&nbsp;&nbsp; {curr_time}", ParagraphStyle('dt', fontName='Arial-Bold', fontSize=8, alignment=TA_RIGHT, leading=12))
    ]

    header_table = Table([[img, text_col, dt_col]], colWidths=[3.0*cm, 12.5*cm, 3.5*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('ALIGN', (2,0), (2,0), 'RIGHT')
    ]))
    els.append(header_table)

    els.append(Spacer(1, 15))
    
    els.append(Paragraph(f"<b>Exam Center : {center_name.upper()}</b>", ParagraphStyle('ecn', fontName='Arial-Bold', fontSize=10, alignment=TA_LEFT)))
    els.append(Spacer(1, 10))

    rooms_data = []
    if center_id:
        rooms = (PA_Exam_Center_Trn.query
                 .filter_by(fk_exam_center_id=center_id)
                 .order_by(PA_Exam_Center_Trn.order_by)
                 .all())
        try:
            allotted_rows = db.session.execute(text("""
                SELECT fk_roomId, COUNT(*) as allotted_count, MIN(RollNo) as min_roll, MAX(RollNo) as max_roll 
                FROM PA_SeatAllotment_Details 
                WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
                GROUP BY fk_roomId
            """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
            allotted_map = {row[0]: (row[1], row[2], row[3]) for row in allotted_rows}
        except Exception:
            allotted_map = {}

        for i, r in enumerate(rooms, 1):
            allotted, min_roll, max_roll = allotted_map.get(r.id, (0, '', ''))
            rooms_data.append({
                "sno": i,
                "room_no": r.room_no or str(i),
                "capacity": r.room_capacity,
                "allotted": allotted,
                "start_roll": min_roll or '',
                "end_roll": max_roll or ''
            })

    hdr = [['S.No', 'Room No.', 'Capacity', 'Allotted', 'Starting Roll No', 'Ending Roll No']]
    rows = []
    for r in rooms_data:
        rows.append([str(r['sno']), str(r['room_no']), str(r['capacity']), str(r['allotted']),
                     str(r['start_roll']), str(r['end_roll'])])
    if not rows:
        rows = [['', 'No rooms configured.', '', '', '', '']]

    col_w = [1.5*cm, 3.5*cm, 3*cm, 3*cm, 4*cm, 4*cm]
    t = Table(hdr + rows, colWidths=col_w, repeatRows=1)
    ts = _grid_style()
    ts.add('ALIGN', (0,1), (-1,-1), 'CENTER')
    t.setStyle(ts)
    els.append(t)

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="Exam_Center_Detail_Report_{center_id}.pdf"'
    return resp

def _pdf_seating_arrangement_report(session_name, et_name, center_name, session_id=None, et_id=None, center_id=None, room_filter_id=None, room_no_filter=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.25*inch, bottomMargin=0.25*inch)
    st = _styles()
    els = []

    try:
        sql = """
            SELECT 
                r.regno, 
                r.s_name, 
                r.f_name, 
                r.mobileno,
                s.Specialization, 
                c.Name AS center_name, 
                rm.RoomNo, 
                rm.RoomCapacity,
                rm.NoRow,
                rm.NoColumn,
                rm.Paper_Dist,
                d.SeatNo, 
                d.RollNo,
                deg.description AS degree_desc,
                et.Description AS et_desc
            FROM PA_SeatAllotment_Details d
            INNER JOIN PA_Registration_Mst r ON d.fk_regid = r.pk_regid
            INNER JOIN PA_Exam_Center_Mst c ON d.fk_examCenterId = c.pk_examCenterId
            INNER JOIN PA_Exam_Center_Trn rm ON d.fk_roomId = rm.pk_id
            INNER JOIN PA_ET_Master et ON d.Fk_ETID = et.Pk_ETID
            LEFT JOIN PA_Specialization_mst s ON r.fk_SId = s.Pk_SID
            LEFT JOIN ACD_Degree_Mst deg ON r.fk_dtypeid = deg.pk_degreeid
            WHERE d.Fk_ETID = :etid AND d.fk_examCenterId = :cid AND d.fk_sessionid = :sid
        """
        params = {'etid': et_id, 'cid': center_id, 'sid': session_id}

        if room_filter_id:
            sql += " AND d.fk_roomId = :roomid"
            params['roomid'] = room_filter_id

        sql += " ORDER BY rm.OrderBy, d.SeatNo"

        rows = db.session.execute(text(sql), params).fetchall()

        rooms = {}
        room_keys = []
        for row in rows:
            k = row.RoomNo
            if k not in rooms:
                rooms[k] = {
                    'capacity': getattr(row, 'RoomCapacity', 0),
                    'no_row': getattr(row, 'NoRow', ''),
                    'no_column': getattr(row, 'NoColumn', ''),
                    'paper_dist': getattr(row, 'Paper_Dist', ''),
                    'cands': []
                }
                room_keys.append(k)
            rooms[k]['cands'].append({
                "reg_no": row.regno,
                "name": row.s_name,
                "father_name": row.f_name,
                "mobile": row.mobileno,
                "seat": row.SeatNo,
                "roll_no": row.RollNo,
                "degree_type": getattr(row, 'degree_desc', '')
            })
    except Exception as e:
        print(f"Error fetching candidates: {e}")
        rooms = {}
        room_keys = []

    first = True
    
    card_r_st = ParagraphStyle('cr', fontName='Helvetica-Bold', fontSize=14, alignment=TA_CENTER)
    card_s_st = ParagraphStyle('cs', fontName='Helvetica-Bold', fontSize=9, textColor=colors.red, alignment=TA_CENTER)
    card_n_st = ParagraphStyle('cn', fontName='Helvetica', fontSize=7.5, alignment=TA_CENTER, leading=9)
    card_un_st = ParagraphStyle('cu', fontName='Helvetica', fontSize=7.5, textColor=colors.gray, alignment=TA_CENTER)
    
    if not rooms:
        _univ_header(els, st)
        els.append(Paragraph("Seating Arrangement Report", st['title']))
        els.append(Spacer(1, 10))
        els.append(Paragraph("No records found for selected criteria.", st['normal']))
    else:
        for rNo in room_keys:
            if not first:
                els.append(PageBreak())
            first = False
            
            roomData = rooms[rNo]
            cands = roomData['cands']
            capacity = roomData['capacity'] or len(cands)
            cols = roomData['no_column'] or 5
            rows_count = roomData['no_row'] or 0
            
            try:
                cols = int(cols)
                if cols <= 0: cols = 5
            except:
                cols = 5
                
            try:
                import math
                rows_count = int(rows_count)
                if rows_count <= 0: rows_count = math.ceil(capacity / cols)
            except:
                import math
                rows_count = math.ceil(capacity / cols)
                
            dist = roomData.get('paper_dist', '')
            is_sequential = dist and 'sequential' in dist.lower()

            els.append(Paragraph(f"{center_name}", ParagraphStyle('ch', fontName='Helvetica-Bold', fontSize=14, alignment=TA_CENTER, backColor=colors.HexColor('#17b1d7'), textColor=colors.white, spaceBefore=10, spaceAfter=5, borderPadding=6)))
            els.append(Paragraph(f"Room No:- {rNo}", ParagraphStyle('rh', fontName='Helvetica-Bold', fontSize=12, alignment=TA_CENTER, spaceAfter=15)))
            
            seatMap = {c['seat']: c for c in cands}
            
            grid_data = []
            
            available_w = A4[0] - 1.0*inch
            card_w = available_w / cols

            total_seats = rows_count * cols
            seatNo = 0

            for j in range(rows_count):
                current_row = []
                
                if is_sequential:
                    for k in range(cols):
                        if k == 0:
                            seatNo = j + 1
                        else:
                            seatNo = (j + 1) + (k * rows_count)
                            
                        seatIdx = total_seats - (seatNo - 1)
                            
                        if seatIdx <= capacity and capacity > 0:
                            c = seatMap.get(seatIdx)
                            if c:
                                card_content = [
                                    [Paragraph(f"{c['roll_no']}", card_r_st)],
                                    [Paragraph(f"Seat: {seatIdx}", card_s_st)],
                                    [Paragraph(f"Reg: {c['reg_no']}", card_n_st)],
                                    [Paragraph(f"{c['name'][:25]}", card_n_st)]
                                ]
                                bg_color = colors.white
                            else:
                                card_content = [
                                    [Paragraph("Vacant", card_un_st)],
                                    [Paragraph(f"Seat: {seatIdx}", card_s_st)],
                                    [Paragraph("&nbsp;", card_n_st)],
                                    [Paragraph("Unallotted", card_un_st)]
                                ]
                                bg_color = colors.white
                                
                            ct = Table(card_content, colWidths=[card_w - 4])
                            ct.setStyle(TableStyle([
                                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                ('BOX', (0,0), (-1,-1), 1, colors.black),
                                ('BACKGROUND', (0,0), (-1,-1), bg_color),
                                ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                                ('TOPPADDING', (0,0), (-1,-1), 1),
                                ('LEFTPADDING', (0,0), (-1,-1), 2),
                                ('RIGHTPADDING', (0,0), (-1,-1), 2),
                            ]))
                            current_row.append(ct)
                        else:
                            current_row.append('')
                else:
                    if j % 2 == 0:
                        for k in range(cols):
                            if j == 0:
                                seatNo = k + 1
                            else:
                                seatNo = (j * cols) + (k + 1)
                                
                            seatIdx = total_seats - (seatNo - 1)
                                
                            if seatIdx <= capacity and capacity > 0:
                                c = seatMap.get(seatIdx)
                                if c:
                                    card_content = [
                                        [Paragraph(f"{c['roll_no']}", card_r_st)],
                                        [Paragraph(f"Seat: {seatIdx}", card_s_st)],
                                        [Paragraph(f"Reg: {c['reg_no']}", card_n_st)],
                                        [Paragraph(f"{c['name'][:25]}", card_n_st)]
                                    ]
                                    bg_color = colors.white
                                else:
                                    card_content = [
                                        [Paragraph("Vacant", card_un_st)],
                                        [Paragraph(f"Seat: {seatIdx}", card_s_st)],
                                        [Paragraph("&nbsp;", card_n_st)],
                                        [Paragraph("Unallotted", card_un_st)]
                                    ]
                                    bg_color = colors.white
                                    
                                ct = Table(card_content, colWidths=[card_w - 4])
                                ct.setStyle(TableStyle([
                                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                    ('BOX', (0,0), (-1,-1), 1, colors.black),
                                    ('BACKGROUND', (0,0), (-1,-1), bg_color),
                                    ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                                    ('TOPPADDING', (0,0), (-1,-1), 1),
                                    ('LEFTPADDING', (0,0), (-1,-1), 2),
                                    ('RIGHTPADDING', (0,0), (-1,-1), 2),
                                ]))
                                current_row.append(ct)
                            else:
                                current_row.append('')
                    else:
                        for k in range(cols):
                            if k == 0:
                                seatNo = (j * cols) + cols
                            else:
                                seatNo = (j * cols) + cols - k
                                
                            seatIdx = total_seats - (seatNo - 1)
                                
                            if seatIdx <= capacity and capacity > 0:
                                c = seatMap.get(seatIdx)
                                if c:
                                    card_content = [
                                        [Paragraph(f"{c['roll_no']}", card_r_st)],
                                        [Paragraph(f"Seat: {seatIdx}", card_s_st)],
                                        [Paragraph(f"Reg: {c['reg_no']}", card_n_st)],
                                        [Paragraph(f"{c['name'][:25]}", card_n_st)]
                                    ]
                                    bg_color = colors.white
                                else:
                                    card_content = [
                                        [Paragraph("Vacant", card_un_st)],
                                        [Paragraph(f"Seat: {seatIdx}", card_s_st)],
                                        [Paragraph("&nbsp;", card_n_st)],
                                        [Paragraph("Unallotted", card_un_st)]
                                    ]
                                    bg_color = colors.white
                                    
                                ct = Table(card_content, colWidths=[card_w - 4])
                                ct.setStyle(TableStyle([
                                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                    ('BOX', (0,0), (-1,-1), 1, colors.black),
                                    ('BACKGROUND', (0,0), (-1,-1), bg_color),
                                    ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                                    ('TOPPADDING', (0,0), (-1,-1), 1),
                                    ('LEFTPADDING', (0,0), (-1,-1), 2),
                                    ('RIGHTPADDING', (0,0), (-1,-1), 2),
                                ]))
                                current_row.append(ct)
                            else:
                                current_row.append('')
                    
                while len(current_row) < cols:
                    current_row.append('')
                grid_data.append(current_row)
                
            if grid_data:
                row_h = 0.9 * inch
                gt = Table(grid_data, colWidths=[card_w]*cols, rowHeights=[row_h]*len(grid_data))
                gt.setStyle(TableStyle([
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                    ('TOPPADDING', (0,0), (-1,-1), 1),
                    ('LEFTPADDING', (0,0), (-1,-1), 2),
                    ('RIGHTPADDING', (0,0), (-1,-1), 2),
                ]))
                els.append(gt)
                
            els.append(Spacer(1, 15))
            els.append(Paragraph("<b>Blackboard is on this side!</b>", ParagraphStyle('bb', fontName='Helvetica-Bold', fontSize=12, alignment=TA_CENTER, backColor=colors.black, textColor=colors.white, borderPadding=8)))

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = 'inline; filename="Seating_Arrangement_Report.pdf"'
    return resp

def _pdf_exam_center_room_wise(session_name, et_name, et_id=None, session_id=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    st = _styles()
    els = []
    _univ_header(els, st)
    els.append(Paragraph("Exam Center Room Wise Report", st['title']))
    els.append(Spacer(1, 4))
    els.append(Paragraph(f"<b>Session :</b> {session_name}   &nbsp;&nbsp;   <b>ET Type :</b> {et_name}", st['bold9']))
    els.append(Spacer(1, 6))

    sql = """
        SELECT 
            c.Name, 
            r.RoomNo, 
            r.RoomCapacity, 
            r.Paper_Dist, 
            r.RoomLocation
        FROM PA_Exam_Center_Mst c
        INNER JOIN PA_Exam_Center_Trn r ON c.pk_examCenterId = r.fk_examCenterId
        WHERE c.fk_ETID = :et_id AND c.fk_SessionId = :sid AND c.IsActive = 1
        ORDER BY c.OrderBy, r.OrderBy
    """
    
    hdr = [['S.No', 'Exam Center', 'Room No.', 'Capacity', 'Distribution Pattern', 'Location']]
    rows = []
    sno = 1
    
    if et_id and session_id:
        results = db.session.execute(text(sql), {'et_id': et_id, 'sid': session_id}).fetchall()
        for r in results:
            rows.append([
                str(sno), 
                Paragraph(r.Name or '', st['normal']), 
                Paragraph(r.RoomNo or '', st['normal']), 
                str(r.RoomCapacity),
                Paragraph(r.Paper_Dist or '', st['normal']), 
                Paragraph(r.RoomLocation or '', st['normal'])
            ])
            sno += 1
            
    if not rows:
        rows = [['', 'No data found for selected Session and ET Type.', '', '', '', '']]

    col_w = [1*cm, 5.8*cm, 1.8*cm, 1.8*cm, 4.4*cm, 3.6*cm]
    t = Table(hdr + rows, colWidths=col_w, repeatRows=1)
    ts = _grid_style()
    ts.add('ALIGN', (0,1),(0,-1), 'CENTER')
    ts.add('ALIGN', (2,1),(3,-1), 'CENTER')
    t.setStyle(ts)
    els.append(t)

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="Exam_Center_Room_Wise_Report_{et_id}.pdf"'
    return resp

def _excel_exam_center_room_wise(session_name, et_name, et_id=None, session_id=None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Room Wise Report"
    
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    alt_fill = PatternFill(start_color="EEF2FB", end_color="EEF2FB", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "Exam Center Room Wise Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Session : {session_name}      ET Type : {et_name}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = center_align
    
    headers = ['S.No', 'Exam Center', 'Room No.', 'Capacity', 'Distribution Pattern', 'Location']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    sql = """
        SELECT 
            c.Name, 
            r.RoomNo, 
            r.RoomCapacity, 
            r.Paper_Dist, 
            r.RoomLocation
        FROM PA_Exam_Center_Mst c
        INNER JOIN PA_Exam_Center_Trn r ON c.pk_examCenterId = r.fk_examCenterId
        WHERE c.fk_ETID = :et_id AND c.fk_SessionId = :sid AND c.IsActive = 1
        ORDER BY c.OrderBy, r.OrderBy
    """
    
    row_idx = 5
    sno = 1
    has_data = False
    
    if et_id and session_id:
        results = db.session.execute(text(sql), {'et_id': et_id, 'sid': session_id}).fetchall()
        for r in results:
            has_data = True
            fill = alt_fill if sno % 2 == 0 else PatternFill(fill_type=None)
            
            c1 = ws.cell(row=row_idx, column=1, value=sno)
            c2 = ws.cell(row=row_idx, column=2, value=r.Name)
            c3 = ws.cell(row=row_idx, column=3, value=r.RoomNo or '')
            c4 = ws.cell(row=row_idx, column=4, value=r.RoomCapacity)
            c5 = ws.cell(row=row_idx, column=5, value=r.Paper_Dist or '')
            c6 = ws.cell(row=row_idx, column=6, value=r.RoomLocation or '')
            
            for cell in [c1, c2, c3, c4, c5, c6]:
                cell.border = thin_border
                cell.fill = fill
                cell.alignment = center_align if cell.column in [1, 3, 4] else left_align
                    
            row_idx += 1
            sno += 1
            
    if not has_data:
        ws.merge_cells('A5:F5')
        cell = ws.cell(row=5, column=1, value="No data found for selected Session and ET Type.")
        cell.alignment = center_align
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="Exam_Center_Room_Wise_Report_{et_id}.xlsx"'
    return resp

def _pdf_admit_card(reg_nos, session_id):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.0*cm, rightMargin=1.0*cm,
                            topMargin=0.8*cm, bottomMargin=0.8*cm)
    st = _styles()
    st['l'] = ParagraphStyle('l', alignment=TA_LEFT, fontSize=8, leading=10)
    st['c'] = ParagraphStyle('c', alignment=TA_CENTER, fontSize=8, leading=10)
    st['cb'] = ParagraphStyle('cb', alignment=TA_CENTER, fontSize=8, fontName='Helvetica-Bold', leading=10)
    st['inst'] = ParagraphStyle('inst', alignment=TA_JUSTIFY, fontSize=7.5, leading=9)

    els = []

    candidates = []
    if reg_nos:
        placeholders = ','.join(['?' for _ in reg_nos])
        try:
            conn = db.session.connection().connection
            cursor = conn.cursor()
            
            sql_positional = f"""
                SELECT 
                    r.regno, 
                    r.s_name, 
                    r.f_name,
                    r.m_name,
                    r.mobileno,
                    s.Specialization, 
                    c.Name AS center_name, 
                    c.Address AS center_address,
                    rm.RoomNo, 
                    d.SeatNo, 
                    d.RollNo,
                    deg.description AS degree_desc,
                    et.Description AS et_desc,
                    r.c_address,
                    r.c_district,
                    r.C_Village,
                    r.c_pincode,
                    doc.imgattach_p,
                    doc.imgattach_s
                FROM PA_Registration_Mst r
                LEFT JOIN PA_SeatAllotment_Details d ON d.fk_regid = r.pk_regid AND d.fk_sessionid = ?
                LEFT JOIN PA_Exam_Center_Mst c ON d.fk_examCenterId = c.pk_examCenterId
                LEFT JOIN PA_Exam_Center_Trn rm ON d.fk_roomId = rm.pk_id
                LEFT JOIN PA_ET_Master et ON d.Fk_ETID = et.Pk_ETID
                LEFT JOIN PA_Specialization_mst s ON r.fk_SId = s.Pk_SID
                LEFT JOIN ACD_Degree_Mst deg ON r.fk_dtypeid = deg.pk_degreeid
                LEFT JOIN PA_Registration_Document doc ON r.pk_regid = doc.fk_regid
                WHERE r.regno IN ({placeholders})
            """
            
            params = (session_id,) + tuple(reg_nos)
            rows = cursor.execute(sql_positional, params).fetchall()

            for row in rows:
                photo_img = ''
                sign_img = ''
                
                if row.imgattach_p:
                    try:
                        p_stream = io.BytesIO(row.imgattach_p)
                        from PIL import Image as PILImage
                        PILImage.open(p_stream).verify()
                        p_stream.seek(0)
                        photo_img = Image(p_stream, width=3.5*cm, height=4.5*cm)
                    except Exception as e:
                        print(f"Error parsing photo for {row.regno}: {e}")
                        
                if row.imgattach_s:
                    try:
                        s_stream = io.BytesIO(row.imgattach_s)
                        from PIL import Image as PILImage
                        PILImage.open(s_stream).verify()
                        s_stream.seek(0)
                        sign_img = Image(s_stream, width=3.5*cm, height=1.2*cm)
                    except Exception as e:
                        print(f"Error parsing signature for {row.regno}: {e}")
                
                c_name = row.center_name if row.center_name else 'N/A'
                candidates.append({
                    "reg_no": row.regno,
                    "name": row.s_name,
                    "father_name": row.f_name,
                    "mother_name": row.m_name,
                    "subject": row.Specialization or 'N/A',
                    "center": f"{c_name} (CCS HAU, HISAR)",
                    "address": f"HouseNo-{row.c_address or ''}, District-{row.c_district or ''}, Village-{row.C_Village or ''}, Pin-{row.c_pincode or ''}",
                    "degree": row.degree_desc or 'N/A',
                    "exam_type": row.et_desc or 'N/A',
                    "roll_no": row.RollNo or 'N/A',
                    "photo": photo_img,
                    "sign": sign_img
                })
        except Exception as e:
            print(f"Error fetching real candidates for Admit Card: {e}")

    if not candidates:
        candidates = [{
            "reg_no": "116351110", "name": "AARTI", "father_name": "JEET SINGH", "mother_name": "PUSHPA DEVI",
            "subject": "Math", "center": "COLLEGE OF COMMUNITY SCIENCE (COCS) (CCS HAU, HISAR)",
            "address": "HouseNo-0,StreetNo-DHANI KHAIRAMPUR ROAD,AddressLandMark-VPO SADALPUR TEHSIL ADAMPUR DISTT HISAR,DistrictName-HISAR,Block/Town-ADAMPUR BL,Ward/Village-Sadelpur",
            "degree": "M. Sc. College of Basic Sciences & Humanities (2025-2026)", "exam_type": "ET-I (2025-26)", "roll_no": "87001",
            "photo": "", "sign": ""
        }]

    affix_text = Paragraph("<b>Self Attested photograph of<br/>the Candidate (Photo to be<br/>same as the one Uploaded<br/>in the Application Form)</b>", st['cb'])
    
    logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
    logo_img = Image(logo_path, width=1.6*cm, height=1.6*cm) if os.path.exists(logo_path) else 'LOGO'
    
    asst_reg_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'asst_reg_sign.png')
    asst_reg_img = Image(asst_reg_path, width=3.0*cm, height=0.8*cm) if os.path.exists(asst_reg_path) else Spacer(1, 0.8*cm)

    for idx, c in enumerate(candidates):
        header_text = f"<b>CCS Haryana Agricultural University,Hisar</b><br/>Provisional Admit Card(Subject to eligibility) for Entrance Test for admission to<br/>{c.get('degree','M. Sc. College of Basic Sciences & Humanities (2025-2026)')}<br/><b>Note: Kindly bring the Admit Card for Entrance Test. Mobile Phone not allowed</b>"

        head_t = Table([
            [logo_img, Paragraph(header_text, st['c']), Paragraph("<b>University's<br/>Copy</b>", st['cb'])]
        ], colWidths=[2.0*cm, 14.5*cm, 2.5*cm])
        head_t.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1)
        ]))
        els.append(head_t)
        
        photo_display = c.get('photo') if c.get('photo') else Spacer(1, 4.5*cm)
        sign_display = c.get('sign') if c.get('sign') else Spacer(1, 1.2*cm)
        
        raw_address = str(c.get('address','')).replace(',', ', ')
        addr_style = ParagraphStyle('addr', alignment=TA_LEFT, fontSize=7.5, leading=9)
        
        data = [
            [Paragraph('<b>Application Id</b>', st['l']), str(c.get('reg_no','')), photo_display, affix_text],
            [Paragraph('<b>Roll No.</b>', st['l']), Paragraph(f"<b>{c.get('roll_no','')}</b>", st['l']), '', ''],
            [Paragraph('<b>Date of Exam</b>', st['l']), '28/06/2025(Saturday)', '', ''],
            [Paragraph('<b>Reporting Time</b>', st['l']), '09:00 AM', '', ''],
            [Paragraph('<b>Time of Exam</b>', st['l']), '10:00 AM  To  12:30 PM', '', ''],
            [Paragraph('<b>Applicant\'s Name</b>', st['l']), str(c.get('name','')).upper(), '', ''],      
            [Paragraph('<b>Father\'s Name</b>', st['l']), str(c.get('father_name','')).upper(), '', ''],  
            [Paragraph('<b>Mother\'s Name</b>', st['l']), str(c.get('mother_name','')).upper(), '', ''],  
            [Paragraph('<b>Subject for Entrance Test</b>', st['l']), str(c.get('subject','')), Paragraph("<b>Photo of the Candidate<br/>Generated Online</b>", st['cb']), ''],

            [Paragraph(f"<b>Entrance Exam Centre</b><br/>{c.get('center','')}", addr_style), '', [Spacer(1, 10), sign_display, Spacer(1, 5), Paragraph("<b>Signature of the Candidate<br/>Generated Online</b>", st['cb'])], [Spacer(1, 35), Paragraph("<b>Signature of the Candidate<br/>(to be signed during test)</b>", st['cb'])]],
            [Paragraph(f"<b>Address of the Candidate</b><br/>{raw_address}", addr_style), '', '', ''],

            [[Paragraph("This is certified that the particulars <br/> given in this admit card i.e. name, photograph, address <br/>  and signature are correct to the best of my knowledge.", st['c']), Spacer(1, 35), Paragraph("<b>Signature of the Candidate</b>", st['cb'])],
             [Spacer(1, 15), asst_reg_img, Spacer(1, 20), Paragraph("<b>Assistant Registrar<br/>(Academic)</b>", st['cb'])],
             [Spacer(1, 25), Paragraph("<b>Signature of Invigilator - I</b>", st['cb'])],
             Paragraph("<b>Thumb Impression of the<br/>Candidate (to be obtained<br/>during test)</b>", st['cb'])], 

            ['', '', [Spacer(1, 25), Paragraph("<b>Signature of Invigilator - II</b>", st['cb'])], '']
        ]

        t = Table(data, colWidths=[5.2*cm, 5.5*cm, 4.3*cm, 4.0*cm])
        t.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),  
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (2,0), (2,-1), 'CENTER'),
            ('SPAN', (2,0), (2,7)), 
            ('SPAN', (3,0), (3,8)), 
            ('SPAN', (0,9), (1,9)), 
            ('SPAN', (0,10), (1,10)), 
            ('SPAN', (2,9), (2,10)), 
            ('SPAN', (3,9), (3,10)), 
            ('SPAN', (0,11), (0,12)), 
            ('SPAN', (1,11), (1,12)), 
            ('SPAN', (3,11), (3,12)), 
            ('VALIGN', (0,11), (0,12), 'TOP'), 
            ('VALIGN', (1,11), (1,12), 'BOTTOM'), 
            ('VALIGN', (2,11), (2,11), 'BOTTOM'), 
            ('VALIGN', (2,12), (2,12), 'BOTTOM'), 
            ('VALIGN', (3,11), (3,12), 'BOTTOM'), 
            ('VALIGN', (2,0), (3,8), 'BOTTOM'),
            ('VALIGN', (2,9), (3,10), 'BOTTOM'), 
        ]))
        els.append(t)
        els.append(Spacer(1, 15))
        els.append(Paragraph("<b>------------------------------------------------------- Cut from here -------------------------------------------------------</b>", st['c']))
        els.append(Spacer(1, 15))

        raw_inst = "1. Please verify the details on Admit card after downloading and intimate discrepancies, if any.<br/>2. Candidate should report at the examination centre 30 minutes before the commencement of the examination to avoid last minute confusion .<b>First 30 minutes will be provided for pre exam formalities.</b><br/>3. No candidate shall be allowed in the centre after the start of exam.<br/>4. The candidate will hand over OMR Sheet, University Admit Card and Question Booklet (for PG exam only) to the Invigilators before leaving the examination hall.<br/>5. Kindly bring your original ID proof (Aadhar card, Voter Id, PAN card, passport, driving license etc.) for entry into the exam centre.<br/>6. Following items are not allowed to carry inside the examination Center/examination room under any circumstances; all ornaments / jewellery etc any metallic item personal belongings like bag etc, wristwatch, camera, calculator, mobile, phone, pager, bluetooth, earphones, purse, logtables, healthband, electronic gadgets with or without built-in-calculators, geometry/plastic pouch, blank or printed papers, written chits etc. <b>Candidates found possessing any of these items will be treated AS AN ACT OF UNFAIR MEANS AND WILL BE DISQUALIFIED FOR HIS/HER RIGHT TO ADMISSION.</b>"

        head_c = Table([
            [logo_img, Paragraph(header_text, st['c']), Paragraph("<b>Candidate's<br/>Copy</b>", st['cb'])]
        ], colWidths=[2.0*cm, 14.5*cm, 2.5*cm])
        
        head_c.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1)
        ]))
        els.append(head_c)

        full_address = str(c.get('address','')).replace(',', ', ')

        data_c = [
            [Paragraph('<b>Application Id</b>', st['l']), str(c.get('reg_no','')), Paragraph('<b>Time of Examination</b>', st['l']), '10:00 AM  To  12:30 PM'],
            [Paragraph('<b>Roll No.</b>', st['l']), Paragraph(f"<b>{c.get('roll_no','')}</b>", st['l']), Paragraph('<b>Applicant\'s Name</b>', st['l']), str(c.get('name','')).upper()],
            [Paragraph('<b>Date of Exam</b>', st['l']), '28/06/2025(Saturday)', Paragraph('<b>Father\'s Name</b>', st['l']), str(c.get('father_name','')).upper()],
            [Paragraph('<b>Reporting Time</b>', st['l']), '09:00 AM', Paragraph('<b>Mother\'s Name</b>', st['l']), str(c.get('mother_name','')).upper()],
            [Paragraph('<b>Subject for Entrance Test</b>', st['l']), str(c.get('subject','')), '', ''],
            [Paragraph('<b>Entrance Exam Centre</b>', st['l']), Paragraph(str(c.get('center','')), addr_style), '', ''],
            [Paragraph('<b>Address of the Candidate</b>', st['l']), Paragraph(full_address, addr_style), '', ''],
            [Paragraph('<b>Instructions</b>', ParagraphStyle('inst_h', alignment=TA_CENTER, fontSize=9, fontName='Helvetica-Bold', backColor=colors.lightgrey)), '', '', ''],     
            [Paragraph(raw_inst, st['inst']), '', '', ''] 
        ]

        tc = Table(data_c, colWidths=[4.2*cm, 5.3*cm, 4*cm, 5.5*cm])
        tc.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),  
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('SPAN', (1,4), (3,4)), 
            ('SPAN', (1,5), (3,5)), 
            ('SPAN', (1,6), (3,6)), 
            ('SPAN', (0,7), (3,7)), 
            ('SPAN', (0,8), (3,8)),   
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))
        els.append(tc)

        if idx < len(candidates) - 1:
            els.append(PageBreak())

    doc.build(els)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = 'inline; filename="Admit_Cards.pdf"'
    return resp

# ─────────────────────────────────────────────────────────────
#  ROUTES FROM NEW CODE
# ─────────────────────────────────────────────────────────────
@seat_allocation_bp.route('/exam-center-report', methods=['GET', 'POST'])
def exam_center_report():
    sessions = AcademicSession.query.filter_by(is_active=True).order_by(AcademicSession.id.desc()).all()
    
    session_id = request.values.get('session_id', type=int)
    et_id = request.values.get('et_id', type=int)
    action = request.values.get('action')
    
    ets = []
    if session_id:
        ets = PA_ET_Master.query.filter_by(fk_session_id=session_id).order_by(PA_ET_Master.id).all()

    centers_data = []
    selected_sess = AcademicSession.query.get(session_id) if session_id else None
    selected_et   = PA_ET_Master.query.get(et_id) if et_id else None
    
    show_grid = False
    
    if action == 'view' and session_id and et_id:
        show_grid = True
        sql = """
            SELECT 
                c.pk_examCenterId, 
                c.Name, 
                c.Code,
                (SELECT COUNT(*) FROM PA_Exam_Center_Trn t WHERE t.fk_examCenterId = c.pk_examCenterId) as room_count,
                (SELECT COUNT(*) FROM PA_SeatAllotment_Details d WHERE d.fk_examCenterId = c.pk_examCenterId AND d.Fk_ETID = :et_id AND d.fk_sessionid = :sid) as cand_count
            FROM PA_Exam_Center_Mst c
            WHERE c.fk_ETID = :et_id AND c.fk_SessionId = :sid AND c.IsActive = 1
            ORDER BY c.OrderBy
        """
        results = db.session.execute(text(sql), {'et_id': et_id, 'sid': session_id}).fetchall()
        for r in results:
            centers_data.append({
                'id': r.pk_examCenterId,
                'name': r.Name,
                'code': r.Code,
                'room_count': r.room_count,
                'cand_count': r.cand_count
            })

    if request.method == 'POST':
        if not session_id or not et_id:
            flash("Please select both Academic Session and ET Type.", "error")
        else:
            if action == 'generate_pdf':
                return _pdf_exam_center_report(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    et_id, session_id)
            if action == 'room_wise_pdf':
                return _pdf_exam_center_room_wise(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    et_id, session_id)
            if action == 'export_excel':
                return _excel_exam_center_report(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    et_id, session_id)
            if action == 'room_wise_excel':
                return _excel_exam_center_room_wise(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    et_id, session_id)
                
    return render_template('seat_allocation/exam_center_report.html',
                           sessions=sessions, ets=ets, centers=centers_data,
                           session_id=session_id, et_id=et_id,
                           selected_sess=selected_sess, selected_et=selected_et,
                           show_grid=show_grid)

@seat_allocation_bp.route('/exam-center-detail-report', methods=['GET', 'POST'])
def exam_center_detail_report():
    sessions = AcademicSession.query.filter_by(is_active=True).order_by(AcademicSession.id.desc()).all()
    
    session_id = request.values.get('session_id', type=int)
    et_id      = request.values.get('et_id', type=int)
    center_id  = request.values.get('exam_center_id', type=int)
    action     = request.values.get('action')

    ets = []
    if session_id:
        ets = PA_ET_Master.query.filter_by(fk_session_id=session_id).order_by(PA_ET_Master.id).all()
        
    centers = []
    if et_id:
        centers = PA_Exam_Center_Mst.query.filter_by(fk_et_id=et_id, is_active=True).order_by(PA_Exam_Center_Mst.order_by).all()
        
    selected_sess   = AcademicSession.query.get(session_id) if session_id else None
    selected_et     = PA_ET_Master.query.get(et_id)         if et_id     else None
    selected_center = PA_Exam_Center_Mst.query.get(center_id) if center_id else None

    et_date_str = 'N/A'
    if selected_et and selected_et.dated:
        et_date_str = selected_et.dated.strftime('%d-%m-%Y')

    show_grid = False
    rooms_data = []
    
    if action == 'view' and session_id and et_id and center_id:
        show_grid = True
        rooms = (PA_Exam_Center_Trn.query
                 .filter_by(fk_exam_center_id=center_id)
                 .order_by(PA_Exam_Center_Trn.order_by)
                 .all())
        try:
            allotted_rows = db.session.execute(text("""
                SELECT fk_roomId, COUNT(*) as allotted_count, MIN(RollNo) as min_roll, MAX(RollNo) as max_roll 
                FROM PA_SeatAllotment_Details 
                WHERE fk_examCenterId = :cid AND Fk_ETID = :etid AND fk_sessionid = :sid
                GROUP BY fk_roomId
            """), {'cid': center_id, 'etid': et_id, 'sid': session_id}).fetchall()
            allotted_map = {row[0]: (row[1], row[2], row[3]) for row in allotted_rows}
        except Exception:
            allotted_map = {}

        for i, r in enumerate(rooms, 1):
            allotted, min_roll, max_roll = allotted_map.get(r.id, (0, '', ''))
            rooms_data.append({
                "sno": i,
                "room_no": r.room_no or str(i),
                "capacity": r.room_capacity,
                "allotted": allotted,
                "start_roll": min_roll or '',
                "end_roll": max_roll or ''
            })

    if request.method == 'POST':
        if not session_id or not et_id or not center_id:
            flash("Please select Academic Session, ET Type, and Exam Center.", "error")
        else:
            purpose = request.form.get('purpose', 'E')
            
            if purpose == 'E':
                return _pdf_exam_center_detail_report(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    selected_center.name       if selected_center else 'N/A',
                    session_id, et_id, center_id, et_date_str)
            elif purpose == 'D':
                return _pdf_exam_center_display_report(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    selected_center.name       if selected_center else 'N/A',
                    session_id, et_id, center_id, et_date_str)
            elif purpose == 'R':
                return _pdf_exam_center_door_placards(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    selected_center.name       if selected_center else 'N/A',
                    session_id, et_id, center_id, et_date_str)
            elif purpose == 'P':
                return _pdf_exam_center_distribution_pattern(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    selected_center.name       if selected_center else 'N/A',
                    session_id, et_id, center_id, et_date_str)
            elif purpose == 'EX':
                return _excel_exam_center_distribution_pattern(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    selected_center.name       if selected_center else 'N/A',
                    session_id, et_id, center_id, et_date_str)
            elif purpose == 'XL':
                return _excel_exam_center_detail_report(
                    selected_sess.session_name if selected_sess else 'N/A',
                    selected_et.description    if selected_et   else 'N/A',
                    selected_center.name       if selected_center else 'N/A',
                    session_id, et_id, center_id)

    return render_template('seat_allocation/exam_center_detail_report.html',
                           sessions=sessions, ets=ets, centers=centers,
                           session_id=session_id, et_id=et_id, center_id=center_id, purpose=request.form.get('purpose', 'E'),
                           selected_sess=selected_sess, selected_et=selected_et, selected_center=selected_center,
                           show_grid=show_grid, rooms=rooms_data)

@seat_allocation_bp.route('/seating-arrangement-report', methods=['GET', 'POST'])
def seating_arrangement_report():
    sessions = AcademicSession.query.filter_by(is_active=True).order_by(AcademicSession.id.desc()).all()
    ets      = PA_ET_Master.query.order_by(PA_ET_Master.id).all()
    centers  = PA_Exam_Center_Mst.query.filter_by(is_active=True).order_by(PA_Exam_Center_Mst.order_by).all()
    if request.method == 'POST':
        action    = request.form.get('action')
        if action == 'generate_pdf':
            sid       = request.form.get('session_id', type=int)
            et_id     = request.form.get('et_id', type=int)
            center_id = request.form.get('exam_center_id', type=int)
            room_id   = request.form.get('room_id', type=int)
            sess      = AcademicSession.query.get(sid)          if sid       else None
            et        = PA_ET_Master.query.get(et_id)            if et_id     else None
            center    = PA_Exam_Center_Mst.query.get(center_id)  if center_id else None
            room      = PA_Exam_Center_Trn.query.get(room_id)    if room_id   else None
            return _pdf_seating_arrangement_report(
                sess.session_name if sess   else 'N/A',
                et.description    if et     else 'N/A',
                center.name       if center else 'N/A',
                sid, et_id, center_id, room_id,
                room.room_no      if room   else None)
    return render_template('seat_allocation/seating_arrangement_report.html',
                           sessions=sessions, ets=ets, centers=centers)

@seat_allocation_bp.route('/roll-no-admit-card-generation', methods=['GET', 'POST'])
def roll_no_admit_card_generation():
    sessions = AcademicSession.query.filter_by(is_active=True).order_by(AcademicSession.id.desc()).all()
    degrees  = Degree.query.filter_by(active=True).order_by(Degree.name).all()
    ets      = PA_ET_Master.query.order_by(PA_ET_Master.id).all()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'generate_admit_cards':
            reg_nos = request.form.getlist('p_cand_sel')
            session_id = request.form.get('session_id', type=int)
            return _pdf_admit_card(reg_nos, session_id)
            
        elif action == 'process_roll_no':
            session_id = request.form.get('session_id', type=int)
            et_id      = request.form.get('et_id', type=int)
            reg_nos    = request.form.getlist('np_cand_sel')
            start_roll_str = request.form.get('roll_start_with')
            
            if not session_id or not et_id:
                flash('Please select Session and ET Type.', 'error')
            elif not reg_nos:
                flash('Please select candidates to process.', 'error')
            elif not start_roll_str or not start_roll_str.isdigit():
                flash('Please provide a valid starting Roll No.', 'error')
            else:
                try:
                    start_roll = int(start_roll_str)
                    
                    placeholders = ','.join([f":reg{i}" for i in range(len(reg_nos))])
                    params = {'etid': et_id, 'sid': session_id}
                    for i, r in enumerate(reg_nos):
                        params[f"reg{i}"] = r
                        
                    sql = f"""
                        SELECT d.fk_regid, r.regno
                        FROM PA_SeatAllotment_Details d
                        INNER JOIN PA_Registration_Mst r ON d.fk_regid = r.pk_regid
                        INNER JOIN PA_Exam_Center_Mst c ON d.fk_examCenterId = c.pk_examCenterId
                        INNER JOIN PA_Exam_Center_Trn rm ON d.fk_roomId = rm.pk_id
                        WHERE d.Fk_ETID = :etid AND d.fk_sessionid = :sid 
                          AND r.regno IN ({placeholders})
                        ORDER BY c.OrderBy, rm.RoomNo, d.SeatNo
                    """
                    rows = db.session.execute(text(sql), params).fetchall()
                    
                    curr_roll = start_roll
                    for row in rows:
                        fk_regid = row.fk_regid
                        db.session.execute(text("""
                            UPDATE PA_Registration_Mst 
                            SET rollno = :rno 
                            WHERE pk_regid = :regid
                        """), {'rno': str(curr_roll), 'regid': fk_regid})
                        
                        db.session.execute(text("""
                            UPDATE PA_SeatAllotment_Details 
                            SET RollNo = :rno 
                            WHERE fk_regid = :regid AND Fk_ETID = :etid AND fk_sessionid = :sid
                        """), {'rno': str(curr_roll), 'regid': fk_regid, 'etid': et_id, 'sid': session_id})
                        
                        curr_roll += 1
                        
                    db.session.commit()
                    flash(f'Successfully processed {len(rows)} candidates. Last Roll No: {curr_roll - 1}', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error during processing: {str(e)}', 'error')

        elif action == 'unprocess_roll_no':
            session_id = request.form.get('session_id', type=int)
            et_id      = request.form.get('p_et_id', type=int)
            
            if not session_id or not et_id:
                flash('Please select Session and ET Type in the Processed section to unprocess.', 'error')
            else:
                try:
                    db.session.execute(text("""
                        UPDATE PA_Registration_Mst 
                        SET rollno = NULL 
                        WHERE pk_regid IN (
                            SELECT fk_regid FROM PA_SeatAllotment_Details 
                            WHERE Fk_ETID = :etid AND fk_sessionid = :sid
                        )
                    """), {'etid': et_id, 'sid': session_id})
                    
                    db.session.execute(text("""
                        UPDATE PA_SeatAllotment_Details 
                        SET RollNo = NULL 
                        WHERE Fk_ETID = :etid AND fk_sessionid = :sid
                    """), {'etid': et_id, 'sid': session_id})
                    
                    db.session.commit()
                    flash('Unprocessed successfully: Roll numbers cleared.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error during unprocess: {str(e)}', 'error')
        else:
            flash('Roll numbers processed successfully.', 'success')
            
    return render_template('seat_allocation/roll_no_admit_card_generation.html',
                           sessions=sessions, degrees=degrees, ets=ets)

@seat_allocation_bp.route('/print-admit-card/<reg_no>/<int:session_id>', methods=['GET'])
def print_admit_card(reg_no, session_id):
    return _pdf_admit_card([reg_no], session_id)

@seat_allocation_bp.route('/seating-arrangement', methods=['GET', 'POST'])
def seating_arrangement():
    sessions = AcademicSession.query.filter_by(is_active=True).order_by(AcademicSession.id.desc()).all()
    ets      = PA_ET_Master.query.order_by(PA_ET_Master.id).all()
    centers  = PA_Exam_Center_Mst.query.filter_by(is_active=True).order_by(PA_Exam_Center_Mst.order_by).all()
    if request.method == 'POST':
        flash('Seating arrangement generated successfully.', 'success')
    return render_template('seat_allocation/seating_arrangement.html', sessions=sessions, ets=ets, centers=centers)

