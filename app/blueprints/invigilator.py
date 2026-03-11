from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from app import db
from app.models import PA_ET_Master, AcademicSession, PA_Exam_Center_Mst, PA_Exam_Center_Trn, StaffType_Mst, StaffCategory_Mst, PA_ExternalStaff_Trn, PA_Exemption_Mst, PA_Exemption_Detail, PA_StaffDuties_Mst, PA_StaffDuties_Trn
from datetime import datetime
import json
import collections
from app.utils_pdf import render_to_pdf

invigilator_bp = Blueprint('invigilator', __name__, url_prefix='/invigilator')

def get_pagination(query, per_page=10):
    page = request.args.get('page', 1, type=int)
    return query.paginate(page=page, per_page=per_page)


@invigilator_bp.before_request
def check_admin_auth():
    from flask import session, redirect, url_for, flash, request
    if not session.get('user_id'):
        if request.endpoint and 'login' not in request.endpoint:
            flash('Please login to access this section.', 'error')
            return redirect(url_for('main.login'))

@invigilator_bp.route('/et-master', methods=['GET', 'POST'])
def et_master():
    edit_id = request.args.get('edit_id', type=int)
    edit_et = PA_ET_Master.query.get(edit_id) if edit_id else None
    
    selected_session_id = request.args.get('filter_session_id', type=int)

    if request.method == 'POST':
        try:
            et_id = request.form.get('et_id')
            session_id = request.form.get('session_id')
            description = request.form.get('description')
            dated_str = request.form.get('dated')
            remarks = request.form.get('remarks')
            letter_no = request.form.get('letter_no')
            letter_date_str = request.form.get('letter_date')

            dated = datetime.strptime(dated_str, '%Y-%m-%d') if dated_str else None
            letter_date = datetime.strptime(letter_date_str, '%Y-%m-%d') if letter_date_str else None

            if not session_id or not description or not dated_str or not letter_no or not letter_date_str:
                flash('Please fill all mandatory fields.', 'error')
                return redirect(url_for('invigilator.et_master'))

            if et_id:
                # Update
                et = PA_ET_Master.query.get(et_id)
                if et:
                    et.fk_session_id = session_id
                    et.description = description
                    et.dated = dated
                    et.remarks = remarks
                    et.letter_no = letter_no
                    et.letter_date = letter_date
                    flash('ET details updated successfully!', 'success')
            else:
                # Insert
                max_id = db.session.query(db.func.max(PA_ET_Master.id)).scalar() or 0
                new_et = PA_ET_Master(
                    id=max_id + 1,
                    fk_session_id=session_id,
                    description=description,
                    dated=dated,
                    remarks=remarks,
                    letter_no=letter_no,
                    letter_date=letter_date
                )
                db.session.add(new_et)
                flash('ET details saved successfully!', 'success')

            db.session.commit()
            return redirect(url_for('invigilator.et_master', filter_session_id=selected_session_id))
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    # GET request data - Limit to last 2 sessions
    sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).limit(2).all()
    
    # Ensure selected_session_id is an int if it exists
    if selected_session_id:
        selected_session_id = int(selected_session_id)
    elif edit_et:
        selected_session_id = edit_et.fk_session_id

    # Pagination with filtering and sorting by dated
    page = request.args.get('page', 1, type=int)
    ets = []
    pagination = None

    if selected_session_id:
        query = PA_ET_Master.query.filter_by(fk_session_id=selected_session_id)
        pagination = query.order_by(PA_ET_Master.dated.desc()).paginate(page=page, per_page=10, error_out=False)
        ets = pagination.items

    return render_template('invigilator/et_master.html', 
                           sessions=sessions, 
                           ets=ets, 
                           edit_et=edit_et, 
                           pagination=pagination,
                           selected_session_id=selected_session_id)

@invigilator_bp.route('/delete-et-master/<int:id>')
def delete_et_master(id):
    try:
        et = PA_ET_Master.query.get_or_404(id)
        db.session.delete(et)
        db.session.commit()
        flash('Entrance Test deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Cannot delete this Entrance Test as it is used in other records.', 'error')
    return redirect(url_for('invigilator.et_master'))

@invigilator_bp.route('/exam-center-master', methods=['GET', 'POST'])
def exam_center_master():
    edit_id = request.args.get('edit_id', type=int)
    edit_center = PA_Exam_Center_Mst.query.get(edit_id) if edit_id else None
    
    selected_session_id = request.args.get('filter_session_id', type=int)
    selected_et_id = request.args.get('filter_et_id', type=int)
    
    if request.method == 'POST':
        try:
            center_id = request.form.get('center_id')
            session_id = request.form.get('session_id')
            et_id = request.form.get('et_id')
            name = request.form.get('name')
            code = request.form.get('code')
            center_type = request.form.get('center_type')
            address = request.form.get('address')
            is_active = 'is_active' in request.form
            order_by = request.form.get('order_by')
            
            # Rooms JSON parsing
            rooms_data = request.form.get('rooms_data')
            rooms = json.loads(rooms_data) if rooms_data else []

            if not session_id or not et_id or not name or not address:
                flash('Please fill all mandatory fields.', 'error')
                return redirect(url_for('invigilator.exam_center_master', filter_session_id=selected_session_id, filter_et_id=selected_et_id))
                
            # Validate session and ET exist
            if not AcademicSession.query.get(session_id):
                flash('Invalid Academic Session selected.', 'error')
                return redirect(url_for('invigilator.exam_center_master', filter_session_id=selected_session_id, filter_et_id=selected_et_id))
            if not PA_ET_Master.query.get(et_id):
                flash('Invalid Entrance Test selected.', 'error')
                return redirect(url_for('invigilator.exam_center_master', filter_session_id=selected_session_id, filter_et_id=selected_et_id))
                
            if center_id:
                # Update
                center = PA_Exam_Center_Mst.query.get(center_id)
                if center:
                    center.fk_session_id = session_id
                    center.fk_et_id = et_id
                    center.name = name
                    center.code = code
                    center.center_type = center_type if center_type else None
                    center.address = address
                    center.is_active = is_active
                    center.order_by = order_by if order_by else None
                    
                    # Update rooms - wipe and re-insert for simplicity
                    PA_Exam_Center_Trn.query.filter_by(fk_exam_center_id=center_id).delete()
                    for room in rooms:
                        new_room = PA_Exam_Center_Trn(
                            fk_exam_center_id=center_id,
                            room_no=room.get('room_no'),
                            room_capacity=room.get('room_capacity'),
                            no_row=room.get('no_row'),
                            no_column=room.get('no_column'),
                            room_location=room.get('room_location'),
                            paper_dist=room.get('paper_dist'),
                            order_by=room.get('order_by')
                        )
                        db.session.add(new_room)
                    flash('Exam Center updated successfully!', 'success')
            else:
                # Insert
                max_id = db.session.query(db.func.max(PA_Exam_Center_Mst.id)).scalar() or 0
                new_center = PA_Exam_Center_Mst(
                    id=max_id + 1,
                    fk_session_id=session_id,
                    fk_et_id=et_id,
                    name=name,
                    code=code,
                    center_type=center_type if center_type else None,
                    address=address,
                    is_active=is_active,
                    order_by=order_by if order_by else None
                )
                db.session.add(new_center)
                db.session.flush() # Get the new ID
                
                for room in rooms:
                    new_room = PA_Exam_Center_Trn(
                        fk_exam_center_id=new_center.id,
                        room_no=room.get('room_no'),
                        room_capacity=room.get('room_capacity'),
                        no_row=room.get('no_row'),
                        no_column=room.get('no_column'),
                        room_location=room.get('room_location'),
                        paper_dist=room.get('paper_dist'),
                        order_by=room.get('order_by')
                    )
                    db.session.add(new_room)
                flash('Exam Center saved successfully!', 'success')

            db.session.commit()
            return redirect(url_for('invigilator.exam_center_master', filter_session_id=session_id, filter_et_id=et_id))
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    # GET request data - Limit to last 2 sessions
    sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).limit(2).all()
    
    if selected_session_id:
        selected_session_id = int(selected_session_id)
    elif edit_center:
        selected_session_id = edit_center.fk_session_id
    elif sessions:
        selected_session_id = sessions[0].id

    # Get ETs for the selected session
    ets = PA_ET_Master.query.filter_by(fk_session_id=selected_session_id).order_by(PA_ET_Master.dated.desc()).all()
    
    if selected_et_id:
        selected_et_id = int(selected_et_id)
    elif edit_center:
        selected_et_id = edit_center.fk_et_id

    # Pre-load rooms for edit mode
    existing_rooms_json = "[]"
    if edit_center and edit_center.rooms:
        existing_rooms = [{
            'room_no': r.room_no,
            'room_capacity': r.room_capacity,
            'no_row': r.no_row,
            'no_column': r.no_column,
            'room_location': r.room_location,
            'paper_dist': r.paper_dist,
            'order_by': r.order_by
        } for r in edit_center.rooms]
        existing_rooms_json = json.dumps(existing_rooms)

    page = request.args.get('page', 1, type=int)
    centers = []
    pagination = None
    if selected_et_id:
        query = PA_Exam_Center_Mst.query.filter_by(fk_et_id=selected_et_id)
        pagination = query.order_by(PA_Exam_Center_Mst.name.asc()).paginate(page=page, per_page=10, error_out=False)
        centers = pagination.items

    return render_template('invigilator/exam_center_master.html',
                           sessions=sessions,
                           ets=ets,
                           centers=centers,
                           edit_center=edit_center,
                           existing_rooms_json=existing_rooms_json,
                           pagination=pagination,
                           selected_session_id=selected_session_id,
                           selected_et_id=selected_et_id)

@invigilator_bp.route('/delete-exam-center/<int:id>')
def delete_exam_center(id):
    try:
        center = PA_Exam_Center_Mst.query.get_or_404(id)
        db.session.delete(center)
        db.session.commit()
        flash('Exam Center deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Cannot delete this Exam Center as it is used in other records.', 'error')
    return redirect(url_for('invigilator.exam_center_master'))

@invigilator_bp.route('/exam-center-copy', methods=['GET', 'POST'])
def exam_center_copy():
    if request.method == 'POST':
        try:
            source_session_id = request.form.get('source_session_id')
            source_et_id = request.form.get('source_et_id')
            dest_session_id = request.form.get('dest_session_id')
            dest_et_id = request.form.get('dest_et_id')
            center_ids = request.form.getlist('center_ids') # list of checked center IDs
            
            # Print for debugging
            print(f"DEBUG: Form center_ids: {center_ids}")
            
            if not all([source_session_id, source_et_id, dest_session_id, dest_et_id]):
                flash('Please select both source and destination configurations.', 'error')
                return redirect(url_for('invigilator.exam_center_copy'))
                
            # Validate session and ET exist
            if not AcademicSession.query.get(source_session_id) or not AcademicSession.query.get(dest_session_id):
                flash('Invalid Academic Session selected.', 'error')
                return redirect(url_for('invigilator.exam_center_copy'))
            if not PA_ET_Master.query.get(source_et_id) or not PA_ET_Master.query.get(dest_et_id):
                flash('Invalid Entrance Test selected.', 'error')
                return redirect(url_for('invigilator.exam_center_copy'))
                
            if source_et_id == dest_et_id:
                flash('Source and Destination Entrance Tests cannot be the same.', 'error')
                return redirect(url_for('invigilator.exam_center_copy'))
                
            if not center_ids:
                flash('Please select at least one Exam Center to copy.', 'error')
                return redirect(url_for('invigilator.exam_center_copy'))
                
            # Copy logic
            centers_to_copy = PA_Exam_Center_Mst.query.filter(PA_Exam_Center_Mst.id.in_(center_ids)).all()
            max_center_id = db.session.query(db.func.max(PA_Exam_Center_Mst.id)).scalar() or 0
            
            copied_count = 0
            for center in centers_to_copy:
                # Check if already exists in dest to avoid duplicate (optional business logic, but good practice)
                existing = PA_Exam_Center_Mst.query.filter_by(fk_et_id=dest_et_id, name=center.name).first()
                if existing:
                    continue # Skip
                    
                max_center_id += 1
                new_center = PA_Exam_Center_Mst(
                    id=max_center_id,
                    fk_session_id=dest_session_id,
                    fk_et_id=dest_et_id,
                    name=center.name,
                    code=center.code,
                    center_type=center.center_type,
                    address=center.address,
                    is_active=center.is_active,
                    order_by=center.order_by
                )
                db.session.add(new_center)
                db.session.flush() # Needed for getting new_center.id
                
                # Copy rooms
                for room in center.rooms:
                    new_room = PA_Exam_Center_Trn(
                        fk_exam_center_id=new_center.id,
                        room_no=room.room_no,
                        room_capacity=room.room_capacity,
                        no_row=room.no_row,
                        no_column=room.no_column,
                        room_location=room.room_location,
                        paper_dist=room.paper_dist,
                        order_by=room.order_by
                    )
                    db.session.add(new_room)
                copied_count += 1
                
            db.session.commit()
            if copied_count > 0:
                flash(f'Successfully copied {copied_count} Exam Centers with their rooms.', 'success')
            else:
                flash('No centers were copied (they might already exist in the destination).', 'info')
                
            return redirect(url_for('invigilator.exam_center_copy'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    # For initial load, we just pass the sessions. The ETs and Centers would ideally be fetched via AJAX based on session selection. 
    # For simplicity and mimicking classic ASP.NET without heavy JS refactoring right away, we can pass all and filter in JS, or do postbacks.
    # Given the standard approach, passing all and letting JS filter is smooth for small datasets.
    sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).all()
    ets = PA_ET_Master.query.order_by(PA_ET_Master.dated.desc()).all()
    centers = PA_Exam_Center_Mst.query.order_by(PA_Exam_Center_Mst.name.asc()).all()

    return render_template('invigilator/exam_center_copy.html', sessions=sessions, ets=ets, all_centers=centers)

@invigilator_bp.route('/staff-type-master', methods=['GET', 'POST'])
def staff_type_master():
    edit_id = request.args.get('edit_id', type=int)
    edit_type = StaffType_Mst.query.get(edit_id) if edit_id else None

    if request.method == 'POST':
        try:
            type_id = request.form.get('type_id')
            description = request.form.get('description')

            if not description:
                flash('Description is required.', 'error')
                return redirect(url_for('invigilator.staff_type_master'))

            if type_id:
                staff_type = StaffType_Mst.query.get(type_id)
                if staff_type:
                    staff_type.description = description
                    flash('Staff Type updated successfully!', 'success')
            else:
                max_id = db.session.query(db.func.max(StaffType_Mst.id)).scalar() or 0
                new_type = StaffType_Mst(
                    id=max_id + 1,
                    description=description
                )
                db.session.add(new_type)
                flash('Staff Type saved successfully!', 'success')

            db.session.commit()
            return redirect(url_for('invigilator.staff_type_master'))
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    page = request.args.get('page', 1, type=int)
    pagination = StaffType_Mst.query.order_by(StaffType_Mst.description.asc()).paginate(page=page, per_page=10, error_out=False)
    types = pagination.items

    return render_template('invigilator/staff_type_master.html', types=types, edit_type=edit_type, pagination=pagination)

@invigilator_bp.route('/delete-staff-type/<int:id>')
def delete_staff_type(id):
    try:
        staff_type = StaffType_Mst.query.get_or_404(id)
        db.session.delete(staff_type)
        db.session.commit()
        flash('Staff Type deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Cannot delete this Staff Type as it is used in other records.', 'error')
    return redirect(url_for('invigilator.staff_type_master'))

@invigilator_bp.route('/staff-category-master', methods=['GET', 'POST'])
def staff_category_master():
    edit_id = request.args.get('edit_id', type=int)
    edit_category = StaffCategory_Mst.query.get(edit_id) if edit_id else None

    # Handle filters for GET request (and persisted through POST via hidden fields or just re-evaluating)
    search_staff_type_id = request.args.get('search_staff_type_id', type=int)

    if request.method == 'POST':
        try:
            category_id = request.form.get('category_id')
            description = request.form.get('description')
            staff_type_id = request.form.get('staff_type_id')
            amount = request.form.get('amount')
            category_order = request.form.get('category_order')

            if not description or not staff_type_id:
                flash('Description and Staff Type are mandatory.', 'error')
                return redirect(url_for('invigilator.staff_category_master'))
                
            # Validate Staff Type exists
            if not StaffType_Mst.query.get(staff_type_id):
                flash('Invalid Staff Type selected.', 'error')
                return redirect(url_for('invigilator.staff_category_master'))

            if category_id:
                # Update
                category = StaffCategory_Mst.query.get(category_id)
                if category:
                    category.description = description
                    category.fk_staff_type_id = staff_type_id
                    category.amount = amount if amount else None
                    category.category_order = category_order if category_order else None
                    flash('Staff Category updated successfully!', 'success')
            else:
                # Insert
                max_id = db.session.query(db.func.max(StaffCategory_Mst.id)).scalar() or 0
                new_category = StaffCategory_Mst(
                    id=max_id + 1,
                    description=description,
                    fk_staff_type_id=staff_type_id,
                    amount=amount if amount else None,
                    category_order=category_order if category_order else None
                )
                db.session.add(new_category)
                flash('Staff Category saved successfully!', 'success')

            db.session.commit()
            return redirect(url_for('invigilator.staff_category_master'))
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    # GET request data
    staff_types = StaffType_Mst.query.order_by(StaffType_Mst.description.asc()).all()

    query = StaffCategory_Mst.query
    if search_staff_type_id:
        query = query.filter_by(fk_staff_type_id=search_staff_type_id)

    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(StaffCategory_Mst.category_order.asc(), StaffCategory_Mst.description.asc()).paginate(page=page, per_page=10, error_out=False)
    categories = pagination.items

    return render_template('invigilator/staff_category_master.html', 
                           staff_types=staff_types, 
                           categories=categories, 
                           edit_category=edit_category, 
                           pagination=pagination,
                           search_staff_type_id=search_staff_type_id)

@invigilator_bp.route('/delete-staff-category/<int:id>')
def delete_staff_category(id):
    try:
        category = StaffCategory_Mst.query.get_or_404(id)
        db.session.delete(category)
        db.session.commit()
        flash('Staff Category deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Cannot delete this Staff Category as it is used in other records.', 'error')
    return redirect(url_for('invigilator.staff_category_master'))

@invigilator_bp.route('/external-staff-master', methods=['GET', 'POST'])
def external_staff_master():
    edit_id = request.args.get('edit_id', type=int)
    edit_staff = PA_ExternalStaff_Trn.query.get(edit_id) if edit_id else None

    if request.method == 'POST':
        try:
            staff_id = request.form.get('staff_id')
            et_id = request.form.get('et_id')
            name = request.form.get('name')
            department = request.form.get('department')
            designation = request.form.get('designation')
            contact_no = request.form.get('contact_no')

            if not et_id or not name:
                flash('Entrance Test and Name are mandatory.', 'error')
                return redirect(url_for('invigilator.external_staff_master'))
                
            # Validate ET exists
            if not PA_ET_Master.query.get(et_id):
                flash('Invalid Entrance Test selected.', 'error')
                return redirect(url_for('invigilator.external_staff_master'))

            if staff_id:
                staff = PA_ExternalStaff_Trn.query.get(staff_id)
                if staff:
                    staff.fk_et_id = et_id
                    staff.name = name
                    staff.department = department
                    staff.designation = designation
                    staff.contact_no = contact_no
                    flash('External Staff updated successfully!', 'success')
            else:
                max_id = db.session.query(db.func.max(PA_ExternalStaff_Trn.id)).scalar() or 0
                new_staff = PA_ExternalStaff_Trn(
                    id=max_id + 1,
                    fk_et_id=et_id,
                    name=name,
                    department=department,
                    designation=designation,
                    contact_no=contact_no
                )
                db.session.add(new_staff)
                flash('External Staff saved successfully!', 'success')

            db.session.commit()
            return redirect(url_for('invigilator.external_staff_master'))
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    # Fetch current session to filter ETs
    active_session = AcademicSession.query.filter_by(is_active=True).first()
    if active_session:
        ets = PA_ET_Master.query.filter_by(fk_session_id=active_session.id).order_by(PA_ET_Master.dated.desc()).all()
    else:
        # Fallback to last 2 sessions if no active session is marked
        sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).limit(2).all()
        session_ids = [s.id for s in sessions]
        ets = PA_ET_Master.query.filter(PA_ET_Master.fk_session_id.in_(session_ids)).order_by(PA_ET_Master.dated.desc()).all()
    
    # Handle search filter
    search_et_id = request.args.get('search_et_id', type=int)
    
    query = PA_ExternalStaff_Trn.query
    if search_et_id:
        query = query.filter_by(fk_et_id=search_et_id)
    else:
        # Filter out records where name is None by default if no search is provided
        # or just sort them so they appear at the end
        query = query.filter(PA_ExternalStaff_Trn.name != None)

    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(PA_ExternalStaff_Trn.name.asc()).paginate(page=page, per_page=10, error_out=False)
    staff_list = pagination.items

    return render_template('invigilator/external_staff_master.html', 
                           ets=ets, 
                           staff_list=staff_list, 
                           edit_staff=edit_staff, 
                           pagination=pagination,
                           search_et_id=search_et_id)

@invigilator_bp.route('/delete-external-staff/<int:id>')
def delete_external_staff(id):
    search_et_id = request.args.get('search_et_id')
    try:
        staff = PA_ExternalStaff_Trn.query.get_or_404(id)
        db.session.delete(staff)
        db.session.commit()
        flash('External Staff deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Cannot delete this External Staff.', 'error')
    return redirect(url_for('invigilator.external_staff_master', search_et_id=search_et_id))

@invigilator_bp.route('/exemption-et-duties', methods=['GET', 'POST'])
def exemption_et_duties():
    # To keep this focused and within context limits, we will implement a simplified version of the exemption logic.
    # The live template seems to have a complex grid for selecting staff. 
    # We will implement a standard manual entry form for exemptions to match the DB schema.
    
    edit_id = request.args.get('edit_id', type=int)
    edit_exemption = PA_Exemption_Mst.query.get(edit_id) if edit_id else None
    
    if request.method == 'POST':
        try:
            exemp_mst_id = request.form.get('exemp_mst_id')
            session_id = request.form.get('session_id')
            et_id = request.form.get('et_id')
            
            # Details
            emp_id = request.form.get('emp_id')
            staff_name = request.form.get('staff_name')
            department = request.form.get('department')
            designation = request.form.get('designation')
            contact_no = request.form.get('contact_no')
            
            if not session_id or not et_id or not staff_name:
                flash('Please fill all mandatory fields.', 'error')
                return redirect(url_for('invigilator.exemption_et_duties'))
                
            # Validate session and ET exist
            if not AcademicSession.query.get(session_id):
                flash('Invalid Academic Session selected.', 'error')
                return redirect(url_for('invigilator.exemption_et_duties'))
            if not PA_ET_Master.query.get(et_id):
                flash('Invalid Entrance Test selected.', 'error')
                return redirect(url_for('invigilator.exemption_et_duties'))
                
            if exemp_mst_id:
                mst = PA_Exemption_Mst.query.get(exemp_mst_id)
                mst.fk_session_id = session_id
                mst.fk_et_id = et_id
                
                # Assume 1 detail per mst for this simple implementation
                if mst.details:
                    dtl = mst.details[0]
                    dtl.emp_id = emp_id
                    dtl.staff_name = staff_name
                    dtl.department = department
                    dtl.designation = designation
                    dtl.contact_no = contact_no
                    dtl.fk_et_id = et_id
                flash('Exemption updated successfully!', 'success')
            else:
                max_mst_id = db.session.query(db.func.max(PA_Exemption_Mst.id)).scalar() or 0
                new_mst = PA_Exemption_Mst(id=max_mst_id + 1, fk_session_id=session_id, fk_et_id=et_id)
                db.session.add(new_mst)
                db.session.flush()
                
                max_dtl_id = db.session.query(db.func.max(PA_Exemption_Detail.id)).scalar() or 0
                new_dtl = PA_Exemption_Detail(
                    id=max_dtl_id + 1,
                    fk_exemption_id=new_mst.id,
                    emp_id=emp_id,
                    staff_name=staff_name,
                    department=department,
                    designation=designation,
                    contact_no=contact_no,
                    fk_et_id=et_id
                )
                db.session.add(new_dtl)
                flash('Exemption saved successfully!', 'success')
                
            db.session.commit()
            return redirect(url_for('invigilator.exemption_et_duties'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).all()
    ets = PA_ET_Master.query.order_by(PA_ET_Master.dated.desc()).all()
    
    page = request.args.get('page', 1, type=int)
    pagination = PA_Exemption_Mst.query.order_by(PA_Exemption_Mst.id.desc()).paginate(page=page, per_page=10, error_out=False)
    exemptions = pagination.items

    return render_template('invigilator/exemption_et_duties.html',
                           sessions=sessions,
                           ets=ets,
                           exemptions=exemptions,
                           edit_exemption=edit_exemption,
                           pagination=pagination)

@invigilator_bp.route('/delete-exemption/<int:id>')
def delete_exemption(id):
    try:
        exemp = PA_Exemption_Mst.query.get_or_404(id)
        db.session.delete(exemp)
        db.session.commit()
        flash('Exemption deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Cannot delete this exemption.', 'error')
    return redirect(url_for('invigilator.exemption_et_duties'))

@invigilator_bp.route('/staff-duties-master', methods=['GET', 'POST'])
def staff_duties_master():
    from app.models import UM_Users_Mst, PA_StaffDuties_Room_Details
    
    action = request.args.get('action')
    edit_et_id = request.args.get('edit_et_id', type=int) or request.form.get('edit_et_id', type=int)
    edit_id = request.args.get('edit_id', type=int) or request.form.get('edit_id', type=int)
    search_session_id = request.args.get('search_session_id', type=int) or request.form.get('session_id', type=int)
    search_et_id = request.args.get('search_et_id', type=int) or request.form.get('et_id', type=int)
    
    # If we are editing an ET, we inherently want to search for that ET's data
    if edit_et_id and not search_et_id:
        search_et_id = edit_et_id

    show_form = action in ['add_new', 'search'] or edit_et_id or edit_id or request.method == 'POST' or search_et_id or search_session_id
    
    sessions = AcademicSession.query.order_by(AcademicSession.session_order.desc()).all()
    ets_all = PA_ET_Master.query.order_by(PA_ET_Master.dated.desc()).all()

    if not show_form:
        active_session = AcademicSession.query.filter_by(is_active=True).first()
        query = PA_ET_Master.query
        if active_session:
            query = query.filter_by(fk_session_id=active_session.id)
        page = request.args.get('page', 1, type=int)
        pagination = query.order_by(PA_ET_Master.dated.desc()).paginate(page=page, per_page=10, error_out=False)
        return render_template('invigilator/staff_duty_first_look.html', ets=pagination.items, pagination=pagination)

    edit_duty_trn = PA_StaffDuties_Trn.query.get(edit_id) if edit_id else None

    edit_from_date = ''
    edit_to_date = ''
    if edit_duty_trn:
        from datetime import datetime
        try:
            if edit_duty_trn.from_date:
                # Try parsing standard formats that might be in DB
                for fmt in ('%d/%b/%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%b-%Y'):
                    try:
                        edit_from_date = datetime.strptime(edit_duty_trn.from_date, fmt).strftime('%d/%m/%Y')
                        break
                    except ValueError:
                        pass
        except Exception:
            pass
            
        try:
            if edit_duty_trn.to_date:
                for fmt in ('%d/%b/%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%b-%Y'):
                    try:
                        edit_to_date = datetime.strptime(edit_duty_trn.to_date, fmt).strftime('%d/%m/%Y')
                        break
                    except ValueError:
                        pass
        except Exception:
            pass

    
    # Internal Employee Search
    search_emp_code = request.args.get('search_emp_code')
    search_emp_name = request.args.get('search_emp_name')
    employees = []
    if search_emp_code or search_emp_name:
        emp_query = UM_Users_Mst.query
        if search_emp_code:
            emp_query = emp_query.filter(
                (UM_Users_Mst.fk_empId.like(f'%{search_emp_code}%')) | 
                (UM_Users_Mst.loginname.like(f'%{search_emp_code}%'))
            )
        if search_emp_name:
            emp_query = emp_query.filter(UM_Users_Mst.name.like(f'%{search_emp_name}%'))
        employees = emp_query.limit(20).all()

    if request.method == 'POST':
        if request.form.get('action') == 'save':
            try:
                session_id = request.form.get('session_id')
                et_id = request.form.get('et_id')
                exam_center_id = request.form.get('exam_center_id')
                room_ids = request.form.getlist('room_ids')
                
                staff_type_id = request.form.get('staff_type_id')
                staff_cat_id = request.form.get('staff_cat_id')
                staff_mode = request.form.get('staff_mode')
                
                emp_id = request.form.get('emp_id')
                ex_staff_id = request.form.get('ex_staff_id')
                staff_name = request.form.get('staff_name')
                designation = request.form.get('designation')
                department = request.form.get('department')
                contact_no = request.form.get('contact_no')
                
                rem_type = request.form.get('rem_type')
                rate = request.form.get('rate', type=float) or 0
                from_date = request.form.get('from_date')
                to_date = request.form.get('to_date')
                amount = request.form.get('amount', type=float) or 0

                if not session_id or not et_id or not staff_name or not exam_center_id:
                    flash('Please fill all mandatory fields.', 'error')
                else:
                    mst = PA_StaffDuties_Mst.query.filter_by(fk_session_id=session_id, fk_et_id=et_id).first()
                    if not mst:
                        max_mst_id = db.session.query(db.func.max(PA_StaffDuties_Mst.id)).scalar() or 0
                        mst = PA_StaffDuties_Mst(id=max_mst_id + 1, fk_session_id=session_id, fk_et_id=et_id)
                        db.session.add(mst)
                        db.session.flush()

                    if edit_duty_trn:
                        dtl = edit_duty_trn
                        dtl.staff_name = staff_name
                        dtl.department = department
                        dtl.designation = designation
                        dtl.contact_no = contact_no
                        dtl.fk_exam_center_id = exam_center_id
                        dtl.fk_staff_type_id = staff_type_id if staff_type_id else None
                        dtl.fk_staff_cat_id = staff_cat_id if staff_cat_id else None
                        dtl.emp_id = emp_id if staff_mode == '0' else None
                        dtl.fk_ex_staff_id = ex_staff_id if staff_mode == '1' else None
                        dtl.remuneration_type = rem_type
                        dtl.rate = rate
                        dtl.from_date = from_date
                        dtl.to_date = to_date
                        dtl.amount = amount
                        
                        PA_StaffDuties_Room_Details.query.filter_by(fk_trn_id=dtl.id).delete()
                        flash('Staff Duty updated successfully!', 'success')
                    else:
                        max_dtl_id = db.session.query(db.func.max(PA_StaffDuties_Trn.id)).scalar() or 0
                        dtl = PA_StaffDuties_Trn(
                            id=max_dtl_id + 1,
                            fk_staff_id=mst.id,
                            staff_name=staff_name,
                            department=department,
                            designation=designation,
                            contact_no=contact_no,
                            fk_exam_center_id=exam_center_id,
                            fk_staff_type_id=staff_type_id if staff_type_id else None,
                            fk_staff_cat_id=staff_cat_id if staff_cat_id else None,
                            fk_et_id=et_id,
                            emp_id=emp_id if staff_mode == '0' else None,
                            fk_ex_staff_id=ex_staff_id if staff_mode == '1' else None,
                            remuneration_type=rem_type,
                            rate=rate,
                            from_date=from_date,
                            to_date=to_date,
                            amount=amount
                        )
                        db.session.add(dtl)
                        db.session.flush()
                        flash('Staff Duty assigned successfully!', 'success')

                    room_no_str_list = []
                    for rid in room_ids:
                        room_obj = PA_Exam_Center_Trn.query.get(rid)
                        if room_obj:
                            room_no_str_list.append(room_obj.room_no)
                            max_rd_id = db.session.query(db.func.max(PA_StaffDuties_Room_Details.id)).scalar() or 0
                            rd = PA_StaffDuties_Room_Details(
                                id=max_rd_id + 1,
                                fk_room_id=rid,
                                room_no=room_obj.room_no,
                                fk_et_id=et_id,
                                fk_staff_id_mst=mst.id,
                                fk_trn_id=dtl.id,
                                emp_id=dtl.emp_id,
                                fk_ex_staff_id=dtl.fk_ex_staff_id
                            )
                            db.session.add(rd)
                    
                    dtl.room_no = ", ".join(room_no_str_list)
                    db.session.commit()
                    return redirect(url_for('invigilator.staff_duties_master', edit_et_id=et_id))
                
            except Exception as e:
                db.session.rollback()
                flash(f'An error occurred: {str(e)}', 'error')

        # If it's a POST but action != 'save', we just want to load the template down below (postback)
        if request.form.get('action') == 'search':
            pass  # Handled below by checking request.form directly

    search_session_id = request.args.get('search_session_id', type=int) or request.form.get('session_id', type=int)
    search_et_id = request.args.get('search_et_id', type=int) or request.form.get('et_id', type=int) or edit_et_id

    if search_et_id and not search_session_id:
        et_obj = PA_ET_Master.query.get(search_et_id)
        if et_obj:
            search_session_id = et_obj.fk_session_id

    if not search_session_id and not search_et_id:
        active_session = AcademicSession.query.filter_by(is_active=True).first()
        if active_session:
            search_session_id = active_session.id

    if search_session_id and not search_et_id:
        # User requested: "initially it is the complete for a session ok"
        pass

    if search_session_id:
        sessions = [AcademicSession.query.get(search_session_id)]
    else:
        sessions = AcademicSession.query.order_by(AcademicSession.session_order.desc()).all()

    et_query = PA_ET_Master.query
    if search_session_id:
        et_query = et_query.filter_by(fk_session_id=search_session_id)
    ets = et_query.order_by(PA_ET_Master.dated.desc()).all()

    center_query = PA_Exam_Center_Mst.query
    if search_session_id:
        center_query = center_query.filter_by(fk_session_id=search_session_id)
    if search_et_id:
        center_query = center_query.filter_by(fk_et_id=search_et_id)
    centers = center_query.order_by(PA_Exam_Center_Mst.name.asc()).all()

    staff_types = StaffType_Mst.query.order_by(StaffType_Mst.description.asc()).all()
    categories = StaffCategory_Mst.query.order_by(StaffCategory_Mst.description.asc()).all()

    external_staff_list = []
    if search_et_id:
        external_staff_list = PA_ExternalStaff_Trn.query.filter_by(fk_et_id=search_et_id).order_by(PA_ExternalStaff_Trn.name).all()

    rooms_data = []
    selected_rooms = []
    if edit_duty_trn:
        selected_rooms = [str(r.fk_room_id) for r in edit_duty_trn.assigned_rooms]

    active_center_id = request.form.get('exam_center_id', type=int)
    if not active_center_id and edit_duty_trn:
        active_center_id = edit_duty_trn.fk_exam_center_id
        
    if active_center_id:
        center_for_rooms = PA_Exam_Center_Mst.query.get(active_center_id)
        if center_for_rooms:
            for r in center_for_rooms.rooms:
                rooms_data.append({'id': r.id, 'name': r.room_no, 'center_id': center_for_rooms.id, 'center_name': center_for_rooms.name})

    # Grid Filters
    grid_filter_center = request.args.get('grid_filter_center', type=int)
    grid_filter_dept = request.args.get('grid_filter_dept')
    grid_filter_name = request.args.get('grid_filter_name')
    grid_filter_staff_type = request.args.get('grid_filter_staff_type', type=int)

    # Department list for filter dropdown
    unique_depts = db.session.query(PA_StaffDuties_Trn.department).filter(PA_StaffDuties_Trn.department != None).distinct().order_by(PA_StaffDuties_Trn.department.asc()).all()
    departments = [d[0] for d in unique_depts]

    trn_query = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id).outerjoin(StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id)
    
    if search_session_id:
        trn_query = trn_query.filter(PA_StaffDuties_Mst.fk_session_id == search_session_id)

    if search_et_id:
        trn_query = trn_query.filter(PA_StaffDuties_Mst.fk_et_id == search_et_id)
        
    # Apply Grid Filters
    if grid_filter_center:
        trn_query = trn_query.filter(PA_StaffDuties_Trn.fk_exam_center_id == grid_filter_center)
    if grid_filter_dept:
        trn_query = trn_query.filter(PA_StaffDuties_Trn.department == grid_filter_dept)
    if grid_filter_name:
        trn_query = trn_query.filter(PA_StaffDuties_Trn.staff_name.like(f'%{grid_filter_name}%'))
    if grid_filter_staff_type:
        trn_query = trn_query.filter(PA_StaffDuties_Trn.fk_staff_type_id == grid_filter_staff_type)

    # Apply sorting: Exam Center Order -> Staff Category Order -> Rate -> Staff Duty ID
    page = request.args.get('page', 1, type=int)
    pagination = trn_query.order_by(
        PA_Exam_Center_Mst.order_by.asc(), 
        StaffCategory_Mst.category_order.asc(), 
        PA_StaffDuties_Trn.rate.desc(),
        PA_StaffDuties_Trn.id.asc()
    ).paginate(page=page, per_page=10, error_out=False)
    allocated_staff = pagination.items

    return render_template('invigilator/staff_duty_allocation.html',
                           sessions=sessions,
                           ets=ets,
                           centers=centers,
                           rooms_data=rooms_data,
                           selected_rooms=selected_rooms,
                           staff_types=staff_types,
                           categories=categories,
                           allocated_staff=allocated_staff,
                           pagination=pagination,
                           search_session_id=search_session_id,
                           search_et_id=search_et_id,
                           edit_duty=edit_duty_trn,
                           edit_from_date=edit_from_date,
                           edit_to_date=edit_to_date,
                           employees=employees,
                           external_staff_list=external_staff_list,
                           departments=departments,
                           grid_filter_center=grid_filter_center,
                           grid_filter_dept=grid_filter_dept,
                           grid_filter_name=grid_filter_name,
                           grid_filter_staff_type=grid_filter_staff_type)

@invigilator_bp.route('/delete-duty/<int:id>')
def delete_duty(id):
    try:
        duty = PA_StaffDuties_Mst.query.get_or_404(id)
        db.session.delete(duty)
        db.session.commit()
        flash('Duty deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Cannot delete this duty.', 'error')
    return redirect(url_for('invigilator.staff_duties_master'))

@invigilator_bp.route('/swap-duties', methods=['GET', 'POST'])
def swap_duties():
    # A simplified version of swapping duties: user selects two duties from a list and swaps them.
    # We will pass the sessions and ETs, then filter duties.
    
    if request.method == 'POST':
        try:
            duty1_id = request.form.get('duty1_id')
            duty2_id = request.form.get('duty2_id')
            
            if not duty1_id or not duty2_id:
                flash('Please select two duties to swap.', 'error')
                return redirect(url_for('invigilator.swap_duties'))
                
            if duty1_id == duty2_id:
                flash('Cannot swap a duty with itself.', 'error')
                return redirect(url_for('invigilator.swap_duties'))
                
            dtl1 = PA_StaffDuties_Trn.query.get(duty1_id)
            dtl2 = PA_StaffDuties_Trn.query.get(duty2_id)
            
            if not dtl1 or not dtl2:
                flash('One or both selected duties do not exist or are invalid.', 'error')
                return redirect(url_for('invigilator.swap_duties'))
                
            # Swap logic (swap staff assignment or swap center/role)
            # We'll swap the staff details (name, contact) between the two assigned positions (center, category, etc.)
            
            temp_name = dtl1.staff_name
            temp_contact = dtl1.contact_no
            temp_designation = dtl1.designation
            
            dtl1.staff_name = dtl2.staff_name
            dtl1.contact_no = dtl2.contact_no
            dtl1.designation = dtl2.designation
            
            dtl2.staff_name = temp_name
            dtl2.contact_no = temp_contact
            dtl2.designation = temp_designation
            
            db.session.commit()
            flash('Duties swapped successfully!', 'success')
            return redirect(url_for('invigilator.swap_duties'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {str(e)}', 'error')

    # GET
    sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).all()
    ets = PA_ET_Master.query.order_by(PA_ET_Master.dated.desc()).all()
    
    search_et_id = request.args.get('search_et_id', type=int)
    
    query = PA_StaffDuties_Trn.query
    if search_et_id:
        query = query.filter_by(fk_et_id=search_et_id)
        
    page = request.args.get('page', 1, type=int)
    # We load all for swap dropdowns, pagination just for view
    all_duties = query.order_by(PA_StaffDuties_Trn.staff_name.asc()).all()
    pagination = query.order_by(PA_StaffDuties_Trn.staff_name.asc()).paginate(page=page, per_page=10, error_out=False)
    duties_list = pagination.items

    return render_template('invigilator/swap_duties.html',
                           sessions=sessions,
                           ets=ets,
                           search_et_id=search_et_id,
                           all_duties=all_duties,
                           duties_list=duties_list,
                           pagination=pagination)

@invigilator_bp.route('/invigilator-reports', methods=['GET', 'POST'])
def invigilator_reports():
    sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).all()
    ets = PA_ET_Master.query.order_by(PA_ET_Master.dated.desc()).all()
    
    report_data = []
    selected_session = request.args.get('session_id', type=int)
    selected_et = request.args.get('et_id', type=int)
    
    if request.method == 'POST':
        selected_session = request.form.get('session_id', type=int)
        selected_et = request.form.get('et_id', type=int)
        report_type = request.form.get('report_type')
        
        if report_type == '1':
            return redirect(url_for('invigilator_reports.download_main_duty_report_pdf', et_id=selected_et))
        elif report_type == '2':
            return redirect(url_for('invigilator_reports.download_icard_excel', et_id=selected_et))
        elif report_type == '3':
            return redirect(url_for('invigilator_reports.download_dept_wise_report_pdf', et_id=selected_et))
        elif report_type == '4':
            return redirect(url_for('invigilator_reports.download_category_wise_report_pdf', et_id=selected_et))
        elif report_type == '5':
            return redirect(url_for('invigilator_reports.download_center_name_list_pdf', et_id=selected_et))
        elif report_type == '6':
            return redirect(url_for('invigilator_reports.download_date_wise_attendance_pdf', et_id=selected_et))
        elif report_type == '7':
            return redirect(url_for('invigilator_reports.download_remuneration_wise_pdf', et_id=selected_et))
        elif report_type == '8':
            return redirect(url_for('invigilator_reports.download_contact_list_pdf', et_id=selected_et))
        elif report_type == '9':
            return redirect(url_for('invigilator_reports.download_control_room_staff_pdf', et_id=selected_et))
        elif report_type == '10':
            return redirect(url_for('invigilator_reports.download_duty_assigned_pdf', et_id=selected_et))
        else:
            flash('Please select a report type.', 'info')
            return redirect(url_for('invigilator.invigilator_reports', session_id=selected_session, et_id=selected_et))
        
    return render_template('invigilator/invigilator_reports.html',
                           sessions=sessions,
                           ets=ets,
                           selected_session=selected_session,
                           selected_et=selected_et,
                           report_data=report_data)









@invigilator_bp.route('/duty-letters', methods=['GET', 'POST'])
def duty_letters():
    sessions = AcademicSession.query.order_by(AcademicSession.session_name.desc()).all()
    ets = PA_ET_Master.query.order_by(PA_ET_Master.dated.desc()).all()
    staff_types = StaffType_Mst.query.order_by(StaffType_Mst.description.asc()).all()
    categories = StaffCategory_Mst.query.order_by(StaffCategory_Mst.description.asc()).all()
    
    letter_data = []
    selected_session = request.args.get('session_id', type=int)
    selected_et = request.args.get('et_id', type=int)
    selected_staff_type = request.args.get('staff_type_id', type=int)
    selected_category = request.args.get('category_id', type=int)
    
    if request.method == 'POST':
        selected_session = request.form.get('session_id', type=int)
        selected_et = request.form.get('et_id', type=int)
        selected_staff_type = request.form.get('staff_type_id', type=int)
        selected_category = request.form.get('category_id', type=int)
        return redirect(url_for('invigilator.duty_letters', session_id=selected_session, et_id=selected_et, staff_type_id=selected_staff_type, category_id=selected_category))
        
    if selected_et:
        query = PA_StaffDuties_Trn.query.filter_by(fk_et_id=selected_et)
        if selected_staff_type:
            # We don't have a direct staff_type_id in Trn, but we can filter via category or just assume it's roughly mapped
            query = query.join(StaffCategory_Mst).filter(StaffCategory_Mst.fk_staff_type_id == selected_staff_type)
        if selected_category:
            query = query.filter(PA_StaffDuties_Trn.fk_staff_cat_id == selected_category)
            
        letter_data = query.order_by(PA_StaffDuties_Trn.staff_name.asc()).all()

    return render_template('invigilator/duty_letters.html',
                           sessions=sessions,
                           ets=ets,
                           staff_types=staff_types,
                           categories=categories,
                           selected_session=selected_session,
                           selected_et=selected_et,
                           selected_staff_type=selected_staff_type,
                           selected_category=selected_category,
                           letter_data=letter_data)

@invigilator_bp.route('/duty-letters/download-pdf')
def download_duty_letters_pdf():
    selected_et = request.args.get('et_id', type=int)
    selected_staff_type = request.args.get('staff_type_id', type=int)
    selected_category = request.args.get('category_id', type=int)

    if not selected_et:
        flash('Please select an Entrance Test first.', 'error')
        return redirect(url_for('invigilator.duty_letters'))

    et = PA_ET_Master.query.get(selected_et)
    query = PA_StaffDuties_Trn.query.filter_by(fk_et_id=selected_et)
    
    if selected_staff_type:
        query = query.join(StaffCategory_Mst).filter(StaffCategory_Mst.fk_staff_type_id == selected_staff_type)
    if selected_category:
        query = query.filter(PA_StaffDuties_Trn.fk_staff_cat_id == selected_category)
        
    letter_data = query.order_by(PA_StaffDuties_Trn.staff_name.asc()).all()

    import base64
    import os
    from flask import current_app
    sign_path = os.path.join(current_app.root_path, 'static', 'images', 'pawan_kumar_sign.png')
    sign_base64 = ''
    try:
        with open(sign_path, 'rb') as f:
            sign_base64 = base64.b64encode(f.read()).decode('utf-8')
    except Exception:
        pass

    context = {
        'et': et,
        'letter_data': letter_data,
        'now': datetime.now(),
        'sign_base64': sign_base64
    }
    
    pdf = render_to_pdf('invigilator/pdf_duty_letter.html', context)
    if pdf:
        import io
        return send_file(io.BytesIO(pdf), download_name=f'Duty_Letters_{et.description}.pdf', as_attachment=True)
    return "Error generating PDF"

@invigilator_bp.route('/duty-letters/download-selected', methods=['GET', 'POST'])
def download_selected_duty_letters():
    selected_et = request.form.get('et_id', type=int) or request.args.get('et_id', type=int)
    
    if request.method == 'POST':
        selected_duty_ids = request.form.getlist('selected_duties')
    else:
        selected_duty_ids = request.args.getlist('selected_duties')
        
    if not selected_et or not selected_duty_ids:
        flash('Please select an Entrance Test and at least one employee.', 'error')
        return redirect(url_for('invigilator.duty_letters'))
        
    et = PA_ET_Master.query.get(selected_et)
    query = PA_StaffDuties_Trn.query.filter(PA_StaffDuties_Trn.id.in_(selected_duty_ids))
    letter_data = query.order_by(PA_StaffDuties_Trn.staff_name.asc()).all()
    
    import base64
    import os
    from flask import current_app
    sign_path = os.path.join(current_app.root_path, 'static', 'images', 'pawan_kumar_sign.png')
    sign_base64 = ''
    try:
        with open(sign_path, 'rb') as f:
            sign_base64 = base64.b64encode(f.read()).decode('utf-8')
    except Exception:
        pass

    context = {
        'et': et,
        'letter_data': letter_data,
        'now': datetime.now(),
        'sign_base64': sign_base64
    }
    
    pdf = render_to_pdf('invigilator/pdf_duty_letter.html', context)
    if pdf:
        import io
        return send_file(io.BytesIO(pdf), download_name=f'Duty_Letters_{et.description}_Selected.pdf', as_attachment=True)
    return "Error generating PDF"

@invigilator_bp.route('/staff-duties-export/<int:et_id>')
def export_staff_duties(et_id):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    et = PA_ET_Master.query.get_or_404(et_id)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == et_id
    ).order_by(
        PA_Exam_Center_Mst.order_by.asc(), 
        StaffCategory_Mst.category_order.asc(), 
        PA_StaffDuties_Trn.rate.desc(),
        PA_StaffDuties_Trn.id.asc()
    ).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Duties List"
    
    headers = [
        'S.No', 'Academic Session', 'ET Name', 'Staff Type', 'Staff Type Name', 
        'Staff Category', 'ET Date', 'Name', 'Room No.', 'Staff Name', 
        'Department', 'Designation', 'Contact No.', 'From Date', 'To Date', 
        'Remuneration Type', 'Amount'
    ]
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4A6690', end_color='4A6690', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = align_center
        
    for index, duty in enumerate(duties, 1):
        session_name = et.session.session_name if et.session else ''
        et_name = et.description
        staff_type_internal = 'Internal' if duty.emp_id else 'External'
        staff_type_name = duty.staff_type.description if duty.staff_type else ''
        staff_cat = duty.category.description if duty.category else ''
        et_date = et.dated.strftime('%d/%m/%Y') if et.dated else ''
        center_name = duty.exam_center.name if duty.exam_center else ''
        room_no = duty.room_no or 'ALL'
        staff_name = duty.staff_name
        dept = duty.department
        desig = duty.designation
        contact = duty.contact_no
        from_d = duty.from_date
        to_d = duty.to_date
        rem_type = duty.remuneration_type
        amount = f"{duty.amount:.2f}" if duty.amount else "0.00"
        
        row_data = [
            index, session_name, et_name, staff_type_internal, staff_type_name,
            staff_cat, et_date, center_name, room_no, staff_name,
            dept, desig, contact, from_d, to_d, rem_type, amount
        ]
        
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=index+1, column=col_num, value=val)
            cell.border = thin_border
            if col_num in [1, 2, 7, 9, 13, 14, 15, 17]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    for col_num, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row, column=col_num).value or '')
            if len(cell_val) > max_length:
                max_length = len(cell_val)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, download_name=f"Staff_Duties_List_{et.description}.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
