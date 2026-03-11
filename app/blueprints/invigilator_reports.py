from flask import Blueprint, request, send_file, flash, redirect, url_for, render_template
from app import db
from app.models import PA_ET_Master, AcademicSession, PA_StaffDuties_Trn, PA_StaffDuties_Mst, PA_Exam_Center_Mst, StaffCategory_Mst
from datetime import datetime
import io
import os
import collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Import ReportLab modules
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm, inch
from reportlab.pdfbase.pdfmetrics import registerFontFamily

invigilator_reports_bp = Blueprint('invigilator_reports', __name__, url_prefix='/invigilator-reports')

def get_logo_path():
    from flask import current_app
    return os.path.join(current_app.root_path, 'static', 'images', 'university_logo.png')

class HeaderFooterDocTemplate(SimpleDocTemplate):
    def __init__(self, filename, title_text="", sub_title_text="", right_text1="", right_text2="", left_text="", left_sub_text="", **kw):
        super().__init__(filename, **kw)
        self.title_text = title_text
        self.sub_title_text = sub_title_text
        self.right_text1 = right_text1
        self.right_text2 = right_text2
        self.left_text = left_text
        self.left_sub_text = left_sub_text
        self.logo_path = get_logo_path()

    def beforePage(self):
        # Draw header
        self.canv.saveState()
        
        # We need a 3-column header logic.
        styles = getSampleStyleSheet()
        
        # Center title (University name)
        title_style = ParagraphStyle(
            'HeaderTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold', # Replace with actual fonts if embedded, but Helvetica-Bold is standard
            fontSize=14.8,
            alignment=1, # Center
        )
        sub_title_style = ParagraphStyle(
            'HeaderSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10.5,
            alignment=1,
            spaceBefore=5
        )
        
        # Left title
        left_style = ParagraphStyle(
            'HeaderLeft',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10.5,
            alignment=0, # Left
        )
        left_sub_style = ParagraphStyle(
            'HeaderLeftSub',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10.5,
            alignment=0,
            spaceBefore=3
        )
        
        # Right text (Time/Date)
        right_style = ParagraphStyle(
            'HeaderRight',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=7.9,
            alignment=2, # Right
        )
        
        left_p = []
        if self.left_text:
            left_p.append(Paragraph(self.left_text, left_style))
        if self.left_sub_text:
            left_p.append(Paragraph(self.left_sub_text, left_sub_style))
            
        logo_p = []
        if os.path.exists(self.logo_path):
            logo_img = Image(self.logo_path, width=60, height=60)
            logo_p.append(logo_img)
            
        center_p = []
        if self.title_text:
            center_p.append(Paragraph(self.title_text, title_style))
        if self.sub_title_text:
            center_p.append(Paragraph(self.sub_title_text, sub_title_style))
            
        right_p = []
        if self.right_text1:
            right_p.append(Paragraph(self.right_text1, right_style))
        if self.right_text2:
            right_p.append(Paragraph(self.right_text2, right_style))

        # Create a 4-column table for the header to prevent overlap and keep logo left
        # Widths: Logo(2.5cm), LeftText(4.5cm), CenterText(9cm), RightText(3cm)
        header_table = Table([[logo_p, left_p, center_p, right_p]], colWidths=[2.5*cm, 4.5*cm, 9*cm, 3*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'LEFT'),
            ('ALIGN', (2,0), (2,0), 'CENTER'),
            ('ALIGN', (3,0), (3,0), 'RIGHT'),
        ]))
        
        # Draw the table on canvas
        w, h = header_table.wrap(self.width, self.topMargin)
        header_table.drawOn(self.canv, self.leftMargin, self.height + self.topMargin - h)
        
        # Draw footer (Page number)
        self.canv.setFont("Helvetica-Oblique", 7)
        self.canv.drawRightString(self.width + self.leftMargin, 0.5 * cm, f"Page {self.canv.getPageNumber()}")
        
        self.canv.restoreState()


@invigilator_reports_bp.route('/main-duty-report')
def download_main_duty_report_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        PA_Exam_Center_Mst.order_by.asc(), 
        StaffCategory_Mst.category_order.asc(), 
        PA_StaffDuties_Trn.rate.desc(),
        PA_StaffDuties_Trn.id.asc()
    ).all()
    
    grouped_duties = collections.defaultdict(lambda: collections.defaultdict(list))
    for duty in duties:
        center_name = duty.exam_center.name if duty.exam_center else 'UNASSIGNED CENTER'
        category_name = duty.category.description if duty.category else 'UNASSIGNED CATEGORY'
        grouped_duties[center_name][category_name].append(duty)

    now = datetime.now()
    left_sub = f"Duty List of {et.description if et else ''}-[{et.session.session_name if et and et.session else ''}]"
    left_sub2 = et.dated.strftime('%d-%m-%Y') if et and et.dated else ''
    
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm,
        rightMargin=1*cm,
        topMargin=4.5*cm, # Space for header
        bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        left_text=left_sub,
        left_sub_text=left_sub2,
        right_text1=f"Time : {now.strftime('%I:%M:%S %p').lower()}",
        right_text2=f"Date : {now.strftime('%m/%d/%Y')}"
    )

    elements = []
    styles = getSampleStyleSheet()
    
    center_heading_style = ParagraphStyle(
        'CenterHeading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10.5, spaceBefore=15, spaceAfter=5
    )
    
    cat_heading_style = ParagraphStyle(
        'CatHeading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10.5
    )
    
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    for center_name, cat_dict in grouped_duties.items():
        elements.append(Paragraph(center_name.upper(), center_heading_style))
        
        table_data = []
        # Header Row
        table_data.append([
            Paragraph("<b>Name Of Officer/Officials</b>", cell_style),
            Paragraph("<b>Department</b>", cell_style),
            Paragraph("<b>Duty Room</b>", cell_style),
            Paragraph("<b>S.No.</b>", cell_style),
            Paragraph("<b>ContactNo</b>", cell_style),
            Paragraph("<b>Staff Type</b>", cell_style)
        ])
        
        for cat_name, items in cat_dict.items():
            # Category Row
            table_data.append([Paragraph(f"<b>{cat_name}</b>", cat_heading_style), "", "", "", "", ""])
            
            for idx, duty in enumerate(items, 1):
                table_data.append([
                    Paragraph(duty.staff_name or '', cell_style),
                    Paragraph(duty.department or '', cell_style),
                    Paragraph(duty.room_no or 'ALL', cell_style),
                    Paragraph(str(idx), cell_style),
                    Paragraph(duty.contact_no or '', cell_style),
                    Paragraph('Internal' if duty.emp_id else 'External', cell_style)
                ])
                
        # Create Table
        t = Table(table_data, colWidths=[4.5*cm, 4.5*cm, 3.5*cm, 1.5*cm, 3*cm, 2*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (3,0), (3,-1), 'CENTER'),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ]))

        # Merge cells and style category header rows
        row_idx = 1
        for cat_name, items in cat_dict.items():
            t.setStyle(TableStyle([
                ('SPAN', (0, row_idx), (-1, row_idx)),
                ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#d0d8f0')),
                ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
            ]))
            row_idx += len(items) + 1

        elements.append(t)
        # Add Page Break if needed? User wants category-wise separation for the Category Wise report, but here it's Main Duty.
        # "for each category separate page(s) should be used" - wait, he said that generally, let's add PageBreak for Center maybe?
        # In Main Duty report, we usually don't break by category unless asked. But the user said "for each category separate page(s) should be used".
        # If he meant Category Wise report, we will do it there.
        elements.append(PageBreak())

    if not elements:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Main_Duty_Report_{et.description}.pdf", mimetype='application/pdf')


@invigilator_reports_bp.route('/category-wise-report')
def download_category_wise_report_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        StaffCategory_Mst.category_order.asc(), 
        PA_Exam_Center_Mst.order_by.asc(),
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()
    
    grouped_duties = collections.defaultdict(list)
    for duty in duties:
        category_name = duty.category.description if duty.category else 'UNASSIGNED CATEGORY'
        grouped_duties[category_name].append(duty)

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm,
        rightMargin=1*cm,
        topMargin=4.5*cm,
        bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        sub_title_text=et.session.session_name if et and et.session else '',
        left_text="Duty List Category Wise",
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    
    cat_heading_style = ParagraphStyle(
        'CatHeading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, spaceBefore=15, spaceAfter=5
    )
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    for category_name, items in grouped_duties.items():
        elements.append(Paragraph(category_name, cat_heading_style))
        
        table_data = []
        table_data.append([
            Paragraph("<b>Name Of Officer/Officials</b>", cell_style),
            Paragraph("<b>Department</b>", cell_style),
            Paragraph("<b>Exam Center Name</b>", cell_style),
            Paragraph("<b>Contact No.</b>", cell_style),
            Paragraph("<b>S.No.</b>", cell_style)
        ])
        
        for idx, duty in enumerate(items, 1):
            table_data.append([
                Paragraph(duty.staff_name or '', cell_style),
                Paragraph(duty.department or '', cell_style),
                Paragraph(duty.exam_center.name if duty.exam_center else 'UNASSIGNED', cell_style),
                Paragraph(duty.contact_no or '', cell_style),
                Paragraph(str(idx), cell_style)
            ])
            
        t = Table(table_data, colWidths=[5*cm, 4.5*cm, 5*cm, 3*cm, 1.5*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (4,0), (4,-1), 'CENTER'),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eef2fb')]),
        ]))
        
        elements.append(t)
        # Separate page for each category
        elements.append(PageBreak())

    if not elements:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Category_Wise_Report_{et.description}.pdf", mimetype='application/pdf')


@invigilator_reports_bp.route('/dept-wise-report')
def download_dept_wise_report_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        PA_StaffDuties_Trn.department.asc(),
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()
    
    grouped_duties = collections.defaultdict(list)
    for duty in duties:
        dept_name = duty.department if duty.department else 'UNASSIGNED DEPARTMENT'
        grouped_duties[dept_name].append(duty)

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm,
        rightMargin=1*cm,
        topMargin=4.5*cm,
        bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        sub_title_text=f"Entrance Test Date {et.dated.strftime('%d/%m/%Y') if et and et.dated else ''}",
        left_text=f"{et.description if et else ''} [{et.session.session_name if et and et.session else ''}]",
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    
    dept_heading_style = ParagraphStyle(
        'DeptHeading', parent=styles['Normal'], fontName='Times-Bold', fontSize=12.4, spaceBefore=15, spaceAfter=5
    )
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    for dept_name, items in grouped_duties.items():
        elements.append(Paragraph(f"Duty List for the Department/Office of {dept_name}", dept_heading_style))
        
        table_data = []
        table_data.append([
            Paragraph("<b>Name Of Officer/Officials</b>", cell_style),
            Paragraph("<b>Name Of Duty</b>", cell_style),
            Paragraph("<b>Exam Center Name</b>", cell_style),
            Paragraph("<b>S.No.</b>", cell_style),
            Paragraph("<b>Designation</b>", cell_style),
            Paragraph("<b>Signature</b>", cell_style)
        ])
        
        for idx, duty in enumerate(items, 1):
            table_data.append([
                Paragraph(duty.staff_name or '', cell_style),
                Paragraph(duty.category.description if duty.category else '', cell_style),
                Paragraph(duty.exam_center.name if duty.exam_center else 'UNASSIGNED', cell_style),
                Paragraph(str(idx), cell_style),
                Paragraph(duty.designation or '', cell_style),
                ""
            ])
            
        t = Table(table_data, colWidths=[4*cm, 3*cm, 4.5*cm, 1.5*cm, 3.5*cm, 2.5*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (3,0), (3,-1), 'CENTER'),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eef2fb')]),
        ]))
        
        elements.append(t)
        elements.append(PageBreak())

    if not elements:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Department_Wise_Report_{et.description}.pdf", mimetype='application/pdf')


@invigilator_reports_bp.route('/center-name-list')
def download_center_name_list_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    centers = PA_Exam_Center_Mst.query.filter_by(fk_et_id=selected_et).order_by(PA_Exam_Center_Mst.order_by.asc(), PA_Exam_Center_Mst.name.asc()).all()

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm,
        rightMargin=1*cm,
        topMargin=4.5*cm,
        bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        left_text="Private School / Center Name List",
        left_sub_text=et.session.session_name if et and et.session else '',
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    table_data = []
    table_data.append([
        Paragraph("<b>S.No.</b>", cell_style),
        Paragraph("<b>Name Of School</b>", cell_style),
        Paragraph("<b>Address</b>", cell_style)
    ])
    
    for idx, center in enumerate(centers, 1):
        table_data.append([
            Paragraph(str(idx), cell_style),
            Paragraph(center.name or '', cell_style),
            Paragraph(center.address or '', cell_style)
        ])
        
    t = Table(table_data, colWidths=[2*cm, 8.5*cm, 8.5*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eef2fb')]),
    ]))
    
    elements.append(t)

    if not centers:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Center_Name_List_{et.description}.pdf", mimetype='application/pdf')


# ─── I-Card Excel (type 2) ────────────────────────────────────────────────────
@invigilator_reports_bp.route('/icard-excel')
def download_icard_excel():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "I-Card"

    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(fill_type='solid', fgColor='2E5090')
    hdr_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Title rows
    ws.merge_cells('A1:H1')
    ws['A1'] = 'CCS Haryana Agricultural University'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:H2')
    et_desc = et.description if et else ''
    session_name = et.session.session_name if et and et.session else ''
    ws['A2'] = f'I-Card Details — {et_desc} [{session_name}]'
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = center_align

    ws.append([])  # blank row

    headers = ['S.No.', 'Name', 'Designation', 'Department', 'Category', 'Exam Center', 'Room No.', 'Contact No.']
    ws.append(headers)
    hdr_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.font = hdr_font_white
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = center_align

    alt_fill = PatternFill(fill_type='solid', fgColor='EEF2FB')
    for idx, duty in enumerate(duties, 1):
        row = [
            idx,
            duty.staff_name or '',
            duty.designation or '',
            duty.department or '',
            duty.category.description if duty.category else '',
            duty.exam_center.name if duty.exam_center else '',
            duty.room_no or 'ALL',
            duty.contact_no or '',
        ]
        ws.append(row)
        data_row = ws.max_row
        fill = alt_fill if idx % 2 == 0 else None
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = border
            cell.alignment = left_align
            if fill:
                cell.fill = fill

    col_widths = [6, 28, 22, 28, 22, 28, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"ICard_{et_desc}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── Date Wise Attendance (type 6) ───────────────────────────────────────────
@invigilator_reports_bp.route('/date-wise-attendance')
def download_date_wise_attendance_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        PA_StaffDuties_Trn.from_date.asc(),
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()

    # Group by from_date (or ET date if from_date is empty)
    grouped = collections.defaultdict(list)
    et_date_str = et.dated.strftime('%d-%m-%Y') if et and et.dated else 'N/A'
    for duty in duties:
        key = duty.from_date or et_date_str
        grouped[key].append(duty)

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm, rightMargin=1*cm, topMargin=4.5*cm, bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        sub_title_text=et.session.session_name if et and et.session else '',
        left_text=f"Date Wise Attendance — {et.description if et else ''}",
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    date_heading_style = ParagraphStyle('DateHeading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, spaceBefore=12, spaceAfter=4)
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    for date_key in sorted(grouped.keys()):
        items = grouped[date_key]
        elements.append(Paragraph(f"Date: {date_key}  ({len(items)} Staff)", date_heading_style))

        table_data = [[
            Paragraph("<b>S.No.</b>", cell_style),
            Paragraph("<b>Name</b>", cell_style),
            Paragraph("<b>Department</b>", cell_style),
            Paragraph("<b>Designation</b>", cell_style),
            Paragraph("<b>Category</b>", cell_style),
            Paragraph("<b>Exam Center</b>", cell_style),
            Paragraph("<b>Signature</b>", cell_style),
        ]]
        for idx, duty in enumerate(items, 1):
            table_data.append([
                Paragraph(str(idx), cell_style),
                Paragraph(duty.staff_name or '', cell_style),
                Paragraph(duty.department or '', cell_style),
                Paragraph(duty.designation or '', cell_style),
                Paragraph(duty.category.description if duty.category else '', cell_style),
                Paragraph(duty.exam_center.name if duty.exam_center else '', cell_style),
                "",
            ])

        t = Table(table_data, colWidths=[1.3*cm, 4.5*cm, 4*cm, 3.5*cm, 3*cm, 3.5*cm, 2.2*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    if not elements:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"Date_Wise_Attendance_{et.description}.pdf",
                     mimetype='application/pdf')


# ─── Remuneration Wise Report (type 7) ───────────────────────────────────────
@invigilator_reports_bp.route('/remuneration-wise')
def download_remuneration_wise_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        PA_StaffDuties_Trn.remuneration_type.asc(),
        PA_StaffDuties_Trn.rate.desc(),
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()

    grouped = collections.defaultdict(list)
    for duty in duties:
        key = duty.remuneration_type or 'UNSPECIFIED'
        grouped[key].append(duty)

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm, rightMargin=1*cm, topMargin=4.5*cm, bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        sub_title_text=et.session.session_name if et and et.session else '',
        left_text=f"Remuneration Wise Report — {et.description if et else ''}",
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    grp_style = ParagraphStyle('GrpHeading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, spaceBefore=12, spaceAfter=4)
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    for rem_type, items in grouped.items():
        total_amount = sum((float(d.amount) if d.amount else 0) for d in items)
        elements.append(Paragraph(f"Remuneration Type: {rem_type}  |  Total Staff: {len(items)}  |  Total Amount: {total_amount:,.2f}", grp_style))

        table_data = [[
            Paragraph("<b>S.No.</b>", cell_style),
            Paragraph("<b>Name</b>", cell_style),
            Paragraph("<b>Department</b>", cell_style),
            Paragraph("<b>Category</b>", cell_style),
            Paragraph("<b>Exam Center</b>", cell_style),
            Paragraph("<b>Rate</b>", cell_style),
            Paragraph("<b>Amount</b>", cell_style),
        ]]
        for idx, duty in enumerate(items, 1):
            table_data.append([
                Paragraph(str(idx), cell_style),
                Paragraph(duty.staff_name or '', cell_style),
                Paragraph(duty.department or '', cell_style),
                Paragraph(duty.category.description if duty.category else '', cell_style),
                Paragraph(duty.exam_center.name if duty.exam_center else '', cell_style),
                Paragraph(str(duty.rate or ''), cell_style),
                Paragraph(str(duty.amount or ''), cell_style),
            ])
        # Total row
        table_data.append([
            "", Paragraph("<b>TOTAL</b>", cell_style), "", "", "", "",
            Paragraph(f"<b>{total_amount:,.2f}</b>", cell_style),
        ])

        t = Table(table_data, colWidths=[1.3*cm, 4.5*cm, 4*cm, 3*cm, 3.5*cm, 2*cm, 2.7*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (5,0), (6,-1), 'RIGHT'),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#d0d8f0')),
        ]))
        elements.append(t)
        elements.append(PageBreak())

    if not elements:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"Remuneration_Wise_{et.description}.pdf",
                     mimetype='application/pdf')


# ─── Contact List (type 8) ────────────────────────────────────────────────────
@invigilator_reports_bp.route('/contact-list')
def download_contact_list_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        PA_Exam_Center_Mst.order_by.asc(),
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm, rightMargin=1*cm, topMargin=4.5*cm, bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        sub_title_text=et.session.session_name if et and et.session else '',
        left_text=f"Contact List — {et.description if et else ''}",
        left_sub_text=et.dated.strftime('%d-%m-%Y') if et and et.dated else '',
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    table_data = [[
        Paragraph("<b>S.No.</b>", cell_style),
        Paragraph("<b>Name</b>", cell_style),
        Paragraph("<b>Designation</b>", cell_style),
        Paragraph("<b>Department</b>", cell_style),
        Paragraph("<b>Category</b>", cell_style),
        Paragraph("<b>Exam Center</b>", cell_style),
        Paragraph("<b>Contact No.</b>", cell_style),
    ]]
    for idx, duty in enumerate(duties, 1):
        table_data.append([
            Paragraph(str(idx), cell_style),
            Paragraph(duty.staff_name or '', cell_style),
            Paragraph(duty.designation or '', cell_style),
            Paragraph(duty.department or '', cell_style),
            Paragraph(duty.category.description if duty.category else '', cell_style),
            Paragraph(duty.exam_center.name if duty.exam_center else '', cell_style),
            Paragraph(duty.contact_no or '', cell_style),
        ])

    t = Table(table_data, colWidths=[1.3*cm, 4*cm, 3.5*cm, 3.5*cm, 2.8*cm, 3.2*cm, 2.7*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eef2fb')]),
    ]))
    elements.append(t)

    if not duties:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"Contact_List_{et.description}.pdf",
                     mimetype='application/pdf')


# ─── Control Room Staff List (type 9) ────────────────────────────────────────
@invigilator_reports_bp.route('/control-room-staff')
def download_control_room_staff_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    # Filter staff whose category description contains "control" (case-insensitive)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et,
        StaffCategory_Mst.description.ilike('%control%')
    ).order_by(
        StaffCategory_Mst.category_order.asc(),
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()

    # Fallback: if no "control" category, get all duties for the ET
    if not duties:
        duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
            PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
        ).outerjoin(
            StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
        ).filter(
            PA_StaffDuties_Mst.fk_et_id == selected_et,
            PA_StaffDuties_Trn.room_no.ilike('%control%')
        ).order_by(PA_StaffDuties_Trn.staff_name.asc()).all()

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm, rightMargin=1*cm, topMargin=4.5*cm, bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        sub_title_text=et.session.session_name if et and et.session else '',
        left_text=f"Control Room Staff List — {et.description if et else ''}",
        left_sub_text=et.dated.strftime('%d-%m-%Y') if et and et.dated else '',
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    table_data = [[
        Paragraph("<b>S.No.</b>", cell_style),
        Paragraph("<b>Name</b>", cell_style),
        Paragraph("<b>Designation</b>", cell_style),
        Paragraph("<b>Department</b>", cell_style),
        Paragraph("<b>Category</b>", cell_style),
        Paragraph("<b>Contact No.</b>", cell_style),
        Paragraph("<b>Signature</b>", cell_style),
    ]]
    for idx, duty in enumerate(duties, 1):
        table_data.append([
            Paragraph(str(idx), cell_style),
            Paragraph(duty.staff_name or '', cell_style),
            Paragraph(duty.designation or '', cell_style),
            Paragraph(duty.department or '', cell_style),
            Paragraph(duty.category.description if duty.category else '', cell_style),
            Paragraph(duty.contact_no or '', cell_style),
            "",
        ])

    t = Table(table_data, colWidths=[1.3*cm, 4.5*cm, 3.5*cm, 3.5*cm, 3*cm, 2.7*cm, 2.5*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eef2fb')]),
    ]))
    elements.append(t)

    if not duties:
        elements.append(Paragraph("No control room staff found for this ET.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"Control_Room_Staff_{et.description}.pdf",
                     mimetype='application/pdf')


# ─── Duty Assigned Report (type 10) ──────────────────────────────────────────
@invigilator_reports_bp.route('/duty-assigned-report')
def download_duty_assigned_pdf():
    selected_et = request.args.get('et_id', type=int)
    if not selected_et:
        return "No ET selected", 400

    et = PA_ET_Master.query.get(selected_et)
    duties = PA_StaffDuties_Trn.query.join(PA_StaffDuties_Mst).outerjoin(
        PA_Exam_Center_Mst, PA_StaffDuties_Trn.fk_exam_center_id == PA_Exam_Center_Mst.id
    ).outerjoin(
        StaffCategory_Mst, PA_StaffDuties_Trn.fk_staff_cat_id == StaffCategory_Mst.id
    ).filter(
        PA_StaffDuties_Mst.fk_et_id == selected_et
    ).order_by(
        PA_Exam_Center_Mst.order_by.asc(),
        StaffCategory_Mst.category_order.asc(),
        PA_StaffDuties_Trn.staff_name.asc()
    ).all()

    now = datetime.now()
    buffer = io.BytesIO()
    doc = HeaderFooterDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=1*cm, rightMargin=1*cm, topMargin=4.5*cm, bottomMargin=1*cm,
        title_text="CCS Haryana Agricultural University",
        sub_title_text=f"{et.session.session_name if et and et.session else ''} — Duty Assigned Report",
        left_text=f"{et.description if et else ''}",
        left_sub_text=et.dated.strftime('%d-%m-%Y') if et and et.dated else '',
        right_text1=f"Date : {now.strftime('%m/%d/%Y')}",
        right_text2=f"Time : {now.strftime('%I:%M:%S %p').lower()}"
    )

    elements = []
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.9)

    table_data = [[
        Paragraph("<b>S.No.</b>", cell_style),
        Paragraph("<b>Name</b>", cell_style),
        Paragraph("<b>Department</b>", cell_style),
        Paragraph("<b>Designation</b>", cell_style),
        Paragraph("<b>Category</b>", cell_style),
        Paragraph("<b>Exam Center</b>", cell_style),
        Paragraph("<b>Room</b>", cell_style),
        Paragraph("<b>Type</b>", cell_style),
    ]]
    for idx, duty in enumerate(duties, 1):
        table_data.append([
            Paragraph(str(idx), cell_style),
            Paragraph(duty.staff_name or '', cell_style),
            Paragraph(duty.department or '', cell_style),
            Paragraph(duty.designation or '', cell_style),
            Paragraph(duty.category.description if duty.category else '', cell_style),
            Paragraph(duty.exam_center.name if duty.exam_center else '', cell_style),
            Paragraph(duty.room_no or 'ALL', cell_style),
            Paragraph('Internal' if duty.emp_id else 'External', cell_style),
        ])

    t = Table(table_data, colWidths=[1.3*cm, 4*cm, 3.5*cm, 3*cm, 2.5*cm, 3*cm, 1.8*cm, 2*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2e5090')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eef2fb')]),
    ]))
    elements.append(t)

    if not duties:
        elements.append(Paragraph("No records found.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"Duty_Assigned_{et.description}.pdf",
                     mimetype='application/pdf')
